import sys
import os
import time
import platform
import math
import threading
import hashlib
import csv
import json
import traceback
import subprocess
import datetime
import html
import glob
import re
import uuid
import logging
import tempfile
import multiprocessing
import tkinter as tk
import tkinter.font as tkfont
from tkinter import messagebox, ttk, filedialog, simpledialog

APP_NAME = "TS-PILE V1.0.12"
APP_TITLE = "PHẦN MỀM TÍNH TOÁN MÓNG CỌC"
APP_AUTHOR = "Tác Giả: Nguyễn Ngọc Dũng\nPhòng QLTK - Khối XD PPP\nTập đoàn SunGroup"
APP_BG = "#F5F7FA"
APP_NAV = "#17324D"
APP_PRIMARY = "#1F6FEB"
APP_SUCCESS = "#198754"
APP_VERSION = "V1.0.12"

# New form typography.
# Yêu cầu V1.0.12:
# - toàn bộ nội dung New form dùng cỡ cơ sở 12 pt;
# - bảng dùng 12 pt, riêng hàng có nguy cơ tràn tự giảm xuống 11 pt;
# - chiều cao hàng mục tiêu khoảng 0.70 lần định dạng trước V1.0.11.
NEW_REPORT_BODY_FONT_PT = 12.0
NEW_REPORT_TABLE_FONT_PT = 12.0
NEW_REPORT_TABLE_OVERFLOW_FONT_PT = 11.0
NEW_REPORT_TABLE_LINE_HEIGHT_PT = 10.5
NEW_REPORT_TABLE_OVERFLOW_LINE_HEIGHT_PT = 9.6
NEW_REPORT_ROW_HEIGHT_SCALE = 0.70
NEW_REPORT_AVAILABLE_WIDTH_PT = 186.0 * 72.0 / 25.4
NEW_REPORT_HTML_CELL_HPAD_PX = 2.0


def _new_report_estimated_text_width_pt(value, font_pt=NEW_REPORT_TABLE_FONT_PT, is_header=False):
    """Estimate the longest visible line width without requiring ReportLab."""
    raw = html.unescape(str(value or ""))
    raw = re.sub(r"<br\s*/?>", "\n", raw, flags=re.IGNORECASE)
    raw = re.sub(r"<[^>]+>", "", raw)
    max_width = 0.0
    for line in raw.splitlines() or [""]:
        units = 0.0
        for ch in line:
            if ch in "ilI1|":
                units += 0.28
            elif ch in ".,:;'`":
                units += 0.24
            elif ch in "-+()[]/":
                units += 0.34
            elif ch.isspace():
                units += 0.25
            elif ch in "MW@%":
                units += 0.85
            elif ch.isdigit():
                units += 0.50
            elif ch.isupper():
                units += 0.62
            else:
                units += 0.50
        width = units * float(font_pt)
        if is_header:
            width *= 1.05
        max_width = max(max_width, width)
    return max_width


def _new_report_row_needs_11pt(values, ncols, is_header=False):
    """Return True only for rows that are likely to overflow at 12 pt."""
    ncols = max(int(ncols or 1), 1)

    # Bảng 12-13 cột rất sát khổ A4; dùng 11 pt để tránh tách số Co/Ct thành hai dòng.
    if ncols >= 12:
        return True

    cell_width = NEW_REPORT_AVAILABLE_WIDTH_PT / ncols
    # Chừa biên an toàn cho sai khác bề rộng glyph giữa Times New Roman/DejaVu/Helvetica.
    usable_width = max(8.0, (cell_width - 4.0) * 0.92)
    return any(
        _new_report_estimated_text_width_pt(value, NEW_REPORT_TABLE_FONT_PT, is_header) > usable_width
        for value in values
    )


def _new_report_row_font_pt(values, ncols, is_header=False):
    return (
        NEW_REPORT_TABLE_OVERFLOW_FONT_PT
        if _new_report_row_needs_11pt(values, ncols, is_header)
        else NEW_REPORT_TABLE_FONT_PT
    )


def _new_report_column_widths_pt(headers, rows, available_width_pt=NEW_REPORT_AVAILABLE_WIDTH_PT):
    """Allocate table columns by actual content instead of equal widths."""
    ncols = max(len(headers or []), max((len(row) for row in rows), default=1), 1)
    desired = [22.0 for _ in range(ncols)]

    if headers:
        header_font = _new_report_row_font_pt(headers, ncols, True)
        for cidx, value in enumerate(headers[:ncols]):
            desired[cidx] = max(
                desired[cidx],
                _new_report_estimated_text_width_pt(value, header_font, True) * 1.10 + 4.0,
            )

    for row in rows:
        padded = list(row) + [""] * max(0, ncols - len(row))
        row_font = _new_report_row_font_pt(padded[:ncols], ncols)
        for cidx, value in enumerate(padded[:ncols]):
            desired[cidx] = max(
                desired[cidx],
                _new_report_estimated_text_width_pt(value, row_font, False) * 1.10 + 4.0,
            )

    total = sum(desired)
    if total <= 0:
        return [available_width_pt / ncols for _ in range(ncols)]

    # Scale to the exact usable page width. Wider content receives a wider column.
    scale = float(available_width_pt) / total
    return [width * scale for width in desired]


def _new_report_pdf_metrics(ncols, font_pt):
    """Return compact leading and vertical padding for a PDF table row."""
    ncols = max(int(ncols or 1), 1)
    font_pt = float(font_pt)
    legacy_leading = 7.2 if ncols >= 10 else (8.2 if ncols >= 7 else 12.3)
    target_height = (legacy_leading + 6.0) * NEW_REPORT_ROW_HEIGHT_SCALE

    if font_pt <= NEW_REPORT_TABLE_OVERFLOW_FONT_PT + 1e-9:
        leading = NEW_REPORT_TABLE_OVERFLOW_LINE_HEIGHT_PT
    else:
        leading = NEW_REPORT_TABLE_LINE_HEIGHT_PT

    # Không ép nhỏ hơn mức này để tránh chữ chạm/đè đường kẻ.
    padding = max(0.0, (target_height - leading) / 2.0)
    return leading, padding


# V1.0.12: New form dùng cỡ cơ sở 12 pt; hàng bảng có nguy cơ tràn tự giảm 11 pt; chiều cao hàng khoảng 70% bản đầu.
# V1.0.11: thu gọn chiều cao hàng còn khoảng 80% và tăng cỡ chữ ô bảng lên 10 pt cho New form.
# V1.0.10: bổ sung phân loại tổ hợp theo TTGH và tổng hợp Max/Min riêng trong báo cáo.
logger = logging.getLogger("N2D.Pile")


def _ts_app_base_dir():
    """Base folder for source mode, PyInstaller, and Nuitka standalone."""
    candidates = []
    try:
        if getattr(sys, "frozen", False) or "__compiled__" in globals():
            candidates.append(os.path.dirname(os.path.abspath(sys.executable)))
    except Exception:
        pass
    try:
        candidates.append(os.path.dirname(os.path.abspath(sys.executable)))
    except Exception:
        pass
    try:
        candidates.append(os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        pass
    for path in candidates:
        if path and os.path.isdir(path):
            return path
    return os.getcwd()


def _ts_icon_path(filename):
    base_dir = _ts_app_base_dir()
    candidates = []
    try:
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            candidates.extend([
                os.path.join(meipass, "assets", "icons", filename),
                os.path.join(meipass, filename),
            ])
    except Exception:
        pass
    candidates.extend([
        os.path.join(base_dir, "assets", "icons", filename),
        os.path.join(base_dir, filename),
        os.path.join(os.getcwd(), "assets", "icons", filename),
    ])
    try:
        candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "icons", filename))
    except Exception:
        pass
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return None


def apply_app_icon(window, icon_stem):
    """Apply window icon in source, PyInstaller, and Nuitka builds."""
    try:
        ico_path = _ts_icon_path(f"{icon_stem}.ico")
        if ico_path and os.name == "nt":
            window.iconbitmap(default=ico_path)
    except Exception:
        pass
    try:
        png_path = _ts_icon_path(f"{icon_stem}.png")
        if png_path:
            import tkinter as _tk
            _icon_image = _tk.PhotoImage(file=png_path)
            refs = getattr(window, "_ts_icon_refs", [])
            refs.append(_icon_image)
            window._ts_icon_refs = refs
            window.iconphoto(True, _icon_image)
    except Exception:
        pass



# ==========================================
# THƯ VIỆN THEME N2D - ĐỒNG BỘ TS-PILE / TS-COL
# ==========================================
# Ghi chú:
# - Bộ màu này chỉ ảnh hưởng giao diện, không ảnh hưởng lõi tính toán.
# - Các theme lấy cảm hứng từ các phong cách phổ biến trong cộng đồng kỹ thuật/lập trình
#   như Catppuccin, Dracula, Nord, Gruvbox, One Dark Pro, Tokyo Night, Ayu,
#   cùng các theme kỹ thuật riêng cho phần mềm kết cấu/nền móng.
N2D_THEME_LIBRARY_VERSION = "20260629.2"

UI_THEME_AQUA = "AQUA_FRESH"
UI_THEME_CLEAN_LIGHT = "CLEAN_LIGHT"
UI_THEME_FLUENT_LIGHT = "FLUENT_LIGHT"
UI_THEME_FLUENT_DARK = "FLUENT_DARK"
UI_THEME_INDUSTRIAL_DARK = "INDUSTRIAL_DARK"
UI_THEME_CATPPUCCIN_LATTE = "CATPPUCCIN_LATTE"
UI_THEME_CATPPUCCIN_MOCHA = "CATPPUCCIN_MOCHA"
UI_THEME_DRACULA = "DRACULA"
UI_THEME_NORD = "NORD"
UI_THEME_GRUVBOX_DARK = "GRUVBOX_DARK"
UI_THEME_ONE_DARK = "ONE_DARK_PRO"
UI_THEME_TOKYO_NIGHT = "TOKYO_NIGHT"
UI_THEME_AYU_LIGHT = "AYU_LIGHT"
UI_THEME_AYU_MIRAGE = "AYU_MIRAGE"
UI_THEME_AYU_DARK = "AYU_DARK"
UI_THEME_OCEAN = "OCEAN_PRO"
UI_THEME_EMERALD = "EMERALD_BRIDGE"
UI_THEME_ROYAL = "ROYAL_BLUE"
UI_THEME_STEEL = "STEEL_GRAY"
UI_THEME_WARM = "WARM_SAND"
UI_THEME_DARK = "DARK_SLATE"
UI_THEME_GRAPHITE = "GRAPHITE_PRO"
UI_THEME_CLASSIC = "CLASSIC_BROWN"

UI_THEME_LABELS = {
    UI_THEME_AQUA: "Aqua Fresh - xanh ngọc sáng",
    UI_THEME_CLEAN_LIGHT: "Clean Light / Material - sáng tối giản",
    UI_THEME_FLUENT_LIGHT: "Fluent Light - Windows 11 sáng",
    UI_THEME_FLUENT_DARK: "Fluent Dark - Windows 11 tối",
    UI_THEME_INDUSTRIAL_DARK: "Industrial Dark - xám công nghiệp",
    UI_THEME_CATPPUCCIN_LATTE: "Catppuccin Latte - pastel sáng",
    UI_THEME_CATPPUCCIN_MOCHA: "Catppuccin Mocha - pastel tối",
    UI_THEME_DRACULA: "Dracula - neon tối",
    UI_THEME_NORD: "Nord - Bắc Âu lạnh",
    UI_THEME_GRUVBOX_DARK: "Gruvbox Dark - retro ấm",
    UI_THEME_ONE_DARK: "One Dark Pro - code tối",
    UI_THEME_TOKYO_NIGHT: "Tokyo Night - neon xanh đêm",
    UI_THEME_AYU_LIGHT: "Ayu Light - tối giản sáng",
    UI_THEME_AYU_MIRAGE: "Ayu Mirage - tối dịu",
    UI_THEME_AYU_DARK: "Ayu Dark - tối gọn",
    UI_THEME_OCEAN: "Ocean Pro - xanh biển kỹ thuật",
    UI_THEME_EMERALD: "Emerald Bridge - xanh cầu đường",
    UI_THEME_ROYAL: "Royal Blue - xanh dương hiện đại",
    UI_THEME_STEEL: "Steel Gray - xám thép chuyên nghiệp",
    UI_THEME_WARM: "Warm Sand - vàng cát nhẹ",
    UI_THEME_DARK: "Dark Slate - tối xanh kỹ thuật",
    UI_THEME_GRAPHITE: "Graphite Pro - tối than cao cấp",
    UI_THEME_CLASSIC: "Classic Brown - theme cũ",
}

UI_THEME_ORDER = [
    UI_THEME_CLEAN_LIGHT,
    UI_THEME_FLUENT_LIGHT,
    UI_THEME_AQUA,
    UI_THEME_OCEAN,
    UI_THEME_EMERALD,
    UI_THEME_ROYAL,
    UI_THEME_STEEL,
    UI_THEME_WARM,
    UI_THEME_CATPPUCCIN_LATTE,
    UI_THEME_CATPPUCCIN_MOCHA,
    UI_THEME_DRACULA,
    UI_THEME_NORD,
    UI_THEME_GRUVBOX_DARK,
    UI_THEME_ONE_DARK,
    UI_THEME_TOKYO_NIGHT,
    UI_THEME_AYU_LIGHT,
    UI_THEME_AYU_MIRAGE,
    UI_THEME_AYU_DARK,
    UI_THEME_FLUENT_DARK,
    UI_THEME_INDUSTRIAL_DARK,
    UI_THEME_DARK,
    UI_THEME_GRAPHITE,
    UI_THEME_CLASSIC,
]
UI_THEME_CHOICES = [UI_THEME_LABELS[k] for k in UI_THEME_ORDER]

# Palette chuẩn dùng chung. Các key bổ sung như tree_row/tree_fg/note không làm hỏng bản cũ;
# tool nào cần thì dùng, tool nào không cần thì bỏ qua.
THEME_PRESETS = {
    UI_THEME_AQUA: {
        "bg": "#edfaff", "panel": "#f8fdff", "sidebar": "#007987", "sidebar2": "#9bdced",
        "accent": "#0097a7", "accent_dark": "#007987", "text": "#073b4c", "muted": "#49656d",
        "button": "#dff6ff", "button_active": "#c7efff", "big_button": "#c9f2ff",
        "tree_head": "#dff6ff", "tree_row": "#ffffff", "tree_alt": "#f3fcff", "tree_fg": "#0b2533",
        "progress": "#00a884", "trough": "#d9f7ef", "border": "#9bdced",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#49656d"
    },
    UI_THEME_CLEAN_LIGHT: {
        "bg": "#f5f7fb", "panel": "#ffffff", "sidebar": "#2563eb", "sidebar2": "#bfdbfe",
        "accent": "#2563eb", "accent_dark": "#1d4ed8", "text": "#111827", "muted": "#64748b",
        "button": "#eaf1ff", "button_active": "#dbeafe", "big_button": "#dbeafe",
        "tree_head": "#e5edff", "tree_row": "#ffffff", "tree_alt": "#f8fafc", "tree_fg": "#111827",
        "progress": "#16a34a", "trough": "#dcfce7", "border": "#cbd5e1",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#64748b"
    },
    UI_THEME_FLUENT_LIGHT: {
        "bg": "#f3f6fb", "panel": "#ffffff", "sidebar": "#1f6feb", "sidebar2": "#a7c7ff",
        "accent": "#0078d4", "accent_dark": "#005a9e", "text": "#202020", "muted": "#5f6b7a",
        "button": "#eef4ff", "button_active": "#d9eaff", "big_button": "#d8e8ff",
        "tree_head": "#e9f1ff", "tree_row": "#ffffff", "tree_alt": "#f7fbff", "tree_fg": "#202020",
        "progress": "#107c10", "trough": "#e7f6e7", "border": "#c7d7ea",
        "entry_bg": "#ffffff", "entry_fg": "#202020", "note": "#5f6b7a"
    },
    UI_THEME_FLUENT_DARK: {
        "bg": "#202020", "panel": "#2b2b2b", "sidebar": "#111827", "sidebar2": "#3b82f6",
        "accent": "#60a5fa", "accent_dark": "#3b82f6", "text": "#f3f4f6", "muted": "#cbd5e1",
        "button": "#303642", "button_active": "#3b4657", "big_button": "#263c58",
        "tree_head": "#374151", "tree_row": "#ffffff", "tree_alt": "#f6f7fb", "tree_fg": "#111827",
        "progress": "#22c55e", "trough": "#193325", "border": "#4b5563",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#cbd5e1"
    },
    UI_THEME_INDUSTRIAL_DARK: {
        "bg": "#24272e", "panel": "#30343b", "sidebar": "#15171c", "sidebar2": "#f59e0b",
        "accent": "#f59e0b", "accent_dark": "#d97706", "text": "#f3f4f6", "muted": "#d1d5db",
        "button": "#3a3f48", "button_active": "#4b5563", "big_button": "#463b24",
        "tree_head": "#3f4652", "tree_row": "#ffffff", "tree_alt": "#f7f7f7", "tree_fg": "#111827",
        "progress": "#22c55e", "trough": "#26382f", "border": "#6b7280",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#d1d5db"
    },
    UI_THEME_CATPPUCCIN_LATTE: {
        "bg": "#eff1f5", "panel": "#ffffff", "sidebar": "#7287fd", "sidebar2": "#b4befe",
        "accent": "#8839ef", "accent_dark": "#7287fd", "text": "#4c4f69", "muted": "#6c6f85",
        "button": "#e6e9ef", "button_active": "#dce0e8", "big_button": "#e0e4ff",
        "tree_head": "#e6e9ef", "tree_row": "#ffffff", "tree_alt": "#f7f8fc", "tree_fg": "#4c4f69",
        "progress": "#40a02b", "trough": "#e7f3e4", "border": "#ccd0da",
        "entry_bg": "#ffffff", "entry_fg": "#4c4f69", "note": "#6c6f85"
    },
    UI_THEME_CATPPUCCIN_MOCHA: {
        "bg": "#1e1e2e", "panel": "#313244", "sidebar": "#11111b", "sidebar2": "#b4befe",
        "accent": "#cba6f7", "accent_dark": "#b4befe", "text": "#cdd6f4", "muted": "#bac2de",
        "button": "#45475a", "button_active": "#585b70", "big_button": "#3f3655",
        "tree_head": "#45475a", "tree_row": "#ffffff", "tree_alt": "#f8f7ff", "tree_fg": "#111827",
        "progress": "#a6e3a1", "trough": "#243845", "border": "#6c7086",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#bac2de"
    },
    UI_THEME_DRACULA: {
        "bg": "#282a36", "panel": "#343746", "sidebar": "#1e1f29", "sidebar2": "#ff79c6",
        "accent": "#bd93f9", "accent_dark": "#ff79c6", "text": "#f8f8f2", "muted": "#d6d6e7",
        "button": "#44475a", "button_active": "#555a70", "big_button": "#4a3b62",
        "tree_head": "#44475a", "tree_row": "#ffffff", "tree_alt": "#fbfaff", "tree_fg": "#111827",
        "progress": "#50fa7b", "trough": "#273b2d", "border": "#6272a4",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#d6d6e7"
    },
    UI_THEME_NORD: {
        "bg": "#eceff4", "panel": "#ffffff", "sidebar": "#2e3440", "sidebar2": "#88c0d0",
        "accent": "#5e81ac", "accent_dark": "#3b4252", "text": "#2e3440", "muted": "#4c566a",
        "button": "#e5e9f0", "button_active": "#d8dee9", "big_button": "#d7e5ef",
        "tree_head": "#e5e9f0", "tree_row": "#ffffff", "tree_alt": "#f8fafc", "tree_fg": "#2e3440",
        "progress": "#8fbcbb", "trough": "#e0f2f0", "border": "#b8c0cc",
        "entry_bg": "#ffffff", "entry_fg": "#2e3440", "note": "#4c566a"
    },
    UI_THEME_GRUVBOX_DARK: {
        "bg": "#282828", "panel": "#3c3836", "sidebar": "#1d2021", "sidebar2": "#fabd2f",
        "accent": "#fabd2f", "accent_dark": "#d79921", "text": "#ebdbb2", "muted": "#d5c4a1",
        "button": "#504945", "button_active": "#665c54", "big_button": "#5a4725",
        "tree_head": "#504945", "tree_row": "#ffffff", "tree_alt": "#fffaf0", "tree_fg": "#111827",
        "progress": "#b8bb26", "trough": "#3d4324", "border": "#7c6f64",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#d5c4a1"
    },
    UI_THEME_ONE_DARK: {
        "bg": "#282c34", "panel": "#323842", "sidebar": "#1f2329", "sidebar2": "#61afef",
        "accent": "#61afef", "accent_dark": "#528bff", "text": "#abb2bf", "muted": "#c8ccd4",
        "button": "#3a404a", "button_active": "#4b5263", "big_button": "#2e465f",
        "tree_head": "#3a404a", "tree_row": "#ffffff", "tree_alt": "#f6f9ff", "tree_fg": "#111827",
        "progress": "#98c379", "trough": "#2f3f32", "border": "#5c6370",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#c8ccd4"
    },
    UI_THEME_TOKYO_NIGHT: {
        "bg": "#1a1b26", "panel": "#24283b", "sidebar": "#11121d", "sidebar2": "#7aa2f7",
        "accent": "#7aa2f7", "accent_dark": "#bb9af7", "text": "#c0caf5", "muted": "#a9b1d6",
        "button": "#2f3549", "button_active": "#3b4261", "big_button": "#283a61",
        "tree_head": "#2f3549", "tree_row": "#ffffff", "tree_alt": "#f5f7ff", "tree_fg": "#111827",
        "progress": "#9ece6a", "trough": "#283b35", "border": "#565f89",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#a9b1d6"
    },
    UI_THEME_AYU_LIGHT: {
        "bg": "#fafafa", "panel": "#ffffff", "sidebar": "#55b4d4", "sidebar2": "#ffb454",
        "accent": "#399ee6", "accent_dark": "#1883c7", "text": "#5c6773", "muted": "#7f8c98",
        "button": "#eef5fa", "button_active": "#dceef8", "big_button": "#dff0fb",
        "tree_head": "#eef5fa", "tree_row": "#ffffff", "tree_alt": "#fafcff", "tree_fg": "#5c6773",
        "progress": "#86b300", "trough": "#edf6d8", "border": "#c7d3df",
        "entry_bg": "#ffffff", "entry_fg": "#5c6773", "note": "#7f8c98"
    },
    UI_THEME_AYU_MIRAGE: {
        "bg": "#1f2430", "panel": "#2b3141", "sidebar": "#171b24", "sidebar2": "#ffcc66",
        "accent": "#ffcc66", "accent_dark": "#ffb454", "text": "#cbccc6", "muted": "#b8c0cc",
        "button": "#343d4d", "button_active": "#414b5f", "big_button": "#4b3f2c",
        "tree_head": "#343d4d", "tree_row": "#ffffff", "tree_alt": "#fffaf0", "tree_fg": "#111827",
        "progress": "#bae67e", "trough": "#33402a", "border": "#5c6773",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#b8c0cc"
    },
    UI_THEME_AYU_DARK: {
        "bg": "#0b0e14", "panel": "#161b22", "sidebar": "#03070d", "sidebar2": "#ffb454",
        "accent": "#e6b450", "accent_dark": "#ff8f40", "text": "#b3b1ad", "muted": "#c8c3bc",
        "button": "#1f2530", "button_active": "#27303d", "big_button": "#3c3020",
        "tree_head": "#1f2530", "tree_row": "#ffffff", "tree_alt": "#fffaf2", "tree_fg": "#111827",
        "progress": "#c2d94c", "trough": "#2c3420", "border": "#3e4b59",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#c8c3bc"
    },
    UI_THEME_OCEAN: {
        "bg": "#eef7fb", "panel": "#ffffff", "sidebar": "#0b4f6c", "sidebar2": "#7cc6d9",
        "accent": "#0b7fab", "accent_dark": "#084c61", "text": "#0b2533", "muted": "#526d7a",
        "button": "#d8eef7", "button_active": "#bfe2ef", "big_button": "#c5e8f4",
        "tree_head": "#d9edf7", "tree_row": "#ffffff", "tree_alt": "#f4fbfe", "tree_fg": "#0b2533",
        "progress": "#0e9f9a", "trough": "#d9f0ee", "border": "#8cc7d8",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#526d7a"
    },
    UI_THEME_EMERALD: {
        "bg": "#eefaf3", "panel": "#ffffff", "sidebar": "#0f6b4b", "sidebar2": "#8bd8b4",
        "accent": "#12805c", "accent_dark": "#0b5d43", "text": "#102d22", "muted": "#4d675c",
        "button": "#def5e8", "button_active": "#c6ebd8", "big_button": "#cdeedb",
        "tree_head": "#d8f0e4", "tree_row": "#ffffff", "tree_alt": "#f5fcf8", "tree_fg": "#102d22",
        "progress": "#16a34a", "trough": "#dcfce7", "border": "#8bd8b4",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#4d675c"
    },
    UI_THEME_ROYAL: {
        "bg": "#f3f7ff", "panel": "#ffffff", "sidebar": "#1e3a8a", "sidebar2": "#93b4ff",
        "accent": "#2563eb", "accent_dark": "#1e40af", "text": "#111827", "muted": "#4b5563",
        "button": "#e0ebff", "button_active": "#cbdcff", "big_button": "#d9e7ff",
        "tree_head": "#dbe8ff", "tree_row": "#ffffff", "tree_alt": "#f6f9ff", "tree_fg": "#111827",
        "progress": "#2563eb", "trough": "#dbeafe", "border": "#9bbcff",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#4b5563"
    },
    UI_THEME_STEEL: {
        "bg": "#f3f5f7", "panel": "#ffffff", "sidebar": "#334155", "sidebar2": "#94a3b8",
        "accent": "#475569", "accent_dark": "#1f2937", "text": "#111827", "muted": "#64748b",
        "button": "#e2e8f0", "button_active": "#cbd5e1", "big_button": "#d7dee8",
        "tree_head": "#e2e8f0", "tree_row": "#ffffff", "tree_alt": "#f8fafc", "tree_fg": "#111827",
        "progress": "#0f766e", "trough": "#d7eee9", "border": "#94a3b8",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#64748b"
    },
    UI_THEME_WARM: {
        "bg": "#fff8ed", "panel": "#fffdf8", "sidebar": "#9a5b13", "sidebar2": "#e7c48a",
        "accent": "#b86b12", "accent_dark": "#88430d", "text": "#3b2f20", "muted": "#715f47",
        "button": "#f8ead2", "button_active": "#efd8b3", "big_button": "#f3dfbd",
        "tree_head": "#f7e8cf", "tree_row": "#ffffff", "tree_alt": "#fffaf2", "tree_fg": "#3b2f20",
        "progress": "#ca8a04", "trough": "#fef3c7", "border": "#ddb978",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#715f47"
    },
    UI_THEME_DARK: {
        "bg": "#111827", "panel": "#1f2937", "sidebar": "#020617", "sidebar2": "#38bdf8",
        "accent": "#38bdf8", "accent_dark": "#0ea5e9", "text": "#e5f2ff", "muted": "#c4d4e8",
        "button": "#263653", "button_active": "#334b73", "big_button": "#284a66",
        "tree_head": "#243957", "tree_row": "#ffffff", "tree_alt": "#f5f7fb", "tree_fg": "#111827",
        "progress": "#22c55e", "trough": "#102033", "border": "#3b82f6",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#c4d4e8"
    },
    UI_THEME_GRAPHITE: {
        "bg": "#171717", "panel": "#262626", "sidebar": "#0a0a0a", "sidebar2": "#737373",
        "accent": "#a3e635", "accent_dark": "#65a30d", "text": "#f5f5f5", "muted": "#d4d4d4",
        "button": "#404040", "button_active": "#525252", "big_button": "#3f4d2a",
        "tree_head": "#404040", "tree_row": "#ffffff", "tree_alt": "#f6f6f6", "tree_fg": "#111827",
        "progress": "#84cc16", "trough": "#2f3b1f", "border": "#737373",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#d4d4d4"
    },
    UI_THEME_CLASSIC: {
        "bg": "#f2eee7", "panel": "#fffaf1", "sidebar": "#9a6a3a", "sidebar2": "#c6aa87",
        "accent": "#8b5e34", "accent_dark": "#6f4420", "text": "#3b2f2f", "muted": "#6b5b4a",
        "button": "#eadfcd", "button_active": "#dfcfb8", "big_button": "#efe1c8",
        "tree_head": "#eadfcd", "tree_row": "#ffffff", "tree_alt": "#fffaf1", "tree_fg": "#3b2f2f",
        "progress": "#16a34a", "trough": "#e4dccf", "border": "#c6aa87",
        "entry_bg": "#ffffff", "entry_fg": "#111827", "note": "#6b5b4a"
    },
}

_THEME_ALIAS_KEYWORDS = [
    (UI_THEME_CATPPUCCIN_LATTE, ("CATPPUCCINLATTE", "LATTE", "PASTELSANG")),
    (UI_THEME_CATPPUCCIN_MOCHA, ("CATPPUCCINMOCHA", "MOCHA", "PASTELTOI", "CATPPUCCIN")),
    (UI_THEME_DRACULA, ("DRACULA", "NEONHONG", "NEON")),
    (UI_THEME_NORD, ("NORD", "BACAULANH", "BACAU", "TUYET", "LANH")),
    (UI_THEME_GRUVBOX_DARK, ("GRUVBOX", "RETRO", "DAT", "VANGAM")),
    (UI_THEME_ONE_DARK, ("ONEDARK", "ONEDARKPRO", "CODETOI")),
    (UI_THEME_TOKYO_NIGHT, ("TOKYONIGHT", "TOKYO", "NEONXANH")),
    (UI_THEME_AYU_LIGHT, ("AYULIGHT", "AYUSANG")),
    (UI_THEME_AYU_MIRAGE, ("AYUMIRAGE", "MIRAGE", "TOIDIU")),
    (UI_THEME_AYU_DARK, ("AYUDARK", "AYUTOI", "AYU")),
    (UI_THEME_INDUSTRIAL_DARK, ("INDUSTRIAL", "CONGNGHIEP", "ETABS", "SAP2000", "AUTOCAD")),
    (UI_THEME_CLEAN_LIGHT, ("CLEAN", "MATERIAL", "TOIGIAN", "SANGTOIGIAN")),
    (UI_THEME_FLUENT_LIGHT, ("FLUENTLIGHT", "WINDOWS11SANG", "FLUENTSANG")),
    (UI_THEME_FLUENT_DARK, ("FLUENTDARK", "WINDOWS11TOI", "FLUENT")),
    (UI_THEME_GRAPHITE, ("GRAPHITE", "CARBON", "THAN")),
    (UI_THEME_DARK, ("DARKSLATE", "SLATE", "TOIXANH")),
    (UI_THEME_WARM, ("WARMSAND", "SAND", "CAT", "VANG")),
    (UI_THEME_STEEL, ("STEEL", "GRAY", "GREY", "XAM", "THEP")),
    (UI_THEME_ROYAL, ("ROYAL", "BLUE", "DUONG")),
    (UI_THEME_EMERALD, ("EMERALD", "BRIDGE", "CAUDUONG", "XANHLA")),
    (UI_THEME_OCEAN, ("OCEAN", "BIEN", "KYTHUAT")),
    (UI_THEME_CLASSIC, ("BROWN", "CLASSIC", "CU")),
    (UI_THEME_AQUA, ("AQUA", "FRESH", "XANHNGOC", "XANH")),
]

def _theme_compact(value):
    s = str(value or "").strip().upper().replace(" ", "_").replace("-", "_").replace("/", "_")
    s = (s.replace("Đ", "D").replace("Ư", "U").replace("Ơ", "O").replace("Ê", "E")
           .replace("Á", "A").replace("À", "A").replace("Ả", "A").replace("Ã", "A").replace("Ạ", "A")
           .replace("Ắ", "A").replace("Ằ", "A").replace("Ẳ", "A").replace("Ẵ", "A").replace("Ặ", "A")
           .replace("Ấ", "A").replace("Ầ", "A").replace("Ẩ", "A").replace("Ẫ", "A").replace("Ậ", "A")
           .replace("Í", "I").replace("Ì", "I").replace("Ỉ", "I").replace("Ĩ", "I").replace("Ị", "I")
           .replace("Ó", "O").replace("Ò", "O").replace("Ỏ", "O").replace("Õ", "O").replace("Ọ", "O")
           .replace("Ố", "O").replace("Ồ", "O").replace("Ổ", "O").replace("Ỗ", "O").replace("Ộ", "O")
           .replace("Ớ", "O").replace("Ờ", "O").replace("Ở", "O").replace("Ỡ", "O").replace("Ợ", "O")
           .replace("Ú", "U").replace("Ù", "U").replace("Ủ", "U").replace("Ũ", "U").replace("Ụ", "U")
           .replace("Ứ", "U").replace("Ừ", "U").replace("Ử", "U").replace("Ữ", "U").replace("Ự", "U")
           .replace("Ý", "Y").replace("Ỳ", "Y").replace("Ỷ", "Y").replace("Ỹ", "Y").replace("Ỵ", "Y"))
    return s.replace("_", "")

def normalize_ui_theme(value=""):
    s = str(value or "").strip()
    if not s:
        return UI_THEME_AQUA
    u = s.upper().replace(" ", "_").replace("-", "_").replace("/", "_")
    if u in UI_THEME_LABELS:
        return u
    cu = _theme_compact(u)
    for key, label in UI_THEME_LABELS.items():
        if _theme_compact(label) == cu:
            return key
    for key, words in _THEME_ALIAS_KEYWORDS:
        if any(w in cu for w in words):
            return key
    return UI_THEME_AQUA

def ui_theme_label(value=""):
    return UI_THEME_LABELS.get(normalize_ui_theme(value), UI_THEME_LABELS[UI_THEME_AQUA])

def get_theme_palette(value=""):
    key = normalize_ui_theme(value)
    return THEME_PRESETS.get(key, THEME_PRESETS[UI_THEME_AQUA])

def make_col_theme_presets():
    'Tạo palette rút gọn cho TS-COL, giữ tương thích với code style cũ.'
    out = {}
    for key in UI_THEME_ORDER:
        pal = dict(THEME_PRESETS[key])
        pal.setdefault("note", pal.get("muted", pal.get("text", "#555555")))
        pal.setdefault("tree_row", "#ffffff")
        pal.setdefault("tree_fg", "#111827")
        out[key] = pal
    return out

# Màu mặc định cho các vùng giao diện khởi động/chưa áp theme động.
APP_BG = THEME_PRESETS[UI_THEME_AQUA]["bg"]
APP_NAV = THEME_PRESETS[UI_THEME_AQUA]["sidebar"]
APP_PRIMARY = THEME_PRESETS[UI_THEME_AQUA]["accent"]
APP_SUCCESS = THEME_PRESETS[UI_THEME_AQUA]["progress"]

# ==========================================
# KHỐI 1: BẢO MẬT & KIỂM TRA BẢN QUYỀN
# ==========================================
SECRET_SALT = "Dung_Dev_Security_Key_2026!@#"

# ==========================================
# QA fix R1/R3: cache kích hoạt để chạy offline có thời hạn + ổn định machine-id
# ==========================================
LICENSE_OFFLINE_GRACE_DAYS = 14
N2D_LICENSE_INFO = {"message": "", "reason": ""}

def _n2d_license_dir():
    base = os.path.join(os.path.expanduser("~"), ".n2d_license")
    try:
        os.makedirs(base, exist_ok=True)
    except Exception:
        pass
    return base

def _license_cache_path():
    return os.path.join(_n2d_license_dir(), "ts_suite_license.json")

def _machine_id_store_path():
    return os.path.join(_n2d_license_dir(), "machine_id.txt")

def _offline_trial_path():
    return os.path.join(_n2d_license_dir(), "ts_suite_offline_trial.json")

def _offline_trial_sig(machine_id, day):
    return hashlib.sha256(f"{machine_id}|{day}|OFFLINE_TRIAL|{SECRET_SALT}".encode("utf-8")).hexdigest()

def _offline_trial_save(machine_id, day=None):
    try:
        if day is None:
            day = datetime.datetime.now().strftime("%Y-%m-%d")
        data = {"machine_id": str(machine_id), "start_day": str(day),
                "sig": _offline_trial_sig(machine_id, day)}
        with open(_offline_trial_path(), "w", encoding="utf-8") as fh:
            json.dump(data, fh)
        return True
    except Exception:
        return False

def _offline_trial_load(machine_id):
    try:
        with open(_offline_trial_path(), "r", encoding="utf-8") as fh:
            data = json.load(fh)
        if str(data.get("machine_id", "")) != str(machine_id):
            return None
        day = str(data.get("start_day", ""))
        if str(data.get("sig", "")) != _offline_trial_sig(machine_id, day):
            return None
        return day
    except Exception:
        return None

def _offline_trial_check(machine_id):
    """Trả (được chạy?, ngày offline còn lại). Nếu chưa có file trial thì tạo mới."""
    day = _offline_trial_load(machine_id)
    if not day:
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        if _offline_trial_save(machine_id, today):
            return True, LICENSE_OFFLINE_GRACE_DAYS
        return False, 0
    try:
        start = datetime.datetime.strptime(day, "%Y-%m-%d")
        delta = (datetime.datetime.now() - start).days
        if delta < 0:
            return False, 0
        remain = max(LICENSE_OFFLINE_GRACE_DAYS - delta, 0)
        return remain > 0, remain
    except Exception:
        return False, 0

def _license_cache_sig(machine_id, day, days_left):
    return hashlib.sha256(f"{machine_id}|{day}|{days_left}|{SECRET_SALT}".encode("utf-8")).hexdigest()

def _license_cache_save(machine_id, days_left):
    try:
        day = datetime.datetime.now().strftime("%Y-%m-%d")
        data = {"machine_id": str(machine_id), "last_ok": day, "days_left": days_left,
                "sig": _license_cache_sig(machine_id, day, days_left)}
        with open(_license_cache_path(), "w", encoding="utf-8") as fh:
            json.dump(data, fh)
    except Exception:
        pass

def _license_cache_clear():
    try:
        os.remove(_license_cache_path())
    except Exception:
        pass

def _license_offline_check(machine_id):
    """Trả (được chạy offline?, ngày bản quyền còn lại, ngày offline còn lại).

    14 ngày chỉ là giới hạn chạy offline. Giá trị trả về cho GUI/app vẫn là
    số ngày bản quyền ước tính còn lại từ lần server xác nhận gần nhất.
    """
    try:
        with open(_license_cache_path(), "r", encoding="utf-8") as fh:
            data = json.load(fh)
        if str(data.get("machine_id", "")) != str(machine_id):
            return False, 0, 0
        day = str(data.get("last_ok", ""))
        if str(data.get("sig", "")) != _license_cache_sig(machine_id, day, data.get("days_left")):
            return False, 0, 0
        last = datetime.datetime.strptime(day, "%Y-%m-%d")
        delta = (datetime.datetime.now() - last).days
        if delta < 0:
            return False, 0, 0
        cached_days = int(float(data.get("days_left")))
        license_remaining = max(cached_days - delta, 0)
        offline_remaining = max(LICENSE_OFFLINE_GRACE_DAYS - delta, 0)
        ok = (cached_days - delta) > 0 and offline_remaining > 0
        return ok, license_remaining, offline_remaining
    except Exception:
        return False, 0, 0

def get_machine_id():
    """Lấy mã máy ổn định để kiểm tra license.

    Ưu tiên UUID Windows để tương thích các máy đã kích hoạt. Nếu PowerShell/WMIC lỗi,
    dùng fallback thay vì tự tắt app không thông báo.
    """
    hwid = ""
    errors = []
    if platform.system() == "Windows":
        for cmd in (
            'powershell "(Get-CimInstance -Class Win32_ComputerSystemProduct).UUID"',
            'wmic csproduct get uuid',
        ):
            try:
                out = subprocess.check_output(cmd, shell=True, stderr=subprocess.DEVNULL).decode(errors="ignore").strip()
                if "wmic" in cmd.lower():
                    parts = [line.strip() for line in out.splitlines() if line.strip() and "uuid" not in line.lower()]
                    out = parts[0] if parts else ""
                if out:
                    hwid = out
                    break
            except Exception as exc:
                errors.append(str(exc))
    if hwid:
        mid = hashlib.sha256((str(hwid) + SECRET_SALT).encode()).hexdigest()
        try:
            with open(_machine_id_store_path(), "w", encoding="utf-8") as fh:
                fh.write(mid)
        except Exception:
            pass
        return mid
    # QA fix R3: dùng machine-id đã lưu trước khi rơi về fallback MAC, tránh lệch máy đã kích hoạt.
    try:
        with open(_machine_id_store_path(), "r", encoding="utf-8") as fh:
            stored = fh.read().strip()
        if stored:
            logger.warning("Không lấy được UUID Windows, dùng machine-id đã lưu. Chi tiết: %s", "; ".join(errors))
            return stored
    except Exception:
        pass
    node = platform.node() or "unknown-node"
    mac = uuid.getnode()
    hwid = f"{platform.system()}|{node}|{mac}|{platform.machine()}"
    logger.warning("Không lấy được UUID Windows, dùng machine-id fallback. Chi tiết: %s", "; ".join(errors))
    return hashlib.sha256((str(hwid) + SECRET_SALT).encode()).hexdigest()

def is_word_installed():
    if platform.system() != "Windows": return False
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, "Word.Application")
        winreg.CloseKey(key)
        return True
    except Exception:
        return False

def get_windows_printers():
    if platform.system() != "Windows": return ["Default Printer"]
    try:
        cmd = 'powershell -ExecutionPolicy Bypass -Command "Get-CimInstance -ClassName Win32_Printer | Select-Object -ExpandProperty Name"'
        out = subprocess.check_output(cmd, shell=True, stderr=subprocess.DEVNULL).decode('utf-8', 'ignore').strip()
        printers = [p.strip() for p in out.split('\n') if p.strip()]
        return ["Default Printer"] + printers if printers else ["Default Printer"]
    except:
        return ["Default Printer"]

try:
    # Lazy dependencies: các thư viện nặng chỉ được import khi thật sự cần.
    np = None
    pd = None
    openpyxl = None
    Alignment = None
    pypdf = None
    Image = None
    ImageTk = None
    Canvas = None
    A4 = None
    pdfmetrics = None
    TTFont = None

    def ensure_numpy():
        global np
        if np is None:
            import numpy as _np
            np = _np
        return np

    def ensure_pandas():
        global pd
        if pd is None:
            import pandas as _pd
            pd = _pd
        return pd

    def ensure_openpyxl():
        global openpyxl, Alignment
        if openpyxl is None:
            import openpyxl as _openpyxl
            from openpyxl.styles import Alignment as _Alignment
            openpyxl = _openpyxl
            Alignment = _Alignment
        return openpyxl, Alignment

    def ensure_pypdf():
        global pypdf
        if pypdf is None:
            import pypdf as _pypdf
            pypdf = _pypdf
        return pypdf

    def ensure_pil():
        global Image, ImageTk
        if Image is None or ImageTk is None:
            from PIL import Image as _Image, ImageTk as _ImageTk
            Image = _Image
            ImageTk = _ImageTk
        return Image, ImageTk

    def ensure_reportlab():
        global Canvas, A4, pdfmetrics, TTFont
        if Canvas is None:
            from reportlab.pdfgen.canvas import Canvas as _Canvas
            from reportlab.lib.pagesizes import A4 as _A4
            from reportlab.pdfbase import pdfmetrics as _pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont as _TTFont
            Canvas = _Canvas
            A4 = _A4
            pdfmetrics = _pdfmetrics
            TTFont = _TTFont
        return Canvas, A4, pdfmetrics, TTFont

    def check_server_trial():
        machine_id = get_machine_id()
        try:
            import requests
            response = requests.post(API_URL, json={"machine_id": machine_id}, timeout=8)
            data = response.json()
            status = str(data.get("status", "")).strip().lower()
            if status == "active":
                _license_cache_save(machine_id, data.get('days_left'))
                N2D_LICENSE_INFO.update({"message": "License active", "reason": "active"})
                return True, data.get('days_left')
            inactive_statuses = {"expired", "inactive", "revoked", "not_found", "blocked", "disabled", "not_active", "not-active", "unregistered"}
            if status in inactive_statuses:
                _license_cache_clear()
                N2D_LICENSE_INFO.update({"message": str(data.get("message", "Hết hạn hoặc chưa kích hoạt") or "Hết hạn hoặc chưa kích hoạt"), "reason": "expired"})
                return False, 0
            raise RuntimeError(f"Phản hồi license không hợp lệ: status={status or 'missing'}")
        except Exception as exc:
            ok_offline, license_days, offline_days = _license_offline_check(machine_id)
            if ok_offline:
                N2D_LICENSE_INFO.update({
                    "message": f"Chế độ offline (còn {offline_days} ngày offline)",
                    "reason": "offline",
                })
                return True, license_days
            ok_trial, trial_days = _offline_trial_check(machine_id)
            if ok_trial:
                N2D_LICENSE_INFO.update({
                    "message": f"Chế độ offline lần đầu (còn {trial_days} ngày)",
                    "reason": "offline_trial",
                })
                return True, trial_days
            N2D_LICENSE_INFO.update({
                "message": "Không kiểm tra được bản quyền online và đã quá 14 ngày dùng offline kể từ lần mở đầu tiên. Vui lòng kết nối mạng rồi mở lại phần mềm.",
                "reason": "offline_trial_expired",
            })
            return False, 0

    def resource_path(relative_path):
        if hasattr(sys, '_MEIPASS'): return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)

    # ==========================================
    # HELPER ĐỌC CLIPBOARD BẢNG TỔ HỢP
    # ==========================================
    _CLIPBOARD_ZERO_TOKENS = {"", "-", "–", "—", "−", "―", "‐", "‑"}

    def _parse_clipboard_number(value, blank_as_zero=False):
        """Đọc một ô số từ Excel/Clipboard theo cả định dạng Việt Nam và quốc tế.

        - Giữ đúng vị trí ô trống hoặc dấu gạch; khi blank_as_zero=True thì xem là 0.
        - Hỗ trợ dấu âm Unicode, ngoặc âm, decimal comma, dấu phân cách hàng nghìn.
        - Không âm thầm lấy một số nằm lẫn trong chuỗi chữ thông thường.
        """
        s = html.unescape(str(value if value is not None else ""))
        s = s.replace("\xa0", " ").strip().strip('"').strip("'")
        s = (s.replace("−", "-").replace("–", "-").replace("—", "-")
               .replace("―", "-").replace("‐", "-").replace("‑", "-"))
        if s in _CLIPBOARD_ZERO_TOKENS:
            return 0.0 if blank_as_zero else None

        negative = False
        if len(s) >= 2 and s.startswith("(") and s.endswith(")"):
            negative = True
            s = s[1:-1].strip()

        # Loại khoảng trắng phân cách hàng nghìn và ký hiệu đơn vị phổ biến ở cuối ô.
        s = re.sub(r"\s+", "", s)
        s = re.sub(r"(?i)(tf|tm|t\.m|kn|knm|kn\.m|n|m)$", "", s)
        if not s:
            return 0.0 if blank_as_zero else None

        # Chấp nhận tên tổ hợp dạng TH1/LC1/Combo1 ở cột chỉ số.
        m_index = re.fullmatch(r"(?i)(?:TH|LC|COMBO)?\s*([-+]?\d+)", s)
        if m_index:
            number = float(m_index.group(1))
            return -number if negative else number

        # Không chấp nhận chuỗi còn chữ ở giữa để tránh trôi cột âm thầm.
        if re.search(r"[A-DF-Za-df-z]", s):
            return None

        # Chuẩn hóa dấu thập phân/hàng nghìn.
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                # 1.234,56 -> 1234.56
                s = s.replace(".", "").replace(",", ".")
            else:
                # 1,234.56 -> 1234.56
                s = s.replace(",", "")
        elif "," in s:
            # Clipboard Excel theo locale Việt Nam thường dùng dấu phẩy thập phân.
            s = s.replace(",", ".")

        # Loại dấu nháy phân cách hàng nghìn còn sót.
        s = s.replace("’", "").replace("`", "")
        try:
            number = float(s)
        except Exception:
            return None
        return -number if negative else number

    def _split_clipboard_row(raw_line):
        """Tách một dòng clipboard nhưng bảo toàn ô trống khi có tab/semicolon."""
        raw_line = str(raw_line or "").rstrip("\r\n")
        if "\t" in raw_line:
            return raw_line.split("\t"), "tab"
        if ";" in raw_line:
            # Dùng csv để xử lý ô có dấu ngoặc kép.
            try:
                return next(csv.reader([raw_line], delimiter=";")), "semicolon"
            except Exception:
                return raw_line.split(";"), "semicolon"
        # Text cách nhau bằng khoảng trắng: không thể bảo toàn ô trống, dùng fallback.
        return re.split(r"\s+", raw_line.strip()), "space"

    def parse_load_clipboard_text(text):
        """Đọc bảng tổ hợp tải trọng từ clipboard.

        Trả về (rows, diagnostics), trong đó rows luôn có 7 cột:
        TH, Hx, Hy, N, Mx, My, Mz.
        """
        source = str(text or "").replace("\x00", "")
        raw_lines = re.split(r"[\r\n\u2028\u2029]+", source)
        rows = []
        rejected = []
        nonempty = 0

        for line_no, raw in enumerate(raw_lines, start=1):
            if not str(raw).strip():
                continue
            nonempty += 1
            cells, mode = _split_clipboard_row(raw)

            # Bỏ các cột trống thừa ở cuối, nhưng vẫn giữ đủ 6/7 cột dữ liệu.
            while len(cells) > 7 and str(cells[-1]).strip() == "":
                cells.pop()

            # Với dữ liệu có cấu trúc (tab/semicolon), bảo toàn dấu gạch và ô trống là 0.
            if mode in {"tab", "semicolon"}:
                parsed = [_parse_clipboard_number(c, blank_as_zero=True) for c in cells]
                if len(parsed) >= 7:
                    first = _parse_clipboard_number(cells[0], blank_as_zero=False)
                    loads = parsed[1:7]
                    if first is not None and all(v is not None for v in loads):
                        row = [first] + loads
                    else:
                        row = None
                elif len(parsed) >= 6:
                    loads = parsed[:6]
                    row = [len(rows) + 1] + loads if all(v is not None for v in loads) else None
                else:
                    row = None
            else:
                # Fallback cho text cách nhau bằng khoảng trắng. Dấu gạch đứng riêng được xem là 0.
                parsed = [_parse_clipboard_number(c, blank_as_zero=True) for c in cells]
                parsed = [v for v in parsed if v is not None]
                if len(parsed) >= 7 and abs(parsed[0] - round(parsed[0])) < 1e-8:
                    row = parsed[:7]
                elif len(parsed) >= 6:
                    row = [len(rows) + 1] + parsed[:6]
                else:
                    row = None

            if row is None:
                # Header được bỏ qua riêng, không tính là lỗi dữ liệu.
                if re.search(r"(?i)\b(TH|HX|HY|MX|MY|MZ|COMBO|TỔ\s*HỢP)\b", str(raw)):
                    continue
                rejected.append((line_no, str(raw)))
                continue

            if abs(float(row[0])) < 1e-12 and all(abs(float(v)) < 1e-12 for v in row[1:]):
                continue
            rows.append(row)

        return rows, {
            "nonempty_lines": nonempty,
            "accepted_rows": len(rows),
            "rejected_rows": rejected,
        }

    # ==========================================
    # 1. PARSER DỮ LIỆU ĐẦU VÀO
    # ==========================================
    def _parse_choice_int(value, default=0):
        """Đọc số lựa chọn ở đầu chuỗi Combobox, ví dụ ``0 - Cọc đóng`` -> 0."""
        try:
            m = re.match(r"\s*([-+]?\d+)", str(value or ""))
            return int(m.group(1)) if m else int(default)
        except Exception:
            return int(default)

    def _section_raw_values(p):
        """Ba trường hình học mặt cắt chuẩn sau Bpx/Bpy trong INPUT MCOC Ver.26.

        - Cọc tròn/ống: D ngoài, D trong, 0.
        - Cọc vuông/chữ nhật: A, B, Cday.

        Các khóa Sec1/Sec2/Sec3 được giữ để tương thích dữ liệu nội bộ cũ, nhưng
        không còn nhánh nhập trực tiếp F/Jx/Jy vì MCOC Ver.26 thực tế chỉ chạy khi
        trường điều khiển mặt cắt (token số 3) bằng 0.
        """
        sec1 = p.get('Sec1', p.get('d_ngoai', 0.0))
        sec2 = p.get('Sec2', p.get('d_trong', 0.0))
        sec3 = p.get('Sec3', p.get('day_vo', 0.0))
        return (float(sec1 or 0.0), float(sec2 or 0.0), float(sec3 or 0.0))

    def _mcoc_standard_section_components(section_flag, p):
        """Tính F, Jx, Jy thân cọc từ hình học chuẩn MCOC Ver.26.

        ``section_flag`` được giữ trong chữ ký để tương thích mã gọi cũ, nhưng giá
        trị hợp lệ duy nhất là 0. Thử nghiệm trực tiếp bằng MCOC Turbo BASIC Ver.26
        cho thấy file có token số 3 khác 0 không chạy, kể cả khi cấp F/Jx/Jy.
        """
        flag = int(round(float(section_flag)))
        if flag != 0:
            raise ValueError(
                f"Trường điều khiển mặt cắt MCOC (token số 3) phải bằng 0; hiện tại = {section_flag}"
            )

        a, b, t = _section_raw_values(p)
        if abs(t) < 1e-12:
            # Cọc tròn/cọc ống: a=D ngoài, b=D trong, c.day=0.
            D = max(a, 0.0)
            di = b if 0.0 < b < D else 0.0
            F_outer = math.pi * D**2 / 4.0
            F_core = math.pi * di**2 / 4.0
            J_outer = math.pi * D**4 / 64.0
            J_core = math.pi * di**4 / 64.0
            return {
                'F_total': F_outer, 'Jx_total': J_outer, 'Jy_total': J_outer,
                'F_shell': max(F_outer - F_core, 0.0), 'F_core': F_core,
                'Jx_shell': max(J_outer - J_core, 0.0), 'Jx_core': J_core,
                'Jy_shell': max(J_outer - J_core, 0.0), 'Jy_core': J_core,
            }

        # Cọc vuông/chữ nhật: a,b là kích thước ngoài; c.day là chiều dày.
        ax = max(a, 0.0)
        by = max(b, 0.0)
        if t > 0.0 and 2.0*t < ax and 2.0*t < by:
            ai = ax - 2.0*t
            bi = by - 2.0*t
        else:
            ai = bi = 0.0
        F_outer = ax * by
        F_core = ai * bi
        Jx_outer = ax * by**3 / 12.0
        Jx_core = ai * bi**3 / 12.0
        Jy_outer = by * ax**3 / 12.0
        Jy_core = bi * ai**3 / 12.0
        return {
            'F_total': F_outer, 'Jx_total': Jx_outer, 'Jy_total': Jy_outer,
            'F_shell': max(F_outer - F_core, 0.0), 'F_core': F_core,
            'Jx_shell': max(Jx_outer - Jx_core, 0.0), 'Jx_core': Jx_core,
            'Jy_shell': max(Jy_outer - Jy_core, 0.0), 'Jy_core': Jy_core,
        }

    def _validate_mcoc_unit_contract(data):
        """Chặn INPUT có dấu hiệu dùng mm/mm²/mm⁴ trong trường MCOC yêu cầu SI."""
        errors = []
        for p in data.get('Piles', []):
            name = p.get('Name', '?')
            bpx = abs(float(p.get('Bpx', 0.0) or 0.0))
            bpy = abs(float(p.get('Bpy', 0.0) or 0.0))
            sec1, sec2, sec3 = _section_raw_values(p)
            a, b, cday = abs(sec1), abs(sec2), abs(sec3)
            fo = abs(float(p.get('Area', 0.0) or 0.0))
            jo = abs(float(p.get('J_xy', 0.0) or 0.0))
            if bpx > 20.0 or bpy > 20.0:
                errors.append(f"Cọc {name}: Bpx/Bpy={bpx:g}/{bpy:g} quá lớn nếu tính bằng m")
            if max(a, b, cday) > 20.0:
                errors.append(f"Cọc {name}: kích thước mặt cắt A/B/Cday có dấu hiệu đang dùng mm")
            if fo > 100.0 or jo > 100.0:
                errors.append(f"Cọc {name}: Fo/Jo có dấu hiệu đang dùng mm²/mm⁴")
        if errors:
            raise ValueError(
                "INPUT MCOC sai hoặc không thống nhất đơn vị. MCOC yêu cầu chiều dài/kích thước bằng m, "
                "diện tích bằng m² và mô men quán tính bằng m⁴.\n" + "\n".join(errors[:10])
            )

    def parse_foundation_input(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                lines = [line.strip() for line in f.readlines() if line.strip()]
            if not lines:
                return "File input rỗng."

            data = {'Global': {}, 'Load_Combos': [], 'Piles': [], 'Config': {}}
            data['Global']['Project_ID'] = lines[0].strip()

            content = " ".join(lines[1:])
            tokens = []
            for t in content.split():
                try:
                    tokens.append(float(t))
                except ValueError:
                    pass

            if len(tokens) < 18:
                return "File input thiếu dữ liệu. Cần ít nhất 18 số đầu tiên."

            n_piles, n_combos = int(tokens[0]), int(tokens[1])
            section_flag_raw = float(tokens[2])
            section_flag = int(round(section_flag_raw))
            if abs(section_flag_raw - section_flag) > 1e-9 or section_flag != 0:
                return (
                    "Trường điều khiển mặt cắt MCOC tại token số 3 phải bằng 0 "
                    f"để tương thích Turbo BASIC Ver.26; hiện tại = {tokens[2]}"
                )
            expected = 18 + 6*n_combos + 16*n_piles
            if len(tokens) < expected:
                return f"Lỗi đọc file: Cần {expected} thông số, chỉ có {len(tokens)}."

            data['Global'].update({
                # Header MCOC Ver.26. Token số 3 bắt buộc bằng 0. Token số 4 là
                # cờ legacy đã được thử nghiệm 0/1/2/3 và không ảnh hưởng kết quả.
                'MCOC_Section_Flag': 0,
                'Section_Logic': 0,  # alias tương thích dữ liệu/call-site cũ
                'Legacy_Header_Flag': tokens[3],
                'Legacy_Reserved_1': tokens[4],
                'Pile_Count_Copy': tokens[5],
                'Legacy_Reserved_2': tokens[7],
                'Kn': tokens[6], 'Bx': tokens[8], 'By': tokens[9], 'Cz': tokens[10],
                'EI_uon': tokens[11], 'Er_uon': tokens[12],
                'EA_nen': tokens[13], 'Er_nen': tokens[14],
                # Thứ tự file INPUT: Md, Mq, m; báo cáo MCOC in Mq, Md, m.
                'md': tokens[15], 'mq': tokens[16], 'm': tokens[17],
            })

            idx = 18
            for _ in range(n_combos):
                data['Load_Combos'].append({
                    'Name': str(len(data['Load_Combos']) + 1),
                    'Hx': tokens[idx], 'Hy': tokens[idx+1], 'N_load': tokens[idx+2],
                    'Mx': tokens[idx+3], 'My': tokens[idx+4], 'Mz': tokens[idx+5],
                })
                idx += 6
            for _ in range(n_piles):
                data['Piles'].append({
                    'Name': str(len(data['Piles']) + 1),
                    'Lo': tokens[idx], 'H': tokens[idx+1],
                    'Bpx': tokens[idx+2], 'Bpy': tokens[idx+3],
                    # Ba trường hình học chuẩn; giữ alias Sec1/Sec2/Sec3 để tương thích ngược.
                    'Sec1': tokens[idx+4], 'Sec2': tokens[idx+5], 'Sec3': tokens[idx+6],
                    'd_ngoai': tokens[idx+4], 'd_trong': tokens[idx+5], 'day_vo': tokens[idx+6],
                    'Area': tokens[idx+7], 'J_xy': tokens[idx+8],
                    'Po': tokens[idx+9], 'Co': tokens[idx+10], 'Ct': tokens[idx+11],
                    'X': tokens[idx+12], 'Y': tokens[idx+13],
                    'Phi': tokens[idx+14], 'Xi': tokens[idx+15],
                })
                idx += 16

            _validate_mcoc_unit_contract(data)
            return data
        except Exception as e:
            return str(e)

    # ==========================================
    # 2. TÍNH TOÁN FEM & TCVN 11823
    # ==========================================
    class PiledRaftFoundation:
        def __init__(self, data):
            self.data = data
            self.K = None
            self.stiffness_warnings = []

        # Nghiệm cọc dài của hồ sơ MCOC (alpha*h >= 4).
        _MCOC_LONG_A0 = 2.441
        _MCOC_LONG_B0 = 1.621
        _MCOC_LONG_C0 = 1.751

        def _add_stiffness_warning(self, msg):
            if msg not in self.stiffness_warnings:
                self.stiffness_warnings.append(msg)

        def _equivalent_diameter(self, p):
            try:
                sec = _mcoc_standard_section_components(
                    0, p
                )
                area = max(float(sec.get('F_total', 0.0)), 0.0)
                return math.sqrt(4.0*area/math.pi) if area > 0 else 1.0
            except Exception:
                return 1.0

        def _mcoc_influence_matrix(self, z, n_terms=96):
            """Ma trận hàm ảnh hưởng của phương trình d4y/dz4 + z*y = 0.

            Dùng chuỗi đạo hàm tại z=0 đúng với các hàm A/B/C/D trong hồ sơ MCOC.
            Hàm này chỉ dùng cho alpha*h < 4 nên chuỗi hội tụ nhanh và ổn định.
            """
            ensure_numpy()
            z = float(z)
            F = np.zeros((4, 4), dtype=float)
            for basis in range(4):
                deriv = np.zeros(n_terms, dtype=float)
                deriv[basis] = 1.0
                for n in range(n_terms - 4):
                    deriv[n + 4] = (-n * deriv[n - 1]) if n >= 1 else 0.0
                for r in range(4):
                    total = 0.0
                    term_power = 1.0
                    factorial = 1.0
                    for k, n in enumerate(range(r, n_terms)):
                        if k > 0:
                            term_power *= z
                            factorial *= k
                        total += deriv[n] * term_power / factorial
                    F[r, basis] = total
            return F

        def _finite_depth_ground_stiffness(self, alpha, EI, h, Co, Ct, Fo, Io):
            """Độ cứng [Q,M] tại mặt đất cho cọc hữu hạn theo công thức MCOC.

            Có xét điều kiện biên tại mũi bằng Co*Io và Ct*Fo. Kết quả dùng quy ước
            chuyển vị cục bộ [u, theta] và nội lực [Q, M] của ma trận A3 trong tài liệu.
            """
            ensure_numpy()
            alpha = float(alpha)
            EI = float(EI)
            h = float(h)
            if alpha <= 0 or EI <= 0 or h <= 0:
                raise ValueError("alpha, EI và chiều sâu h phải > 0 khi tính cọc hữu hạn")
            z = alpha*h
            F = self._mcoc_influence_matrix(z)
            D = np.diag([1.0, alpha, -EI*alpha**2, -EI*alpha**3])
            T = D @ F @ np.linalg.inv(D)
            T11, T12 = T[:2, :2], T[:2, 2:]
            T21, T22 = T[2:, :2], T[2:, 2:]

            # Trạng thái vật lý: [y, dy/dx, M, Q]. Điều kiện mũi theo FF của tài liệu.
            S = np.array([[0.0, float(Co)*float(Io)],
                          [-float(Ct)*float(Fo), 0.0]], dtype=float)
            lhs = T22 - S @ T12
            rhs = S @ T11 - T21
            f0 = np.linalg.solve(lhs, rhs)  # [M,Q] theo [y,dy/dx]

            # Đổi sang quy ước MCOC [Q,M] theo [u,theta], theta=-dy/dx.
            K = np.array([[-f0[1, 0], f0[1, 1]],
                          [-f0[0, 0], f0[0, 1]]], dtype=float)
            K = 0.5*(K + K.T)
            eig = np.linalg.eigvalsh(K)
            if not np.all(np.isfinite(K)) or eig[0] <= 0:
                raise ValueError(
                    f"Độ cứng cọc hữu hạn không xác định dương (alpha*h={z:.4f}). "
                    "Kiểm tra Co, Ct, Fo, Io, m, Bp và EI."
                )
            return K

        def get_pile_stiffness(self, alpha, EI, Lo, h, Co, Ct, Fo, Io):
            """Trả rho2, -rho3, rho4 tại đầu cọc theo hồ sơ MCOC."""
            ensure_numpy()
            alpha = max(float(alpha), 1e-12)
            EI = max(float(EI), 1e-12)
            Lo = max(float(Lo), 0.0)
            h = max(float(h), 0.0)
            le = alpha*h

            if le >= 4.0:
                d_HH = self._MCOC_LONG_A0/(alpha**3*EI)
                d_HM = self._MCOC_LONG_B0/(alpha**2*EI)
                d_MM = self._MCOC_LONG_C0/(alpha*EI)
            else:
                K_ground = self._finite_depth_ground_stiffness(alpha, EI, h, Co, Ct, Fo, Io)
                D_ground = np.linalg.inv(K_ground)
                d_HH = float(D_ground[0, 0])
                d_HM = float(0.5*(D_ground[0, 1] + D_ground[1, 0]))
                d_MM = float(D_ground[1, 1])

            # Đoạn tự do Lo theo ma trận A*delta*A + B của tài liệu.
            D_HH = d_HH + 2.0*d_HM*Lo + d_MM*Lo**2 + Lo**3/(3.0*EI)
            D_HM = d_HM + d_MM*Lo + Lo**2/(2.0*EI)
            D_MM = d_MM + Lo/EI
            D = np.array([[D_HH, D_HM], [D_HM, D_MM]], dtype=float)
            K = np.linalg.inv(D)
            K = 0.5*(K + K.T)
            return float(K[0, 0]), float(K[0, 1]), float(K[1, 1])

        def get_group_efficiency(self):
            ensure_numpy()
            eff = {}
            for pi in self.data['Piles']: eff[pi['Name']] = {'v': 1.0, 'h': 1.0}
            if not self.data.get('Config', {}).get('Group_Effect_Enabled', False): return eff
                
            method = self.data.get('Config', {}).get('Group_Method', "Hệ số Poulos")
            N = len(self.data['Piles'])
            if N <= 1: return eff
            
            d_eq_sum = 0
            for pi in self.data['Piles']:
                deq = self._equivalent_diameter(pi)
                d_eq_sum += deq
            d_eq = d_eq_sum / N

            if method == "Hệ số Converse-Labarre":
                S_min = 99999.0
                for i, p1 in enumerate(self.data['Piles']):
                    for j, p2 in enumerate(self.data['Piles']):
                        if i != j:
                            s = math.hypot(p1['X']-p2['X'], p1['Y']-p2['Y'])
                            if s < S_min: S_min = s
                if S_min < 1e-3: S_min = d_eq
                x_coords = set(round(p['X'], 1) for p in self.data['Piles'])
                y_coords = set(round(p['Y'], 1) for p in self.data['Piles'])
                m, n = max(len(x_coords), 1), max(len(y_coords), 1)
                if abs(m * n - N) > N * 0.5:
                    m = int(round(math.sqrt(N)))
                    n = math.ceil(N / m) if m > 0 else 1
                theta = math.degrees(math.atan(d_eq / S_min))
                eta_cl = 1.0 - (theta / 90.0) * ((n-1)*m + (m-1)*n) / (m * n)
                eta_cl = max(0.2, min(1.0, eta_cl))
                for k in eff: eff[k]['v'] = eff[k]['h'] = eta_cl
                    
            elif method == "Móng khối tương đương":
                X_vals = [p['X'] for p in self.data['Piles']]; Y_vals = [p['Y'] for p in self.data['Piles']]
                Bx, By = max(X_vals) - min(X_vals) + d_eq, max(Y_vals) - min(Y_vals) + d_eq
                Ug, Up = 2 * (Bx + By), N * math.pi * d_eq
                eta_raft = min(1.0, Ug / Up)
                for k in eff: eff[k]['v'] = eff[k]['h'] = eta_raft

            elif method == "Hệ số Poulos":
                s_d_v, alpha_v = [1.0, 2.0, 3.0, 4.0, 5.0, 10.0, 50.0], [1.0, 0.60, 0.40, 0.30, 0.20, 0.10, 0.0]
                s_d_h, alpha_h = [1.0, 2.0, 3.0, 4.0, 5.0, 10.0, 50.0], [1.0, 0.60, 0.40, 0.25, 0.15, 0.00, 0.0]
                for i, pi in enumerate(self.data['Piles']):
                    sum_a_v, sum_a_h = 0.0, 0.0
                    xi, yi = pi['X'], pi['Y']
                    deq_i = self._equivalent_diameter(pi)
                        
                    for j, pj in enumerate(self.data['Piles']):
                        if i == j: continue
                        s = math.hypot(xi-pj['X'], yi-pj['Y'])
                        sum_a_v += np.interp(s / deq_i, s_d_v, alpha_v)
                        sum_a_h += np.interp(s / deq_i, s_d_h, alpha_h)
                    eff[pi['Name']]['v'] = 1.0 / (1.0 + sum_a_v)
                    eff[pi['Name']]['h'] = 1.0 / (1.0 + sum_a_h)
            return eff

        def get_shadow_effs(self, combo):
            ensure_numpy()
            eff = {p['Name']: 1.0 for p in self.data['Piles']}
            Hx, Hy = combo['Hx'], combo['Hy']
            H_mag = math.hypot(Hx, Hy)
            if H_mag < 1e-3: return eff
            
            vx, vy = Hx / H_mag, Hy / H_mag
            d_eq_sum, N = 0, len(self.data['Piles'])
            for pi in self.data['Piles']:
                deq = self._equivalent_diameter(pi)
                d_eq_sum += deq
            B = d_eq_sum / N if N > 0 else 1.0
            
            projs = [{'Name': p['Name'], 'loc': p['X'] * vx + p['Y'] * vy} for p in self.data['Piles']]
            projs.sort(key=lambda item: item['loc'], reverse=True)
            
            rows, current_row = [], []
            if not projs: return eff
            current_row_start = projs[0]['loc']
            
            for item in projs:
                if current_row_start - item['loc'] <= 1.5 * B: current_row.append(item)
                else:
                    rows.append(current_row); current_row = [item]; current_row_start = item['loc']
            if current_row: rows.append(current_row)
                
            S_avg, S_sum, count = 3.0 * B, 0, 0
            if len(rows) > 1:
                for i in range(len(rows)-1):
                    l1 = sum(item['loc'] for item in rows[i]) / len(rows[i])
                    l2 = sum(item['loc'] for item in rows[i+1]) / len(rows[i+1])
                    S_sum += abs(l1 - l2); count += 1
                if count > 0: S_avg = S_sum / count
                
            SB_ratio = max(3.0, min(5.0, S_avg / B))
            pm_r1 = np.interp(SB_ratio, [3.0, 5.0], [0.8, 1.0])
            pm_r2 = np.interp(SB_ratio, [3.0, 5.0], [0.4, 0.85])
            pm_r3 = np.interp(SB_ratio, [3.0, 5.0], [0.3, 0.7])
            
            for row_idx, row in enumerate(rows):
                if row_idx == 0: pm = pm_r1
                elif row_idx == 1: pm = pm_r2
                else: pm = pm_r3
                for item in row: eff[item['Name']] = pm
            return eff

        def _pile_properties(self, p, g, eff_v=1.0, eff_h=1.0, shadow_pm=1.0):
            Lo = float(p.get('Lo', 0.0) or 0.0)
            h = float(p.get('H', 0.0) or 0.0)
            if Lo < 0 or h <= 0:
                raise ValueError(f"Cọc {p['Name']}: Lo phải >= 0 và H phải > 0")

            Fo = float(p.get('Area', 0.0) or 0.0)
            Io = float(p.get('J_xy', 0.0) or 0.0)
            if Fo <= 0 or Io <= 0:
                raise ValueError(f"Cọc {p['Name']}: Fo và Io tại mũi cọc phải > 0")

            sec = _mcoc_standard_section_components(0, p)
            if sec['F_total'] <= 0 or sec['Jx_total'] <= 0 or sec['Jy_total'] <= 0:
                raise ValueError(f"Cọc {p['Name']}: F/Jx/Jy thân cọc không hợp lệ")

            Ev_uon = float(g.get('EI_uon', 0.0) or 0.0)
            Er_uon = float(g.get('Er_uon', 0.0) or 0.0)
            Ev_nen = float(g.get('EA_nen', 0.0) or 0.0)
            Er_nen = float(g.get('Er_nen', 0.0) or 0.0)
            if Ev_uon <= 0 or Ev_nen <= 0:
                raise ValueError("Ev.uốn và Ev.nén phải > 0")
            EF = Ev_nen*sec['F_shell'] + Er_nen*sec['F_core']
            EIx = Ev_uon*sec['Jx_shell'] + Er_uon*sec['Jx_core']
            EIy = Ev_uon*sec['Jy_shell'] + Er_uon*sec['Jy_core']
            if EF <= 0 or EIx <= 0 or EIy <= 0:
                raise ValueError(f"Cọc {p['Name']}: EF/EIx/EIy tính ra <= 0")

            Kn_raw = float(g.get('Kn', 0.0) or 0.0)
            if abs(Kn_raw) < 1e-8:
                Kn = 0
            elif abs(Kn_raw - 1.0) < 1e-8:
                Kn = 1
            else:
                raise ValueError(f"Hệ số Kn phải bằng 0 hoặc 1, hiện tại = {Kn_raw}")

            Po = float(p.get('Po', 0.0) or 0.0)
            Co = float(p.get('Co', 0.0) or 0.0)
            Ct = float(p.get('Ct', 0.0) or 0.0)
            if Kn == 0:
                if Po <= 0:
                    raise ValueError(f"Cọc {p['Name']}: Kn=0 nhưng Po<=0")
                ln = Lo + 0.007*EF/Po
            else:
                if Co <= 0:
                    raise ValueError(f"Cọc {p['Name']}: Kn=1 nhưng Co<=0")
                ln = Lo + h + EF/(Co*Fo)
            if ln <= 0:
                raise ValueError(f"Cọc {p['Name']}: chiều dài chịu nén quy đổi ln<=0")
            rho1 = (EF/ln)*float(eff_v)

            m_soil = float(g.get('m', 0.0) or 0.0)
            Bpx = float(p.get('Bpx', 0.0) or 0.0)
            Bpy = float(p.get('Bpy', 0.0) or 0.0)
            if m_soil <= 0 or Bpx <= 0 or Bpy <= 0:
                raise ValueError(f"Cọc {p['Name']}: m, Bpx và Bpy phải > 0")
            alpha_x = (m_soil*Bpx/EIx)**0.2
            alpha_y = (m_soil*Bpy/EIy)**0.2
            le_x = alpha_x*h
            le_y = alpha_y*h

            rho2x, minus_rho3x, rho4x = self.get_pile_stiffness(alpha_x, EIx, Lo, h, Co, Ct, Fo, Io)
            rho2y, minus_rho3y, rho4y = self.get_pile_stiffness(alpha_y, EIy, Lo, h, Co, Ct, Fo, Io)
            rho3x = -minus_rho3x
            rho3y = -minus_rho3y

            scale_h = float(eff_h)*float(shadow_pm)
            rho2x *= scale_h; rho3x *= scale_h; rho4x *= scale_h
            rho2y *= scale_h; rho3y *= scale_h; rho4y *= scale_h
            rho5 = 0.1*(rho4x + rho4y)

            if min(le_x, le_y) < 4.0:
                axis = 'x' if le_x <= le_y else 'y'
                self._add_stiffness_warning(
                    f"Cọc {p['Name']}: H={h:.3f} m; alpha_x*H={le_x:.3f}, "
                    f"alpha_y*H={le_y:.3f}; phương {axis} < 4. "
                    "Đã tính theo nghiệm cọc hữu hạn của hồ sơ MCOC và xét Co, Ct tại mũi; "
                    "giá trị 4 là chiều sâu tính đổi, không phải chiều dài 4 m."
                )

            A3 = np.array([
                [rho1, 0.0, 0.0, 0.0, 0.0, 0.0],
                [0.0, rho2x, 0.0, 0.0, 0.0, -rho3x],
                [0.0, 0.0, rho2y, 0.0, rho3y, 0.0],
                [0.0, 0.0, 0.0, rho5, 0.0, 0.0],
                [0.0, 0.0, rho3y, 0.0, rho4y, 0.0],
                [0.0, -rho3x, 0.0, 0.0, 0.0, rho4x],
            ], dtype=float)

            phi = math.radians(float(p.get('Phi', 0.0) or 0.0))
            psi = math.radians(float(p.get('Xi', 0.0) or 0.0))
            sp, cp = math.sin(phi), math.cos(phi)
            ss, cs = math.sin(psi), math.cos(psi)
            R = np.array([
                [sp*cs, sp*ss, cp],
                [-cp*cs, -cp*ss, sp],
                [ss, -cs, 0.0],
            ], dtype=float)
            A2 = np.zeros((6, 6), dtype=float)
            A2[:3, :3] = R
            A2[3:, 3:] = R

            x = float(p.get('X', 0.0) or 0.0)
            y = float(p.get('Y', 0.0) or 0.0)
            Abar = np.array([[0.0, 0.0, y],
                             [0.0, 0.0, -x],
                             [-y, x, 0.0]], dtype=float)
            A1 = np.block([[np.eye(3), Abar],
                           [np.zeros((3, 3)), np.eye(3)]])
            T = A2 @ A1

            return {
                'x': x, 'y': y, 'A1': A1, 'A2': A2, 'A3': A3, 'T': T,
                'rho1': rho1, 'rho2x': rho2x, 'rho3x': rho3x, 'rho4x': rho4x,
                'rho2y': rho2y, 'rho3y': rho3y, 'rho4y': rho4y, 'rho5': rho5,
                'alpha_x': alpha_x, 'alpha_y': alpha_y, 'le_x': le_x, 'le_y': le_y,
                'EF': EF, 'EIx': EIx, 'EIy': EIy,
            }

        def _build_K_matrix(self, g, efficiencies, shadow_effs):
            ensure_numpy()
            K = np.zeros((6, 6))

            # Ma trận bệ cùng làm việc Γrp theo hồ sơ MCOC gốc.
            # Ký hiệu trong tài liệu:
            #   an, bn, hn : kích thước bệ theo X, Y, Z  -> Ax, By, Cz
            #   mo         : hệ số tỷ lệ nền quanh bệ     -> mbe (trong code cũ là mq)
            #   mn         : hệ số tỷ lệ nền tại đáy bệ   -> md
            #   Cn = mn.hn
            # Theo sơ đồ khối và benchmark chương trình MCOC: khi m_bệ/mbe (Mq) = 0,
            # MCOC bỏ toàn bộ Γrp, kể cả khi Md khác 0. Md chỉ được xét khi Mq > 0.
            mbe = g.get('mq', 0.0)  # tên cũ trong parser; bản chất là m_be, không phải mô men Mq
            md = g.get('md', 0.0)
            an = g.get('Bx', 0.0)
            bn = g.get('By', 0.0)
            hn = g.get('Cz', 0.0)
            if mbe > 0 and an > 0 and bn > 0 and hn > 0:
                # Thành phần nền quanh bệ theo chiều cao hn=Cz
                Sbx_Fc = bn * mbe * (hn ** 2) / 2.0
                Sbx_Sc = bn * mbe * (hn ** 3) / 6.0
                Sbx_Ic = bn * mbe * (hn ** 4) / 12.0

                Sby_Fc = an * mbe * (hn ** 2) / 2.0
                Sby_Sc = an * mbe * (hn ** 3) / 6.0
                Sby_Ic = an * mbe * (hn ** 4) / 12.0

                # Thành phần nền đáy bệ
                Cn = md * hn

                # u, v
                K[0, 0] += Sbx_Fc
                K[1, 1] += Sby_Fc

                # liên kết ngang - xoay do nền quanh bệ
                K[0, 4] += Sbx_Sc
                K[4, 0] += Sbx_Sc
                K[1, 3] += -Sby_Sc
                K[3, 1] += -Sby_Sc

                # w và xoay do nền đáy bệ + nền quanh bệ
                K[2, 2] += Cn * an * bn
                K[3, 3] += Sby_Ic + Cn * an * (bn ** 3) / 12.0
                K[4, 4] += Sbx_Ic + Cn * (an ** 3) * bn / 12.0

                # xoắn quanh Z do nền quanh bệ
                K[5, 5] += (Sby_Fc * (an ** 2) + Sbx_Fc * (bn ** 2)) / 12.0
            
            pile_cache = {}
            for p in self.data['Piles']:
                eff_v = efficiencies[p['Name']]['v']
                eff_h = efficiencies[p['Name']]['h']
                pm = shadow_effs[p['Name']]
                pp = self._pile_properties(p, g, eff_v, eff_h, pm)
                pile_cache[p['Name']] = pp
                T = pp['T']
                K += T.T @ pp['A3'] @ T
            K = 0.5*(K + K.T)
            return K, pile_cache

        def solve(self):
            ensure_numpy()
            g = self.data['Global']
            efficiencies = self.get_group_efficiency()
            shadow_enabled = self.data.get('Config', {}).get('Shadow_Effect', False)
            base_shadow = {p['Name']: 1.0 for p in self.data['Piles']}
            self.K, _ = self._build_K_matrix(g, efficiencies, base_shadow)

            displacements, forces, verifications = [], [], []
            for combo in self.data['Load_Combos']:
                shadow_effs = self.get_shadow_effs(combo) if shadow_enabled else base_shadow
                K_combo, pile_cache = self._build_K_matrix(g, efficiencies, shadow_effs)
                P = np.array([
                    combo['Hx'], combo['Hy'], combo['N_load'],
                    combo['Mx'], combo['My'], combo['Mz'],
                ], dtype=float)
                cond = np.linalg.cond(K_combo)
                if not np.isfinite(cond) or cond > 1e12:
                    raise ValueError(
                        f"Ma trận độ cứng điều kiện kém ở tổ hợp {combo['Name']} (cond={cond:.2e}). "
                        "Kiểm tra đơn vị, Co, Ct, m, E, F/J, bố trí và góc cọc."
                    )
                Delta = np.linalg.solve(K_combo, P)
                u, v, w, rx, ry, rz = Delta
                displacements.append({
                    'T.H': combo['Name'], 'X': u, 'Y': v, 'Z': w,
                    'Fix': rx, 'Fiy': ry, 'Fiz': rz,
                })

                P_recalc = K_combo @ Delta
                verifications.append({
                    'Name': combo['Name'], 'Original': P.copy(),
                    'Recalculated': P_recalc, 'Error': P_recalc - P,
                })

                for p in self.data['Piles']:
                    pp = pile_cache[p['Name']]
                    local_force = pp['A3'] @ pp['T'] @ Delta
                    forces.append({
                        'T.C': p['Name'], 'T.H': combo['Name'],
                        'N': float(local_force[0]), 'Q2': float(local_force[1]),
                        'Q3': float(local_force[2]), 'M1': float(local_force[3]),
                        'M2': float(local_force[4]), 'M3': float(local_force[5]),
                    })
            return displacements, forces, verifications, efficiencies

    def compute_equivalent_stiffness(K):
        ensure_numpy()
        labels = ['Kx', 'Ky', 'Kz', 'Kmx', 'Kmy', 'Kmz']
        truc_tiep = {labels[i]: K[i, i] for i in range(6)}
        try:
            F = np.linalg.inv(K)
            tuong_duong = {labels[i]: (1.0 / F[i, i]) if abs(F[i, i]) > 1e-15 else None for i in range(6)}
        except np.linalg.LinAlgError: tuong_duong = {labels[i]: None for i in range(6)}
        return {'labels': labels, 'truc_tiep': truc_tiep, 'tuong_duong': tuong_duong}

    def convert_stiffness_to_rmbridge(eq, ton_to_kN=9.80665):
        td = eq['tuong_duong']
        mapping = {'Cx': td['Kz'], 'Cy': td['Kx'], 'Cz': td['Ky'], 'CMx': td['Kmz'], 'CMy': td['Kmx'], 'CMz': td['Kmy']}
        return {k: (v * ton_to_kN if v is not None else None) for k, v in mapping.items()}

    # ==========================================
    # 2B. LIÊN KẾT DỮ LIỆU NỘI LỰC SANG TS-COL
    # ==========================================
    def export_forces_to_n2d_col_csv(filepath, forces, source_file="", item="", limit_state="CĐ"):
        'Xuất nội lực cọc theo đúng template CSV mà TS-COL đọc được.\n\n        Quy ước truyền dữ liệu:\n        - TS-PILE: N, Q2, Q3 tính bằng tấn; M1, M2, M3 tính bằng tấn.m.\n        - TS-COL: N_tf, Vux_tf, Vuy_tf, Tu_tfm, Mux_tfm, Muy_tfm.\n        - Mapping: Q2->Vux, Q3->Vuy, M1->Tu, M2->Mux, M3->Muy.\n\n        Hàm này là bước cầu nối trước khi tích hợp import trực tiếp lõi TS-COL.\n        '
        fields = ["SourceFile", "Item", "Pile", "Combo", "TTGH", "N_tf", "Vux_tf", "Vuy_tf", "Tu_tfm", "Mux_tfm", "Muy_tfm"]
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            for r in forces or []:
                writer.writerow({
                    "SourceFile": source_file,
                    "Item": item,
                    "Pile": r.get("T.C", ""),
                    "Combo": r.get("T.H", ""),
                    "TTGH": limit_state,
                    "N_tf": f"{float(r.get('N', 0.0) or 0.0):.6g}",
                    "Vux_tf": f"{float(r.get('Q2', 0.0) or 0.0):.6g}",
                    "Vuy_tf": f"{float(r.get('Q3', 0.0) or 0.0):.6g}",
                    "Tu_tfm": f"{float(r.get('M1', 0.0) or 0.0):.6g}",
                    "Mux_tfm": f"{float(r.get('M2', 0.0) or 0.0):.6g}",
                    "Muy_tfm": f"{float(r.get('M3', 0.0) or 0.0):.6g}",
                })
        return filepath

    # ==========================================
    # 3. CÁC HÀM XUẤT BÁO CÁO 
    # ==========================================
    TTGH_GROUP_DEFINITIONS = (
        ("SD", "TTGH Sử dụng", "TTGHSD"),
        ("CD", "TTGH Cường độ", "TTGHCĐ"),
        ("DB", "TTGH Đặc biệt", "TTGHĐB"),
    )

    def _parse_combo_selection_spec(spec, max_combo=None):
        """Phân tích chuỗi chọn tổ hợp, ví dụ ``1-4,7,9-11``.

        Quy ước:
        - ``0`` hoặc chuỗi rỗng: không chọn tổ hợp nào.
        - Cho phép dấu phẩy, chấm phẩy và khoảng trắng làm dấu phân cách.
        - Không cho phép số âm, khoảng đảo chiều hoặc số vượt quá max_combo.
        """
        raw = str(spec or "").strip()
        if not raw or raw == "0":
            return []
        if re.search(r"(^|[,;\s])0($|[,;\s])", raw):
            raise ValueError("Giá trị 0 chỉ được dùng riêng để bỏ qua cả nhóm TTGH.")
        normalized = raw.replace("–", "-").replace("—", "-").replace("−", "-")
        parts = [p for p in re.split(r"[,;\s]+", normalized) if p]
        selected = set()
        for part in parts:
            if "-" in part:
                bits = part.split("-")
                if len(bits) != 2 or not bits[0].isdigit() or not bits[1].isdigit():
                    raise ValueError(f"Khoảng tổ hợp không hợp lệ: {part!r}")
                start, end = int(bits[0]), int(bits[1])
                if start <= 0 or end <= 0:
                    raise ValueError(f"Số thứ tự tổ hợp phải lớn hơn 0: {part!r}")
                if end < start:
                    raise ValueError(f"Khoảng tổ hợp bị đảo: {part!r}")
                selected.update(range(start, end + 1))
            else:
                if not part.isdigit():
                    raise ValueError(f"Số thứ tự tổ hợp không hợp lệ: {part!r}")
                value = int(part)
                if value <= 0:
                    raise ValueError("Số thứ tự tổ hợp phải lớn hơn 0.")
                selected.add(value)
        values = sorted(selected)
        if max_combo is not None and values and values[-1] > int(max_combo):
            raise ValueError(
                f"Tổ hợp {values[-1]} vượt quá số tổ hợp lớn nhất hiện có ({int(max_combo)})."
            )
        return values

    def _compress_combo_ids(combo_ids):
        values = sorted({int(v) for v in (combo_ids or []) if int(v) > 0})
        if not values:
            return "0"
        parts = []
        start = prev = values[0]
        for value in values[1:]:
            if value == prev + 1:
                prev = value
                continue
            parts.append(str(start) if start == prev else f"{start}-{prev}")
            start = prev = value
        parts.append(str(start) if start == prev else f"{start}-{prev}")
        return ",".join(parts)

    def _build_ttgh_groups(specs, max_combo=None):
        """Chuẩn hóa ba nhóm TTGH và chặn một tổ hợp nằm trong nhiều nhóm."""
        specs = dict(specs or {})
        groups = []
        owner = {}
        for code, label, short_label in TTGH_GROUP_DEFINITIONS:
            raw_spec = str(specs.get(code, "0") or "0").strip()
            combo_ids = _parse_combo_selection_spec(raw_spec, max_combo=max_combo)
            for combo_id in combo_ids:
                if combo_id in owner:
                    raise ValueError(
                        f"Tổ hợp {combo_id} đang được chọn đồng thời cho {owner[combo_id]} và {label}."
                    )
                owner[combo_id] = label
            groups.append({
                "code": code,
                "label": label,
                "short_label": short_label,
                "combo_ids": combo_ids,
                "spec": _compress_combo_ids(combo_ids),
            })
        if not any(group["combo_ids"] for group in groups):
            raise ValueError("Chưa chọn tổ hợp nào. Nhập ít nhất một nhóm khác 0.")
        return groups

    def _combo_index(value):
        try:
            return int(round(float(str(value).strip())))
        except Exception:
            return None

    def _filter_forces_by_combo_ids(forces, combo_ids):
        selected = {int(v) for v in (combo_ids or [])}
        if not selected:
            return []
        return [row for row in (forces or []) if _combo_index(row.get("T.H")) in selected]

    def _force_summary_items(forces):
        """Trả về các dòng cực trị theo đúng bảng tổng hợp hiện hữu của TS-PILE."""
        ensure_pandas()
        df_forces = pd.DataFrame(forces or [])
        required = {"N", "Q2", "Q3", "M1", "M2", "M3", "T.C", "T.H"}
        if df_forces.empty or not required.issubset(set(df_forces.columns)):
            return []
        def get_max_abs(df, col):
            return df.loc[df[col].abs().idxmax()]
        return [
            ("Nmin", df_forces.loc[df_forces['N'].idxmin()]),
            ("Nmax", df_forces.loc[df_forces['N'].idxmax()]),
            ("Q2max", get_max_abs(df_forces, 'Q2')),
            ("Q3max", get_max_abs(df_forces, 'Q3')),
            ("M2max", get_max_abs(df_forces, 'M2')),
            ("M3max", get_max_abs(df_forces, 'M3')),
        ]

    def _iter_ttgh_force_groups(forces, report_config):
        if not report_config or not report_config.get("p7_ttgh", False):
            return []
        output = []
        for group in report_config.get("ttgh_groups", []) or []:
            combo_ids = list(group.get("combo_ids", []) or [])
            subset = _filter_forces_by_combo_ids(forces, combo_ids)
            if subset:
                actual_ids = sorted({_combo_index(row.get("T.H")) for row in subset})
                actual_ids = [v for v in actual_ids if v is not None]
                resolved = dict(group)
                resolved["actual_combo_ids"] = actual_ids
                resolved["actual_spec"] = _compress_combo_ids(actual_ids)
                output.append((resolved, subset))
        return output

    def _write_equivalent_stiffness_block(f, eq, show_direct=False, pw=70):
        labels = eq['labels']; tt = eq['truc_tiep']; td = eq['tuong_duong']
        f.write("\n")
        f.write("ĐỘ CỨNG TƯƠNG ĐƯƠNG CỦA MÓNG CỌC".center(pw) + "\n\n")
        
        if show_direct:
            f.write("  (1) Độ cứng trực tiếp (Đường chéo ma trận K - CHƯA khử liên kết - THAM KHẢO):\n\n")
            f.write("   " + "".join([f"{lb:>11}" for lb in labels]) + "\n")
            f.write("   " + "".join([f"{tt[lb]:11.0f}" for lb in labels]) + "\n\n")
            f.write("  (2) Độ cứng tương đương ĐÃ KHỬ LIÊN KẾT (NÊN DÙNG):\n")
        else:
            f.write("  Độ cứng tương đương ĐÃ KHỬ LIÊN KẾT (P = Keq * Delta):\n\n")

        f.write("   " + "".join([f"{lb:>11}" for lb in labels]) + "\n")
        f.write("   " + "".join([f"{td[lb]:11.0f}" if td[lb] is not None else f"{'N/A':>11}" for lb in labels]) + "\n")
        f.write("\nĐơn vị: Kx[t/m], Ky[t/m], Kz[t/m], Kmx[t.m/rad], Kmy[t.m/rad], Kmz[t.m/rad]\n\n")

    def _write_rmbridge_block(f, rm, pw=70):
        labels = ['Cx', 'Cy', 'Cz', 'CMx', 'CMy', 'CMz']
        f.write("\n")
        f.write("ĐỘ CỨNG TƯƠNG ĐƯƠNG QUY ĐỔI CHO RM BRIDGE (ĐỔI ĐƠN VỊ KN)".center(pw) + "\n\n")
        f.write("   " + "".join([f"{lb:>11}" for lb in labels]) + "\n")
        f.write("   " + "".join([f"{rm[lb]:11.0f}" if rm[lb] is not None else f"{'N/A':>11}" for lb in labels]) + "\n")
        f.write("\nĐơn vị: Cx[kN/m], Cy[kN/m], Cz[kN/m], CMx[kN.m/rad], CMy[kN.m/rad], CMz[kN.m/rad]\n")
        f.write("* Lưu ý: Lò xo này tính từ ma trận cơ sở, chưa bao gồm P-multiplier (nếu có).\n\n")

    def _section_report_schema(data):
        """Nhãn ba trường hình học mặt cắt trong bảng thông số cọc."""
        return ['A', 'B', 'Cday'], ''

    def _compact_report_number(value, decimals=3):
        """Định dạng kích thước gọn, không giữ các số 0 thừa."""
        value = float(value or 0.0)
        if abs(value) < 0.5 * 10 ** (-decimals):
            value = 0.0
        return f"{value:.{decimals}f}".rstrip('0').rstrip('.')

    def _pile_report_description(pile):
        """Tên cọc ngắn gọn dùng thống nhất trong TXT, Word/PDF và Excel."""
        sec1, sec2, sec3 = _section_raw_values(pile)
        if abs(sec3) < 1e-5:
            d_outer = _compact_report_number(sec1, 3)
            if sec2 > 1e-5:
                d_inner = _compact_report_number(sec2, 3)
                return f"Cọc ống D={d_outer}m, d={d_inner}m"
            return f"Cọc khoan nhồi D={d_outer}m"

        a_cm = _compact_report_number(sec1 * 100.0, 2)
        b_cm = _compact_report_number(sec2 * 100.0, 2)
        return f"Cọc BTCT {a_cm}x{b_cm}cm"

    def export_to_legacy_doc(data, displacements, forces, verifications, efficiencies, filepath, 
                             equivalent_stiffness=None, rm_bridge_stiffness=None, 
                             company="TẬP ĐOÀN SUNGROUP", department="KHỐI XD PPP - PHÒNG QLTK", author='TSDUNGVN - 2026',
                             report_config=None):
        ensure_pandas()
        if report_config is None:
            report_config = {'p1':True, 'p2':True, 'p3':True, 'p4':True, 'p5':True, 'p6':True, 'p7':True, 'p7_ttgh':False, 'ttgh_groups':[], 'p8':True, 'p8_opt':2, 'p9':True, 'p9_direct':False}

        with open(filepath, 'w', encoding='utf-8') as f:
            # ---------------------------------------------
            # THUẬT TOÁN CĂN LỀ TÊN CÔNG TY & PHÒNG BAN
            # ---------------------------------------------
            page_width = 70
            c_lines = company.split('\n')
            d_lines = department.split('\n')
            
            max_len = 0
            for line in c_lines + d_lines:
                if len(line) > max_len: max_len = len(line)
                    
            for line in c_lines:
                padding = (max_len - len(line)) // 2
                f.write(" " * padding + line + "\n")
            for line in d_lines:
                padding = (max_len - len(line)) // 2
                f.write(" " * padding + line + "\n")
            
            f.write("\n\n")
            f.write("CHƯƠNG TRÌNH TÍNH MÓNG CỌC".center(page_width) + "\n\n\n")
            
            g = data['Global']
            f.write(f"Công trình : {g.get('Project_ID', 'Cau')}".center(page_width) + "\n \n")
            
            if report_config['p1']:
                f.write(f"        Kn = {g['Kn']:5.2f}   Ax = {g['Bx']:6.2f}   By = {g['By']:6.2f}    Cz = {g['Cz']:4.2f}\n")
                f.write(f"  E v.uon = {g['EI_uon']:8.0f}   E r.uon = {g.get('Er_uon', 0):2.0f}   E v.nen  = {g['EA_nen']:8.0f}   E r.nen = {g.get('Er_nen', 0):2.0f} \n")
                f.write(f"         Mq = {g['mq']:3.0f} (t/m4)    Md = {g['md']:3.0f} (t/m4)    m = {g['m']:4.0f} (t/m4)\n \n")

            if report_config['p2']:
                f.write("CÁC TỔ HỢP TẢI TRỌNG TÍNH TOÁN".center(page_width) + "\n\n")
                f.write(f"  {'T.T':^2} {'Hx':^9} {'Hy':^9} {'P':^9} {'Mx':^9} {'My':^9} {'Mz':^10}\n\n")
                for c in data['Load_Combos']: f.write(f"  {c['Name']:>2} {c['Hx']:9.2f} {c['Hy']:9.2f} {c['N_load']:9.2f} {c['Mx']:9.2f} {c['My']:9.2f} {c['Mz']:9.2f}\n")

            if report_config['p3']:
                f.write("\n")
                f.write("THÔNG SỐ CỌC".center(page_width) + "\n\n")
                sec_labels, _ = _section_report_schema(data)
                f.write(f" T.C   Lo  H    Bpx    Bpy  {sec_labels[0]:>5} {sec_labels[1]:>5} {sec_labels[2]:>5}    Fo  Jo   Po  Co   Ct\n\n")
                
                first = data['Piles'][0]
                f.write(f" {first['Name']:<3} {first['Lo']:4.2f} {first['H']:5.2f} {first['Bpx']:5.2f} {first['Bpy']:5.2f} {first['d_ngoai']:5.2f} {first['d_trong']:5.2f} {first['day_vo']:5.2f} {first['Area']:5.2f} {first['J_xy']:5.2f} {first['Po']:4.0f} {first['Co']:6.0f} {first['Ct']:6.0f}\n")
                
                for i in range(1, len(data['Piles'])):
                    p = data['Piles'][i]
                    is_same = (abs(p['Lo']-first['Lo'])<1e-5 and abs(p['H']-first['H'])<1e-5 and 
                               abs(p['Bpx']-first['Bpx'])<1e-5 and abs(p['Bpy']-first['Bpy'])<1e-5 and 
                               abs(p['d_ngoai']-first['d_ngoai'])<1e-5 and abs(p['d_trong']-first['d_trong'])<1e-5 and 
                               abs(p['day_vo']-first['day_vo'])<1e-5 and abs(p['Area']-first['Area'])<1e-5 and 
                               abs(p['J_xy']-first['J_xy'])<1e-5 and abs(p['Po']-first['Po'])<1e-5 and 
                               abs(p['Co']-first['Co'])<1e-5 and abs(p['Ct']-first['Ct'])<1e-5)
                    if is_same:
                        f.write(f" {p['Name']:<3}                                n t  \n")
                    else:
                        f.write(f" {p['Name']:<3} {p['Lo']:4.2f} {p['H']:5.2f} {p['Bpx']:5.2f} {p['Bpy']:5.2f} {p['d_ngoai']:5.2f} {p['d_trong']:5.2f} {p['day_vo']:5.2f} {p['Area']:5.2f} {p['J_xy']:5.2f} {p['Po']:4.0f} {p['Co']:6.0f} {p['Ct']:6.0f}\n")
                f.write("\n")
                
                f.write(_pile_report_description(first).center(page_width) + "\n \n")

            if report_config['p4']:
                f.write("TOẠ ĐỘ ĐẦU CỌC".center(page_width) + "\n\n")
                f.write(f"            {'T.C':^2}  {'X':^8} {'Y':^8} {'Phi':^7} {'Xi':^7}\n\n")
                for p in data['Piles']: f.write(f"            {p['Name']:>2} {p['X']:8.2f} {p['Y']:8.2f} {p['Phi']:7.3f} {p['Xi']:7.2f}\n")

                if data.get('Config', {}).get('Group_Effect_Enabled', False):
                    method = data.get('Config', {}).get('Group_Method', '')
                    f.write("\n")
                    f.write(f"HỆ SỐ SUY GIẢM ĐỘ CỨNG NHÓM (Theo: {method})".center(page_width) + "\n\n")
                    f.write("            T.C     Eff_V(Dọc)   Eff_H(Ngang)\n\n")
                    for p in data['Piles']: f.write(f"            {p['Name']:>2}        {efficiencies[p['Name']]['v']:.3f}        {efficiencies[p['Name']]['h']:.3f}\n")

                if data.get('Config', {}).get('Shadow_Effect', False):
                    f.write("\n")
                    f.write("LƯU Ý: ĐÃ XÉT HIỆU ỨNG BÓNG RÂM THEO TCVN 11823".center(page_width) + "\n")
                    f.write("Độ cứng ngang của cọc thay đổi động theo từng tổ hợp dựa trên hướng lực đẩy.".center(page_width) + "\n\n")

            if report_config['p5']:
                f.write("\n")
                f.write("CHUYỂN VỊ BỆ CỌC".center(page_width) + "\n\n")
                f.write(f"        {'T.H':^2} {'X':^9} {'Y':^9} {'Z':^9} {'Fix':^9} {'Fiy':^9} {'Fiz':^9}\n\n")
                for d in displacements: f.write(f"        {d['T.H']:>2} {d['X']:9.5f} {d['Y']:9.5f} {d['Z']:9.5f} {d['Fix']:9.5f} {d['Fiy']:9.5f} {d['Fiz']:9.5f}\n")

            if report_config['p6']:
                f.write("\n")
                f.write("NỘI LỰC ĐẦU CỌC".center(page_width) + "\n\n")
                f.write(f"  {'T.C':<1}  {'T.H':<1}  {'N':^6}  {'Q2':^7}  {'Q3':^8.5} {'M1':^9.5} {'M2':^10.5} {'M3':^11}\n\n")
                df_forces = pd.DataFrame(forces)
                for pile_name, group in df_forces.groupby('T.C', sort=False):
                    first_row = True
                    for _, row in group.iterrows():
                        p_str = f"  {pile_name:>2}" if first_row else "    "
                        f.write(f"{p_str}  {str(row['T.H']):>2} {row['N']:8.2f} {row['Q2']:8.2f} {row['Q3']:8.2f} {row['M1']:9.3f} {row['M2']:10.3f} {row['M3']:10.3f}\n")
                        first_row = False

            if report_config['p7']:
                f.write("\n")
                f.write("BẢNG TỔNG KẾT NỘI LỰC".center(page_width) + "\n\n")
                f.write(f"{'Loại':<5} {'T.C':^2}  {'T.H':^2} {'N':^8} {'Q2':^8} {'Q3':^8} {'M1':^9} {'M2':^10} {'M3':^10}\n\n")
                for name, r in _force_summary_items(forces):
                    f.write(f"{name:<5} {str(r['T.C']):>2}  {str(r['T.H']):>2} {r['N']:8.2f} {r['Q2']:8.2f} {r['Q3']:8.2f} {r['M1']:9.3f} {r['M2']:10.3f} {r['M3']:10.3f}\n")

            ttgh_groups = _iter_ttgh_force_groups(forces, report_config)
            if ttgh_groups:
                f.write("\n")
                f.write("TỔNG HỢP NỘI LỰC THEO TRẠNG THÁI GIỚI HẠN".center(page_width) + "\n")
                for group, subset in ttgh_groups:
                    f.write("\n")
                    heading = f"{group['label']} ({group['short_label']}) - Tổ hợp {group['actual_spec']}"
                    f.write(heading.center(page_width) + "\n\n")
                    f.write(f"{'Loại':<5} {'T.C':^2}  {'T.H':^2} {'N':^8} {'Q2':^8} {'Q3':^8} {'M1':^9} {'M2':^10} {'M3':^10}\n\n")
                    for name, r in _force_summary_items(subset):
                        f.write(f"{name:<5} {str(r['T.C']):>2}  {str(r['T.H']):>2} {r['N']:8.2f} {r['Q2']:8.2f} {r['Q3']:8.2f} {r['M1']:9.3f} {r['M2']:10.3f} {r['M3']:10.3f}\n")

            if report_config['p8']:
                f.write("\n")
                f.write("TÍNH TOÁN KIỂM TRA".center(page_width) + "\n")
                f.write("SO SÁNH VỚI MA TRẬN TẢI TRỌNG BAN ĐẦU".center(page_width) + "\n\n")
                
                if report_config['p8_opt'] == 1:
                    f.write(f"  {'T.T':^2} {'Hx':^9} {'Hy':^9} {'P':^9} {'Mx':^9} {'My':^9} {'Mz':^10}\n\n")
                    for v in verifications:
                        f.write(f"  {v['Name']:>2} {v['Recalculated'][0]:9.2f} {v['Recalculated'][1]:9.2f} {v['Recalculated'][2]:9.2f} {v['Recalculated'][3]:9.2f} {v['Recalculated'][4]:9.2f} {v['Recalculated'][5]:9.2f}\n")
                else:
                    labels = ['Hx', 'Hy', 'P', 'Mx', 'My', 'Mz']
                    for v in verifications:
                        f.write(f"  [Tổ hợp {v['Name']}]\n            " + "".join([f"{lb:>10}" for lb in labels]) + "\n")
                        f.write(" Ban đầu: " + "".join([f"{v['Original'][i]:10.0f}" for i in range(6)]) + "\n Ngược:   " + "".join([f"{v['Recalculated'][i]:10.0f}" for i in range(6)]) + "\n Sai số:  " + "".join([f"{'OK':>10}" if abs(v['Error'][i]) < 0.001 else f"{v['Error'][i]:10.4f}" for i in range(6)]) + "\n\n")

            if report_config['p9'] and equivalent_stiffness is not None: 
                _write_equivalent_stiffness_block(f, equivalent_stiffness, report_config.get('p9_direct', False), pw=page_width)
                if rm_bridge_stiffness is not None: _write_rmbridge_block(f, rm_bridge_stiffness, pw=page_width)

            # ---------------------------------------------
            # THUẬT TOÁN CĂN LỀ: TÊN NGƯỜI LẬP NẰM TRÊN 1 DÒNG VÀ Ở GÓC PHẢI
            # ---------------------------------------------
            safe_author = author.replace('\n', ' ').replace('\r', '').strip()
            f.write("\n\n")
            f.write(safe_author.rjust(page_width) + "\n")
            f.write(("_" * min(len(safe_author), 30)).rjust(page_width) + "\n")


    # ==========================================
    # 3B. FORM BÁO CÁO MỚI - DỌC, THEO STYLE TS-COL
    # ==========================================
    def _mcoc_fmt(value, nd=2):
        try:
            if value is None:
                return ""
            if isinstance(value, str):
                return value
            v = float(value)
            if abs(v) >= 100000:
                return f"{v:,.0f}"
            if abs(v) >= 1000:
                return f"{v:,.1f}"
            return f"{v:,.{nd}f}"
        except Exception:
            return str(value)

    def _mcoc_html_multiline(text):
        return html.escape(str(text if text is not None else "")).replace("\n", "<br/>")

    def _mcoc_html_pre(text):
        return html.escape(str(text if text is not None else "")).replace(" ", "&nbsp;").replace("\n", "<br/>")

    def _mcoc_center_header_lines(company="", department=""):
        """Căn giữa tên công ty/phòng ban theo đúng thuật toán form classic: dòng ngắn hơn thêm 1/2 hiệu số ký tự ở đầu."""
        lines = []
        for block in (company, department):
            for line in str(block or "").split('\n'):
                s = line.strip()
                if s:
                    lines.append(s)
        if not lines:
            return []
        max_len = max(len(s) for s in lines)
        return [(" " * ((max_len - len(s)) // 2)) + s for s in lines]

    def _mcoc_author_lines(author=""):
        lines = [line.strip() for line in str(author or "").replace('\r', '').split('\n') if line.strip()]
        if not lines:
            lines = ["Người thực hiện"]
        return lines

    def _mcoc_author_underline(author=""):
        lines = _mcoc_author_lines(author)
        return "_" * max(len(line) for line in lines)

    def _mcoc_is_nt_row(row):
        try:
            if len(row) < 2:
                return False
            marker = str(row[1] or "").strip().lower().replace(" ", "")
            return marker in ("nt", "nhưtrên", "nhutrên", "nhutren") and all(str(x or "").strip() == "" for x in row[2:])
        except Exception:
            return False

    def _mcoc_summary_force_rows(forces):
        rows = []
        for name, r in _force_summary_items(forces):
            rows.append([name, r['T.C'], r['T.H'], _mcoc_fmt(r['N']), _mcoc_fmt(r['Q2']), _mcoc_fmt(r['Q3']), _mcoc_fmt(r['M1'], 3), _mcoc_fmt(r['M2'], 3), _mcoc_fmt(r['M3'], 3)])
        return rows

    def _mcoc_report_sections(data, displacements, forces, verifications, efficiencies,
                              equivalent_stiffness=None, rm_bridge_stiffness=None, report_config=None):
        ensure_pandas()
        if report_config is None:
            report_config = {'p1': True, 'p2': True, 'p3': True, 'p4': True, 'p5': True, 'p6': True, 'p7': True, 'p7_ttgh': False, 'ttgh_groups': [], 'p8': True, 'p8_opt': 1, 'p9': True, 'p9_direct': False}
        g = data['Global']
        sections = []

        if report_config.get('p1', True):
            rows = [
                [f". Hệ số tỉ lệ nền Kn = {_mcoc_fmt(g.get('Kn'))}"],
                [f". Kích thước bệ: Ax={_mcoc_fmt(g.get('Bx'))} (m); By={_mcoc_fmt(g.get('By'))} (m); Cz={_mcoc_fmt(g.get('Cz'))} (m)"],
                [". Modul đàn hồi vật liệu cọc:"],
                [f"+ Ev uốn = {_mcoc_fmt(g.get('EI_uon'), 0)} (T/m2);"],
                [f"+ Er uốn = {_mcoc_fmt(g.get('Er_uon', 0), 0)} (T/m2);"],
                [f"+ Ev nén = {_mcoc_fmt(g.get('EA_nen'), 0)} (T/m2);"],
                [f"+ Er nén = {_mcoc_fmt(g.get('Er_nen', 0), 0)} (T/m2)"],
                [f". Hệ số nền quanh bệ mq={_mcoc_fmt(g.get('mq'), 0)} (T/m4)"],
                [f". Hệ số nền đáy bệ md={_mcoc_fmt(g.get('md'), 0)} (T/m4)"],
                [f". Hệ số nền khu vực cọc làm việc: m={_mcoc_fmt(g.get('m'), 0)} (T/m4)"],
            ]
            sections.append(("1. Thông tin cơ bản", [], rows, ""))

        if report_config.get('p2', True):
            rows = [[c['Name'], _mcoc_fmt(c['Hx']), _mcoc_fmt(c['Hy']), _mcoc_fmt(c['N_load']), _mcoc_fmt(c['Mx']), _mcoc_fmt(c['My']), _mcoc_fmt(c['Mz'])] for c in data['Load_Combos']]
            sections.append(("2. Tổ hợp tải trọng", ["T.H", "Hx", "Hy", "P", "Mx", "My", "Mz"], rows, "Đơn vị: lực tấn, mô men tấn.m."))

        if report_config.get('p3', True):
            rows = []
            first = data['Piles'][0]
            sec_labels, _ = _section_report_schema(data)
            headers = ["T.C", "Lo", "H", "Bpx", "Bpy", sec_labels[0], sec_labels[1], sec_labels[2], "Fo", "Jo", "Po", "Co", "Ct"]
            for i, p in enumerate(data['Piles']):
                if i > 0:
                    is_same = (abs(p['Lo']-first['Lo'])<1e-5 and abs(p['H']-first['H'])<1e-5 and
                               abs(p['Bpx']-first['Bpx'])<1e-5 and abs(p['Bpy']-first['Bpy'])<1e-5 and
                               abs(p['d_ngoai']-first['d_ngoai'])<1e-5 and abs(p['d_trong']-first['d_trong'])<1e-5 and
                               abs(p['day_vo']-first['day_vo'])<1e-5 and abs(p['Area']-first['Area'])<1e-5 and
                               abs(p['J_xy']-first['J_xy'])<1e-5 and abs(p['Po']-first['Po'])<1e-5 and
                               abs(p['Co']-first['Co'])<1e-5 and abs(p['Ct']-first['Ct'])<1e-5)
                    if is_same:
                        rows.append([p['Name'], "nt", "", "", "", "", "", "", "", "", "", "", ""])
                        continue
                sec1, sec2, sec3 = _section_raw_values(p)
                rows.append([p['Name'], _mcoc_fmt(p['Lo']), _mcoc_fmt(p['H']), _mcoc_fmt(p['Bpx']), _mcoc_fmt(p['Bpy']), _mcoc_fmt(sec1), _mcoc_fmt(sec2), _mcoc_fmt(sec3), _mcoc_fmt(p['Area']), _mcoc_fmt(p['J_xy']), _mcoc_fmt(p['Po'], 0), _mcoc_fmt(p['Co'], 0), _mcoc_fmt(p['Ct'], 0)])
            note = _pile_report_description(first)
            sections.append(("3. Thông số cọc", headers, rows, note))

        if report_config.get('p4', True):
            rows = [[p['Name'], _mcoc_fmt(p['X']), _mcoc_fmt(p['Y']), _mcoc_fmt(p['Phi'], 3), _mcoc_fmt(p['Xi'])] for p in data['Piles']]
            sections.append(("4. Tọa độ đầu cọc", ["T.C", "X", "Y", "Phi", "Xi"], rows, ""))
            if data.get('Config', {}).get('Group_Effect_Enabled', False):
                method = data.get('Config', {}).get('Group_Method', '')
                rows = [[p['Name'], _mcoc_fmt(efficiencies[p['Name']]['v'], 3), _mcoc_fmt(efficiencies[p['Name']]['h'], 3)] for p in data['Piles']]
                sections.append(("4A. Hệ số suy giảm độ cứng nhóm", ["T.C", "Eff_V", "Eff_H"], rows, f"Phương pháp: {method}"))
            if data.get('Config', {}).get('Shadow_Effect', False):
                sections.append(("4B. Ghi chú hiệu ứng bóng râm", [], [["Đã xét hiệu ứng bóng râm theo hướng lực ngang từng tổ hợp."]], ""))

        if report_config.get('p5', True):
            rows = [[d['T.H'], _mcoc_fmt(d['X'], 5), _mcoc_fmt(d['Y'], 5), _mcoc_fmt(d['Z'], 5), _mcoc_fmt(d['Fix'], 5), _mcoc_fmt(d['Fiy'], 5), _mcoc_fmt(d['Fiz'], 5)] for d in displacements]
            sections.append(("5. Chuyển vị bệ cọc", ["T.H", "X", "Y", "Z", "Fix", "Fiy", "Fiz"], rows, ""))

        if report_config.get('p6', True):
            rows = []
            df_forces = pd.DataFrame(forces)
            for pile_name, group in df_forces.groupby('T.C', sort=False):
                first_row = True
                for _, row in group.iterrows():
                    rows.append([pile_name if first_row else "", row['T.H'], _mcoc_fmt(row['N']), _mcoc_fmt(row['Q2']), _mcoc_fmt(row['Q3']), _mcoc_fmt(row['M1'], 3), _mcoc_fmt(row['M2'], 3), _mcoc_fmt(row['M3'], 3)])
                    first_row = False
            sections.append(("6. Nội lực đầu cọc", ["T.C", "T.H", "N", "Q2", "Q3", "M1", "M2", "M3"], rows, "Đơn vị: N, Q2, Q3 tính bằng tấn; M1, M2, M3 tính bằng tấn.m."))

        if report_config.get('p7', True):
            sections.append(("7. Bảng tổng kết nội lực", ["Loại", "T.C", "T.H", "N", "Q2", "Q3", "M1", "M2", "M3"], _mcoc_summary_force_rows(forces), ""))

        for section_index, (group, subset) in enumerate(_iter_ttgh_force_groups(forces, report_config), start=1):
            suffix = chr(ord('A') + section_index - 1)
            title = f"7{suffix}. Tổng kết nội lực - {group['label']} ({group['short_label']})"
            note = f"Các tổ hợp được tổng hợp: {group['actual_spec']}."
            sections.append((title, ["Loại", "T.C", "T.H", "N", "Q2", "Q3", "M1", "M2", "M3"], _mcoc_summary_force_rows(subset), note))

        if report_config.get('p8', True):
            if int(report_config.get('p8_opt', 1)) == 1:
                rows = [[v['Name'], _mcoc_fmt(v['Recalculated'][0]), _mcoc_fmt(v['Recalculated'][1]), _mcoc_fmt(v['Recalculated'][2]), _mcoc_fmt(v['Recalculated'][3]), _mcoc_fmt(v['Recalculated'][4]), _mcoc_fmt(v['Recalculated'][5])] for v in verifications]
                sections.append(("8. Tính toán kiểm tra", ["T.H", "Hx", "Hy", "P", "Mx", "My", "Mz"], rows, "Kết quả tính ngược P = K.Δ."))
            else:
                rows = []
                labels = ['Hx', 'Hy', 'P', 'Mx', 'My', 'Mz']
                for v in verifications:
                    rows.append([f"TH {v['Name']}", "Ban đầu"] + [_mcoc_fmt(v['Original'][i]) for i in range(6)])
                    rows.append(["", "Tính ngược"] + [_mcoc_fmt(v['Recalculated'][i]) for i in range(6)])
                    rows.append(["", "Sai số"] + ["OK" if abs(v['Error'][i]) < 0.001 else _mcoc_fmt(v['Error'][i], 4) for i in range(6)])
                sections.append(("8. Tính toán kiểm tra", ["Tổ hợp", "Loại"] + labels, rows, "So sánh với ma trận tải trọng ban đầu."))

        if report_config.get('p9', True) and equivalent_stiffness is not None:
            labels = equivalent_stiffness['labels']; tt = equivalent_stiffness['truc_tiep']; td = equivalent_stiffness['tuong_duong']
            if report_config.get('p9_direct', False):
                sections.append(("9A. Độ cứng trực tiếp", labels, [[_mcoc_fmt(tt[lb], 0) for lb in labels]], "Đường chéo ma trận K, chưa khử liên kết - chỉ dùng tham khảo."))
                title = "9B. Độ cứng tương đương đã khử liên kết"
            else:
                title = "9. Độ cứng tương đương đã khử liên kết"
            sections.append((title, labels, [[_mcoc_fmt(td[lb], 0) if td[lb] is not None else "N/A" for lb in labels]], "Đơn vị: Kx, Ky, Kz [t/m]; Kmx, Kmy, Kmz [t.m/rad]."))
            if rm_bridge_stiffness is not None:
                rm_labels = ['Cx', 'Cy', 'Cz', 'CMx', 'CMy', 'CMz']
                sections.append(("9C. Độ cứng quy đổi cho RM Bridge", rm_labels, [[_mcoc_fmt(rm_bridge_stiffness[lb], 0) if rm_bridge_stiffness[lb] is not None else "N/A" for lb in rm_labels]], "Đơn vị: Cx, Cy, Cz [kN/m]; CMx, CMy, CMz [kN.m/rad]."))
        return sections

    def _mcoc_section_to_html(title, headers, rows, note):
        note_html = f"<div class='note'>{_mcoc_html_multiline(note)}</div>" if note else ""
        if not headers:
            table = "".join(f"<p>{_mcoc_html_multiline(' '.join(map(str, row)))}</p>" for row in rows)
            return f"<h2>{_mcoc_html_multiline(title)}</h2>{note_html}{table}"

        ncols = len(headers)

        def row_attr(values, is_header=False):
            return " class='font-11'" if _new_report_row_needs_11pt(values, ncols, is_header) else ""

        head = "".join(f"<th>{_mcoc_html_multiline(h)}</th>" for h in headers)
        head_row = f"<tr{row_attr(headers, True)}>{head}</tr>"
        body_parts = []

        if str(title).startswith("3. Thông số cọc"):
            for idx, row in enumerate(rows):
                padded = list(row) + [""] * max(0, ncols - len(row))
                attr = row_attr(padded)
                if idx > 0 and _mcoc_is_nt_row(padded):
                    body_parts.append(
                        f"<tr{attr}><td>{_mcoc_html_multiline(padded[0])}</td>"
                        f"<td colspan='{ncols - 1}' class='nt-cell'>nt</td></tr>"
                    )
                else:
                    body_parts.append(
                        f"<tr{attr}>" +
                        "".join(f"<td>{_mcoc_html_multiline(c)}</td>" for c in padded[:ncols]) +
                        "</tr>"
                    )
        elif str(title).startswith("6. Nội lực đầu cọc"):
            i = 0
            while i < len(rows):
                row = list(rows[i]) + [""] * max(0, ncols - len(rows[i]))
                pile_name = str(row[0] or "")
                attr = row_attr(row)
                if pile_name:
                    span = 1
                    j = i + 1
                    while j < len(rows):
                        nxt = list(rows[j]) + [""] * max(0, ncols - len(rows[j]))
                        if str(nxt[0] or "").strip():
                            break
                        span += 1
                        j += 1
                    body_parts.append(
                        f"<tr{attr}>"
                        + f"<td rowspan='{span}'>{_mcoc_html_multiline(pile_name)}</td>"
                        + "".join(f"<td>{_mcoc_html_multiline(c)}</td>" for c in row[1:ncols])
                        + "</tr>"
                    )
                    for k in range(i + 1, j):
                        row2 = list(rows[k]) + [""] * max(0, ncols - len(rows[k]))
                        body_parts.append(
                            f"<tr{row_attr(row2)}>" +
                            "".join(f"<td>{_mcoc_html_multiline(c)}</td>" for c in row2[1:ncols]) +
                            "</tr>"
                        )
                    i = j
                else:
                    body_parts.append(
                        f"<tr{attr}>" +
                        "".join(f"<td>{_mcoc_html_multiline(c)}</td>" for c in row[:ncols]) +
                        "</tr>"
                    )
                    i += 1
        else:
            for row in rows:
                padded = list(row) + [""] * max(0, ncols - len(row))
                body_parts.append(
                    f"<tr{row_attr(padded)}>" +
                    "".join(f"<td>{_mcoc_html_multiline(c)}</td>" for c in padded[:ncols]) +
                    "</tr>"
                )

        col_widths = _new_report_column_widths_pt(headers, rows)
        total_width = sum(col_widths) or 1.0
        colgroup = "<colgroup>" + "".join(
            f"<col style='width:{100.0 * width / total_width:.3f}%'>"
            for width in col_widths
        ) + "</colgroup>"
        table = f"<table>{colgroup}<thead>{head_row}</thead><tbody>{''.join(body_parts)}</tbody></table>"
        return f"<h2>{_mcoc_html_multiline(title)}</h2>{note_html}{table}"

    def export_to_new_doc(data, displacements, forces, verifications, efficiencies, filepath,
                          equivalent_stiffness=None, rm_bridge_stiffness=None,
                          company="TẬP ĐOÀN SUNGROUP", department="KHỐI XD PPP - PHÒNG QLTK", author='TSDUNGVN - 2026',
                          report_config=None):
        sections = _mcoc_report_sections(data, displacements, forces, verifications, efficiencies, equivalent_stiffness, rm_bridge_stiffness, report_config)
        g = data['Global']
        date_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        section_html = [_mcoc_section_to_html(title, headers, rows, note) for title, headers, rows, note in sections]
        header_lines = _mcoc_center_header_lines(company, department)
        header_html = "\n".join(header_lines)
        author_lines = _mcoc_author_lines(author)
        author_html = "<br>".join(f"<u>{_mcoc_html_multiline(line)}</u>" for line in author_lines)
        underline_html = ""
        html_doc = f"""<!doctype html>
<html lang="vi"><head><meta charset="utf-8">
<style>
@page {{ size: A4 portrait; margin: 15mm 12mm; }}
body {{ font-family: "Times New Roman", serif; font-size: {NEW_REPORT_BODY_FONT_PT:g}pt; color: #111; line-height: 1.20; }}
.header {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 10px; }}
.company {{ font-weight: bold; text-transform: uppercase; white-space: pre; }} .date {{ text-align: right; font-size: {NEW_REPORT_BODY_FONT_PT:g}pt; }}
h1 {{ text-align: center; font-size: 18pt; text-transform: uppercase; margin: 12px 0 6px 0; }}
.sub {{ text-align: center; font-size: {NEW_REPORT_BODY_FONT_PT:g}pt; font-style: italic; margin-bottom: 10px; }}
h2 {{ font-size: 13.5pt; text-transform: uppercase; border-bottom: 1px solid #333; padding-bottom: 3px; margin: 13px 0 6px 0; }}
table {{ width: 100%; table-layout: fixed; border-collapse: collapse; margin: 5px 0 9px 0; font-size: {NEW_REPORT_TABLE_FONT_PT:g}pt; }}
tr {{ height: {NEW_REPORT_TABLE_LINE_HEIGHT_PT:g}pt; }}
th, td {{ border: 1px solid #333; padding: 0 {NEW_REPORT_HTML_CELL_HPAD_PX:g}px; vertical-align: middle; text-align: center; font-size: {NEW_REPORT_TABLE_FONT_PT:g}pt; line-height: {NEW_REPORT_TABLE_LINE_HEIGHT_PT:g}pt; mso-line-height-rule: exactly; white-space: nowrap; }}
tr.font-11 {{ height: {NEW_REPORT_TABLE_OVERFLOW_LINE_HEIGHT_PT:g}pt; }}
tr.font-11 th, tr.font-11 td {{ font-size: {NEW_REPORT_TABLE_OVERFLOW_FONT_PT:g}pt; line-height: {NEW_REPORT_TABLE_OVERFLOW_LINE_HEIGHT_PT:g}pt; }}
th {{ background: #e9eef7; font-weight: bold; }} .nt-cell {{ text-align: center; }}
.note {{ margin: 4px 0 6px 0; font-size: {NEW_REPORT_BODY_FONT_PT:g}pt; font-style: italic; }} .sign {{ margin-top: 18px; font-size: {NEW_REPORT_BODY_FONT_PT:g}pt; text-align: right; }}
</style></head><body>
<div class="header"><div class="company">{_mcoc_html_multiline(header_html)}</div><div class="date">Ngày xuất báo cáo:<br>{date_str}</div></div>
<h1>Báo cáo tính toán nội lực cọc</h1>
<div class="sub">Công trình: {_mcoc_html_multiline(g.get('Project_ID', ''))}</div>
{''.join(section_html)}
<div class="sign">{author_html}</div>
</body></html>"""
        with open(filepath, 'w', encoding='utf-8-sig') as f:
            f.write(html_doc)

    def export_to_new_text(data, displacements, forces, verifications, efficiencies, filepath,
                           equivalent_stiffness=None, rm_bridge_stiffness=None,
                           company="TẬP ĐOÀN SUNGROUP", department="KHỐI XD PPP - PHÒNG QLTK", author='TSDUNGVN - 2026',
                           report_config=None):
        sections = _mcoc_report_sections(data, displacements, forces, verifications, efficiencies, equivalent_stiffness, rm_bridge_stiffness, report_config)
        g = data['Global']
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"{company}\n{department}\n\n")
            f.write("BÁO CÁO TÍNH TOÁN NỘI LỰC CỌC\n")
            f.write(f"Công trình: {g.get('Project_ID', '')}\n")
            f.write(f"Ngày xuất báo cáo: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n")
            for title, headers, rows, note in sections:
                f.write(title.upper() + "\n")
                if note:
                    f.write(note + "\n")
                if headers:
                    f.write("\t".join(map(str, headers)) + "\n")
                for row in rows:
                    f.write("\t".join(map(str, row)) + "\n")
                f.write("\n")
            safe_author = str(author or "").replace("\r", "").strip()
            f.write("\n" + safe_author + "\n")
            last_author_line = [line for line in safe_author.split("\n") if line.strip()]
            if last_author_line:
                f.write(("_" * len(last_author_line[-1].strip())) + "\n")

    def _mcoc_register_pdf_font():
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            candidates = [
                r"C:\Windows\Fonts\times.ttf", r"C:\Windows\Fonts\timesnewroman.ttf", r"C:\Windows\Fonts\arial.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            ]
            for fp in candidates:
                if os.path.exists(fp):
                    pdfmetrics.registerFont(TTFont("BaoCaoMCOC", fp))
                    return "BaoCaoMCOC"
        except Exception:
            pass
        return "Helvetica"

    def export_to_new_pdf(data, displacements, forces, verifications, efficiencies, filepath,
                          equivalent_stiffness=None, rm_bridge_stiffness=None,
                          company="TẬP ĐOÀN SUNGROUP", department="KHỐI XD PPP - PHÒNG QLTK", author='TSDUNGVN - 2026',
                          report_config=None):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
        sections = _mcoc_report_sections(data, displacements, forces, verifications, efficiencies, equivalent_stiffness, rm_bridge_stiffness, report_config)
        g = data['Global']
        font_name = _mcoc_register_pdf_font()
        doc = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=13*mm, bottomMargin=13*mm)
        styles = getSampleStyleSheet()
        base = ParagraphStyle(
            "mcoc_base",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=NEW_REPORT_BODY_FONT_PT,
            leading=14.4,
        )
        table_cell_12 = ParagraphStyle(
            "mcoc_table_cell_12",
            parent=base,
            fontSize=NEW_REPORT_TABLE_FONT_PT,
            leading=NEW_REPORT_TABLE_LINE_HEIGHT_PT,
            spaceBefore=0,
            spaceAfter=0,
        )
        table_cell_11 = ParagraphStyle(
            "mcoc_table_cell_11",
            parent=base,
            fontSize=NEW_REPORT_TABLE_OVERFLOW_FONT_PT,
            leading=NEW_REPORT_TABLE_OVERFLOW_LINE_HEIGHT_PT,
            spaceBefore=0,
            spaceAfter=0,
        )
        title_style = ParagraphStyle("mcoc_title", parent=base, fontSize=16.5, leading=19.5, alignment=1, spaceAfter=5)
        sub_style = ParagraphStyle("mcoc_sub", parent=base, fontSize=NEW_REPORT_BODY_FONT_PT, leading=14.4, alignment=1, italic=True)
        h_style = ParagraphStyle("mcoc_h", parent=base, fontSize=13.0, leading=15.0, spaceBefore=7, spaceAfter=4)
        story = []
        date_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        header_lines = _mcoc_center_header_lines(company, department)
        header_pdf_html = "<br/>".join(_mcoc_html_pre(line) for line in header_lines)
        header_tbl = Table([
            [Paragraph(f"<b>{header_pdf_html}</b>", base), Paragraph(f"Ngày xuất báo cáo:<br/>{date_str}", base)]
        ], colWidths=[118*mm, 54*mm])
        header_tbl.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP"), ("ALIGN", (1,0), (1,0), "RIGHT")]))
        story.append(header_tbl)
        story.append(Paragraph("<b>BÁO CÁO TÍNH TOÁN NỘI LỰC CỌC</b>", title_style))
        story.append(Paragraph(f"Công trình: {_mcoc_html_multiline(g.get('Project_ID', ''))}", sub_style))
        story.append(Spacer(1, 4))

        available_mm = 186.0
        for title, headers, rows, note in sections:
            story.append(Paragraph(f"<b>{_mcoc_html_multiline(title)}</b>", h_style))
            if note:
                story.append(Paragraph(_mcoc_html_multiline(note), base))
            if not rows:
                continue

            span_cmds = []
            row_font_pts = []
            if headers:
                ncols = max(len(headers), 1)
                header_font = _new_report_row_font_pt(headers, ncols, True)
                header_style = table_cell_11 if header_font == NEW_REPORT_TABLE_OVERFLOW_FONT_PT else table_cell_12
                data_rows = [[Paragraph(_mcoc_html_pre(h), header_style) for h in headers]]
                row_font_pts.append(header_font)

                for row in rows:
                    padded = list(row) + [""] * max(0, ncols - len(row))
                    row_font = _new_report_row_font_pt(padded[:ncols], ncols)
                    row_style = table_cell_11 if row_font == NEW_REPORT_TABLE_OVERFLOW_FONT_PT else table_cell_12
                    data_rows.append([Paragraph(_mcoc_html_pre(v), row_style) for v in padded[:ncols]])
                    row_font_pts.append(row_font)

                if str(title).startswith("3. Thông số cọc"):
                    for ridx, row in enumerate(rows, start=1):
                        padded = list(row) + [""] * max(0, ncols - len(row))
                        if ridx > 1 and _mcoc_is_nt_row(padded):
                            row_style = (
                                table_cell_11
                                if row_font_pts[ridx] == NEW_REPORT_TABLE_OVERFLOW_FONT_PT
                                else table_cell_12
                            )
                            data_rows[ridx][1] = Paragraph("nt", row_style)
                            for cidx in range(2, ncols):
                                data_rows[ridx][cidx] = Paragraph("", row_style)
                            span_cmds.append(("SPAN", (1, ridx), (ncols - 1, ridx)))

                if str(title).startswith("6. Nội lực đầu cọc"):
                    ridx = 1
                    while ridx <= len(rows):
                        row = list(rows[ridx - 1]) + [""] * max(0, ncols - len(rows[ridx - 1]))
                        if str(row[0] or "").strip():
                            start_row = ridx
                            end_row = ridx
                            nxt = ridx + 1
                            while nxt <= len(rows):
                                row_next = list(rows[nxt - 1]) + [""] * max(0, ncols - len(rows[nxt - 1]))
                                if str(row_next[0] or "").strip():
                                    break
                                end_row = nxt
                                nxt += 1
                            if end_row > start_row:
                                span_cmds.append(("SPAN", (0, start_row), (0, end_row)))
                            ridx = nxt
                        else:
                            ridx += 1

                col_widths = _new_report_column_widths_pt(headers, rows, available_mm * mm)
            else:
                ncols = 1
                joined_rows = [" ".join(map(str, row)) for row in rows]
                data_rows = []
                for joined in joined_rows:
                    row_font = _new_report_row_font_pt([joined], ncols)
                    row_style = table_cell_11 if row_font == NEW_REPORT_TABLE_OVERFLOW_FONT_PT else table_cell_12
                    data_rows.append([Paragraph(_mcoc_html_pre(joined), row_style)])
                    row_font_pts.append(row_font)
                col_widths = [available_mm * mm]

            tbl = Table(data_rows, colWidths=col_widths, repeatRows=1 if headers else 0)
            style_cmds = [
                ("GRID", (0,0), (-1,-1), 0.35, colors.black),
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E9EEF7")) if headers else ("BACKGROUND", (0,0), (-1,-1), colors.white),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                ("ALIGN", (0,0), (-1,-1), "CENTER"),
                ("FONTNAME", (0,0), (-1,-1), font_name),
                ("LEFTPADDING", (0,0), (-1,-1), 1.5),
                ("RIGHTPADDING", (0,0), (-1,-1), 1.5),
            ] + span_cmds

            for ridx, row_font in enumerate(row_font_pts):
                _, vertical_padding = _new_report_pdf_metrics(ncols, row_font)
                style_cmds.extend([
                    ("TOPPADDING", (0, ridx), (-1, ridx), vertical_padding),
                    ("BOTTOMPADDING", (0, ridx), (-1, ridx), vertical_padding),
                ])

            tbl.setStyle(TableStyle(style_cmds))
            story.append(tbl)
            story.append(Spacer(1, 3))
        story.append(Spacer(1, 12))
        author_pdf_html = "<br/>".join(f"<u>{_mcoc_html_multiline(line)}</u>" for line in _mcoc_author_lines(author))
        story.append(Paragraph(author_pdf_html, ParagraphStyle("sign", parent=base, alignment=2)))
        doc.build(story)


    def export_to_excel(data, displacements, forces, verifications, efficiencies, filepath, 
                        equivalent_stiffness=None, rm_bridge_stiffness=None, 
                        company="TẬP ĐOÀN SUNGROUP", department="KHỐI XD PPP - PHÒNG QLTK", author='TSDUNGVN - 2026',
                        report_config=None):
        ensure_pandas()
        ensure_openpyxl()
        if report_config is None:
            report_config = {'p1':True, 'p2':True, 'p3':True, 'p4':True, 'p5':True, 'p6':True, 'p7':True, 'p7_ttgh':False, 'ttgh_groups':[], 'p8':True, 'p8_opt':2, 'p9':True, 'p9_direct':False}

        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Báo cáo nội lực cọc"; ws.views.sheetView[0].showGridLines = True
        
        # Format Excel Header
        c_lines = company.split('\n')
        d_lines = department.split('\n')
        for line in c_lines: ws.append([line.strip()])
        for line in d_lines: ws.append([line.strip()])
        
        ws.append([]); ws.append(["CHƯƠNG TRÌNH TÍNH MÓNG CỌC"]); ws.append([f"Công trình: {data['Global'].get('Project_ID', 'Cau')}"]); ws.append([])
        
        g = data['Global']
        
        if report_config['p1']:
            ws.append(["THÔNG SỐ HÌNH HỌC & ĐỊA CHẤT"])
            ws.append(["Kn", g['Kn'], "Ax", g['Bx'], "By", g['By'], "Cz", g['Cz']])
            ws.append(["E v.uon", g['EI_uon'], "E r.uon", g.get('Er_uon', 0), "E v.nen", g['EA_nen'], "E r.nen", g.get('Er_nen', 0)])
            ws.append(["Mq (t/m4)", g['mq'], "Md (t/m4)", g['md'], "m (t/m4)", g['m']]); ws.append([])

        if report_config['p2']:
            ws.append(["CÁC TỔ HỢP TẢI TRỌNG TÍNH TOÁN"]); ws.append(["T.T", "Hx", "Hy", "P", "Mx", "My", "Mz"])
            for c in data['Load_Combos']: ws.append([c['Name'], c['Hx'], c['Hy'], c['N_load'], c['Mx'], c['My'], c['Mz']])
            ws.append([])

        if report_config['p3']:
            sec_labels, _ = _section_report_schema(data)
            ws.append(["THÔNG SỐ CỌC"]); ws.append(["T.C", "Lo", "H", "Bpx", "Bpy", sec_labels[0], sec_labels[1], sec_labels[2], "Fo", "Jo", "Po", "Co", "Ct"])
            
            first = data['Piles'][0]
            ws.append([first['Name'], first['Lo'], first['H'], first['Bpx'], first['Bpy'], first['d_ngoai'], first['d_trong'], first['day_vo'], first['Area'], first['J_xy'], first['Po'], first['Co'], first['Ct']])
            
            for i in range(1, len(data['Piles'])):
                p = data['Piles'][i]
                is_same = (abs(p['Lo']-first['Lo'])<1e-5 and abs(p['H']-first['H'])<1e-5 and 
                           abs(p['Bpx']-first['Bpx'])<1e-5 and abs(p['Bpy']-first['Bpy'])<1e-5 and 
                           abs(p['d_ngoai']-first['d_ngoai'])<1e-5 and abs(p['d_trong']-first['d_trong'])<1e-5 and 
                           abs(p['day_vo']-first['day_vo'])<1e-5 and abs(p['Area']-first['Area'])<1e-5 and 
                           abs(p['J_xy']-first['J_xy'])<1e-5 and abs(p['Po']-first['Po'])<1e-5 and 
                           abs(p['Co']-first['Co'])<1e-5 and abs(p['Ct']-first['Ct'])<1e-5)
                if is_same:
                    ws.append([p['Name'], "như trên"])
                else:
                    ws.append([p['Name'], p['Lo'], p['H'], p['Bpx'], p['Bpy'], p['d_ngoai'], p['d_trong'], p['day_vo'], p['Area'], p['J_xy'], p['Po'], p['Co'], p['Ct']])
            
            ws.append(["Ghi chú:", _pile_report_description(first)])
            ws.append([])

        if report_config['p4']:
            ws.append(["TOẠ ĐỘ ĐẦU CỌC"]); ws.append(["T.C", "X", "Y", "Phi", "Xi"])
            for p in data['Piles']: ws.append([p['Name'], p['X'], p['Y'], p['Phi'], p['Xi']])
            ws.append([])

            if data.get('Config', {}).get('Group_Effect_Enabled', False):
                method = data.get('Config', {}).get('Group_Method', '')
                ws.append([f"HỆ SỐ SUY GIẢM ĐỘ CỨNG NHÓM (Theo: {method})"])
                ws.append(["T.C", "Eff_V(Dọc)", "Eff_H(Ngang)"])
                for p in data['Piles']: ws.append([p['Name'], round(efficiencies[p['Name']]['v'], 3), round(efficiencies[p['Name']]['h'], 3)])
                ws.append([])

            if data.get('Config', {}).get('Shadow_Effect', False):
                ws.append(["LƯU Ý: ĐÃ XÉT HIỆU ỨNG BÓNG RÂM THEO TCVN 11823"])
                ws.append(["Độ cứng ngang của cọc thay đổi động theo từng tổ hợp dựa trên hướng lực đẩy."])
                ws.append([])

        if report_config['p5']:
            ws.append(["CHUYỂN VỊ BỆ CỌC"]); ws.append(["T.H", "X", "Y", "Z", "Fix", "Fiy", "Fiz"])
            for d in displacements: ws.append([d['T.H'], round(d['X'], 5), round(d['Y'], 5), round(d['Z'], 5), round(d['Fix'], 5), round(d['Fiy'], 5), round(d['Fiz'], 5)])
            ws.append([])

        if report_config['p6']:
            ws.append(["NỘI LỰC ĐẦU CỌC"]); ws.append(["T.C", "T.H", "N", "Q2", "Q3", "M1", "M2", "M3"])
            df_forces = pd.DataFrame(forces)
            for pile_name, group in df_forces.groupby('T.C', sort=False):
                first_row = True
                for _, row in group.iterrows():
                    ws.append([pile_name if first_row else "", row['T.H'], row['N'], row['Q2'], row['Q3'], row['M1'], row['M2'], row['M3']])
                    first_row = False
            ws.append([])

        if report_config['p7']:
            ws.append(["BẢNG TỔNG KẾT NỘI LỰC"]); ws.append(["Loại", "T.C", "T.H", "N", "Q2", "Q3", "M1", "M2", "M3"])
            for name, r in _force_summary_items(forces):
                ws.append([name, r['T.C'], r['T.H'], r['N'], r['Q2'], r['Q3'], r['M1'], r['M2'], r['M3']])
            ws.append([])

        ttgh_groups = _iter_ttgh_force_groups(forces, report_config)
        if ttgh_groups:
            ws.append(["TỔNG HỢP NỘI LỰC THEO TRẠNG THÁI GIỚI HẠN"])
            for group, subset in ttgh_groups:
                ws.append([f"{group['label']} ({group['short_label']}) - Tổ hợp {group['actual_spec']}"])
                ws.append(["Loại", "T.C", "T.H", "N", "Q2", "Q3", "M1", "M2", "M3"])
                for name, r in _force_summary_items(subset):
                    ws.append([name, r['T.C'], r['T.H'], r['N'], r['Q2'], r['Q3'], r['M1'], r['M2'], r['M3']])
                ws.append([])

        if report_config['p8']:
            ws.append(["TÍNH TOÁN KIỂM TRA (TÍNH NGƯỢC THỬ LẠI BÀI TOÁN P = K * Delta)"])
            if report_config['p8_opt'] == 1:
                ws.append(["Tổ hợp", "Hx", "Hy", "P", "Mx", "My", "Mz"])
                for v in verifications:
                    ws.append([f"TH {v['Name']}", v['Recalculated'][0], v['Recalculated'][1], v['Recalculated'][2], v['Recalculated'][3], v['Recalculated'][4], v['Recalculated'][5]])
            else:
                ws.append(["Tổ hợp", "Loại tải", "Hx", "Hy", "P", "Mx", "My", "Mz"])
                for v in verifications:
                    ws.append([f"TH {v['Name']}", "Ban đầu", v['Original'][0], v['Original'][1], v['Original'][2], v['Original'][3], v['Original'][4], v['Original'][5]])
                    ws.append(["", "Tính ngược", v['Recalculated'][0], v['Recalculated'][1], v['Recalculated'][2], v['Recalculated'][3], v['Recalculated'][4], v['Recalculated'][5]])
                    ws.append(["", "Sai số"] + ["OK" if abs(v['Error'][i]) < 0.001 else round(v['Error'][i], 4) for i in range(6)])

        if report_config['p9'] and equivalent_stiffness is not None:
            ws.append([]); ws.append(["ĐỘ CỨNG TƯƠNG ĐƯƠNG CỦA MÓNG CỌC"])
            labels = equivalent_stiffness['labels']; tt = equivalent_stiffness['truc_tiep']; td = equivalent_stiffness['tuong_duong']
            
            if report_config.get('p9_direct', False):
                ws.append(["(1) Độ cứng trực tiếp (Đường chéo ma trận K - Chưa khử liên kết - Tham khảo)"]); ws.append(labels); ws.append([tt[lb] for lb in labels]); ws.append([])
                ws.append(["(2) Độ cứng tương đương ĐÃ KHỬ LIÊN KẾT (Nên dùng)"])
            else:
                ws.append(["Độ cứng tương đương ĐÃ KHỬ LIÊN KẾT (P = Keq * Delta)"])

            ws.append(labels); ws.append([td[lb] if td[lb] is not None else "N/A" for lb in labels]); ws.append([])
            ws.append(["Đơn vị: Kx[t/m], Ky[t/m], Kz[t/m], Kmx[t.m/rad], Kmy[t.m/rad], Kmz[t.m/rad]"])
            
            if rm_bridge_stiffness is not None:
                ws.append([]); ws.append(["ĐỘ CỨNG TƯƠNG QUY ĐỔI CHO RM BRIDGE"])
                rm_labels = ['Cx', 'Cy', 'Cz', 'CMx', 'CMy', 'CMz']
                ws.append(rm_labels); ws.append([rm_bridge_stiffness[lb] if rm_bridge_stiffness[lb] is not None else "N/A" for lb in rm_labels])
                ws.append(["Đơn vị: Cx[kN/m], Cy[kN/m], Cz[kN/m], CMx[kN.m/rad], CMy[kN.m/rad], CMz[kN.m/rad]"])
                ws.append(["* Lưu ý: Lò xo này tính từ ma trận cơ sở, chưa bao gồm P-multiplier (nếu có)."])

        ws.append([]); 
        
        # Excel format for author
        a_lines = author.split('\n')
        for line in a_lines:
            ws.append(["", "", "", "", "", "", line.strip()])
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=13):
            for cell in row:
                if isinstance(cell.value, (int, float)): cell.alignment = Alignment(horizontal='right')
        wb.save(filepath)

    def export_to_pdf_from_text(doc_path, pdf_path):
        ensure_reportlab()
        font_name = "Courier"
        try:
            candidates = [os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", f) for f in ["cour.ttf", "consola.ttf", "seguisb.ttf", "arial.ttf"]]
            for fp in candidates:
                if os.path.exists(fp):
                    pdfmetrics.registerFont(TTFont('VNFont', fp))
                    font_name = 'VNFont'
                    break
        except: pass
        with open(doc_path, 'r', encoding='utf-8') as f: lines = f.readlines()
        c = Canvas(pdf_path, pagesize=A4)
        width, height = A4
        font_size, line_height, y = 9.0, 12.0, height - 50
        left_margin = max(25, (width - max([len(line.rstrip('\n')) for line in lines]) * 0.6 * font_size) / 2)
        for line in lines:
            c.setFont(font_name, font_size)
            c.drawString(left_margin, y, line.rstrip('\n'))
            y -= line_height
            if y < 50: c.showPage(); c.setFont(font_name, font_size); y = height - 50
        c.save()

    def merge_pdfs(pdf_list, output_path):
        ensure_pypdf()
        try: merger = pypdf.PdfWriter()
        except AttributeError: merger = pypdf.PdfMerger()
        for pdf in pdf_list: merger.append(pdf)
        merger.write(output_path); merger.close()

    # ==========================================
    # HÀM AN TOÀN TUYỆT ĐỐI GHI ĐÈ FILE TRÙNG
    # ==========================================
    def safe_save_file(root, check_filepath, export_func, *args, **kwargs):
        import time
        while True:
            try:
                export_func(*args, **kwargs)
                return True
            except (PermissionError, IOError):
                result = [False]
                event = threading.Event()
                def ask_user():
                    res = messagebox.askyesno(
                        "File đang bị khóa",
                        f"File sau đang được mở bởi phần mềm khác (Word/Excel):\n{os.path.basename(check_filepath)}\n\n"
                        "Tool sẽ TỰ ĐỘNG ĐÓNG file này để ghi đè. Bạn có đồng ý không?\n\n"
                        "(Lưu ý: Nếu đã bấm Yes mà vẫn hiện bảng này, vui lòng tự đóng file bằng tay rồi bấm Yes)"
                    )
                    result[0] = res
                    event.set()
                
                # Gọi messagebox từ luồng chính để tránh lỗi sập Tkinter
                root.after(0, ask_user)
                event.wait()
                
                if result[0]:
                    try:
                        abs_path = os.path.abspath(check_filepath)
                        filepath_ps = abs_path.replace("'", "''")
                        if check_filepath.lower().endswith(('.doc', '.docx')):
                            ps_script = f"""
                            try {{
                                $word = [System.Runtime.InteropServices.Marshal]::GetActiveObject('Word.Application')
                                foreach ($doc in $word.Documents) {{
                                    if ($doc.FullName.ToLower() -eq '{abs_path.lower()}'.ToLower()) {{
                                        $doc.Close($false)
                                    }}
                                }}
                            }} catch {{ }}
                            """
                            subprocess.run(["powershell", "-ExecutionPolicy", "Bypass", "-Command", ps_script], creationflags=0x08000000)
                        elif check_filepath.lower().endswith(('.xls', '.xlsx')):
                            ps_script = f"""
                            try {{
                                $excel = [System.Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application')
                                foreach ($wb in $excel.Workbooks) {{
                                    if ($wb.FullName.ToLower() -eq '{abs_path.lower()}'.ToLower()) {{
                                        $wb.Close($false)
                                    }}
                                }}
                            }} catch {{ }}
                            """
                            subprocess.run(["powershell", "-ExecutionPolicy", "Bypass", "-Command", ps_script], creationflags=0x08000000)
                    except:
                        pass
                    
                    time.sleep(1.0) # Đợi 1 giây để Windows nhả file lock
                    continue
                else:
                    return False

    # ==========================================
    # 4. GIAO DIỆN PHẦN MỀM (TKINTER GUI) - CHUẨN TAB
    # ==========================================
    class PileDesignToolApp:
        def __init__(self, root, days_left_info):
            self.root = root
            self.days_left = days_left_info
            self.theme_key = UI_THEME_AQUA
            self.pal = THEME_PRESETS[self.theme_key]
            self.root.title(APP_NAME)
            self.root.geometry("1020x740")
            self.root.minsize(980, 700)
            self._maximize_main_window()
            apply_app_icon(self.root, "ts_pile")
            self.file_list = []
            self.file_display_names = []
            self.project_file_path = ""
            self.last_col_csv = ""
            self.last_col_forces = []
            self.workflow_buttons = []
            self.workflow_steps = []
            self.setup_style()
            self.create_widgets()
            self.apply_theme(self.theme_key)
            for _delay in (80, 350, 900):
                try:
                    self.root.after(_delay, self._maximize_main_window)
                except Exception:
                    pass
            try:
                self.root.bind_all("<Control-o>", lambda _e: (self.open_project_file(), "break"))
                self.root.bind_all("<Control-O>", lambda _e: (self.open_project_file(), "break"))
                self.root.bind_all("<Control-s>", lambda _e: (self.save_project_file(), "break"))
                self.root.bind_all("<Control-S>", lambda _e: (self.save_project_file(), "break"))
            except Exception:
                pass

        def _maximize_main_window(self):
            """Mở cửa sổ chính ở trạng thái phóng to, không dùng fullscreen không viền."""
            try:
                self.root.state("zoomed")
                return
            except Exception:
                pass
            try:
                self.root.wm_state("zoomed")
                return
            except Exception:
                pass
            try:
                self.root.attributes("-zoomed", True)
                return
            except Exception:
                pass
            try:
                sw = self.root.winfo_screenwidth()
                sh = self.root.winfo_screenheight()
                self.root.geometry(f"{sw}x{sh}+0+0")
            except Exception:
                pass

        def setup_style(self):
            try:
                default_font = tkfont.nametofont("TkDefaultFont")
                default_font.configure(family="Segoe UI", size=10)
                self.root.option_add("*Font", default_font)
            except Exception:
                pass
            self.style = ttk.Style(self.root)
            try:
                self.style.theme_use("clam")
            except Exception:
                pass

        def apply_theme(self, theme_key=None):
            self.theme_key = normalize_ui_theme(theme_key or self.theme_key)
            pal = THEME_PRESETS.get(self.theme_key, THEME_PRESETS[UI_THEME_AQUA])
            self.pal = pal
            try:
                self.root.configure(bg=pal["bg"])
            except Exception:
                pass
            st = self.style
            st.configure("TFrame", background=pal["bg"])
            st.configure("Panel.TFrame", background=pal["panel"])
            st.configure("TLabel", font=("Segoe UI", 10), background=pal["bg"], foreground=pal["text"])
            st.configure("Title.TLabel", background=pal["bg"], foreground=pal["accent_dark"], font=("Segoe UI", 18, "bold"))
            st.configure("Subtitle.TLabel", background=pal["bg"], foreground=pal["muted"], font=("Segoe UI", 9))
            st.configure("Header.TLabel", font=("Segoe UI", 17, "bold"), background=pal["bg"], foreground=pal["accent_dark"])
            st.configure("SubHeader.TLabel", font=("Segoe UI", 9), background=pal["bg"], foreground=pal["muted"])
            st.configure("TLabelframe", background=pal["bg"], bordercolor=pal["border"], lightcolor=pal["border"], darkcolor=pal["border"])
            st.configure("TLabelframe.Label", font=("Segoe UI", 10, "bold"), background=pal["bg"], foreground=pal["accent_dark"])
            st.configure("TNotebook", background=pal["bg"], bordercolor=pal["border"])
            st.configure("TNotebook.Tab", font=("Segoe UI", 8, "bold"), padding=(5, 3), background=pal["button"], foreground=pal["text"])
            st.map("TNotebook.Tab", background=[("selected", pal["panel"]), ("active", pal["button_active"])])
            st.configure("TButton", font=("Segoe UI", 9), padding=(5, 4), background=pal["button"], foreground=pal["text"], bordercolor=pal["border"])
            st.map("TButton", background=[("active", pal["button_active"]), ("pressed", pal["big_button"])], foreground=[("disabled", "#888888")])
            st.configure("Small.TButton", font=("Segoe UI", 9), padding=(4, 3), background=pal["button"], foreground=pal["text"])
            st.configure("Big.TButton", font=("Segoe UI", 10, "bold"), padding=(7, 5), background=pal["big_button"], foreground=pal["text"])
            st.map("Big.TButton", background=[("active", pal["button_active"]), ("pressed", pal["border"])])
            st.configure("Primary.TButton", font=("Segoe UI", 10, "bold"), padding=(12, 8), background=pal["big_button"], foreground=pal["text"])
            st.configure("About.TButton", font=("Segoe UI", 10, "bold"), padding=(8, 6), foreground=pal["accent_dark"], background=pal["button"])
            st.map("About.TButton", background=[("active", pal["button_active"])])
            st.configure("Horizontal.TProgressbar", thickness=10, background=pal["progress"], troughcolor=pal["trough"], lightcolor=pal["progress"], darkcolor=pal["progress"], bordercolor=pal["border"])
            st.configure("Status.Horizontal.TProgressbar", thickness=16, background=pal["progress"], troughcolor=pal["trough"], lightcolor=pal["progress"], darkcolor=pal["progress"], bordercolor=pal["border"])
            st.configure("TEntry", font=("Segoe UI", 10), fieldbackground=pal["entry_bg"], background=pal["entry_bg"], foreground=pal["entry_fg"], insertcolor=pal["entry_fg"])
            st.configure("TCombobox", font=("Segoe UI", 10), fieldbackground=pal["entry_bg"], background=pal["entry_bg"], foreground=pal["entry_fg"], selectbackground=pal["entry_bg"], selectforeground=pal["entry_fg"])
            st.map("TCombobox", fieldbackground=[("readonly", pal["entry_bg"]), ("!disabled", pal["entry_bg"])], background=[("readonly", pal["entry_bg"]), ("!disabled", pal["entry_bg"])], foreground=[("readonly", pal["entry_fg"]), ("!disabled", pal["entry_fg"])])
            st.configure("Treeview", font=("Segoe UI", 10), rowheight=26, background=pal.get("tree_row", "white"), fieldbackground=pal.get("tree_row", "white"), foreground=pal.get("tree_fg", "#111111"), bordercolor=pal["border"])
            st.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background=pal["tree_head"], foreground=pal["text"], bordercolor=pal["border"])
            st.map("Treeview", background=[("selected", pal["accent"])], foreground=[("selected", "white")])
            st.configure("TCheckbutton", background=pal["bg"], foreground=pal["text"], font=("Segoe UI", 10))
            st.configure("TRadiobutton", background=pal["bg"], foreground=pal["text"], font=("Segoe UI", 10))
            self._apply_tk_theme(self.root)
            self._refresh_special_theme_colors()

        def _refresh_special_theme_colors(self):
            pal = self.pal
            try:
                self.left_frame.configure(bg=pal["sidebar"])
                sep = getattr(self, "sidebar_separator", None)
                def _paint_sidebar(widget):
                    try:
                        if widget is sep:
                            widget.configure(bg=pal["sidebar2"])
                        elif widget.winfo_class() == "Frame":
                            widget.configure(bg=pal["sidebar"])
                        elif widget.winfo_class() == "Label":
                            widget.configure(bg=pal["sidebar"], fg="white")
                    except Exception:
                        pass
                    for sub in widget.winfo_children():
                        _paint_sidebar(sub)
                _paint_sidebar(self.left_frame)
            except Exception:
                pass
            try:
                self.btn_run.configure(bg=pal["accent"], fg="white", activebackground=pal["accent_dark"], activeforeground="white")
            except Exception:
                pass
            try:
                self._update_workflow_buttons()
            except Exception:
                pass
            try:
                self.status_label.configure(fg=pal["accent_dark"], bg=pal["bg"])
            except Exception:
                pass
            try:
                self.header_frame.configure(bg=pal["bg"])
                self.title_frame.configure(bg=pal["bg"])
                self.about_frame.configure(bg=pal["bg"])
            except Exception:
                pass

        def _apply_tk_theme(self, widget):
            pal = self.pal
            try:
                cls = widget.winfo_class()
            except Exception:
                return
            try:
                if cls in ("Frame", "Toplevel"):
                    widget.configure(bg=pal["bg"])
                elif cls == "Label":
                    widget.configure(bg=pal["bg"], fg=pal["text"])
                elif cls == "Button":
                    widget.configure(bg=pal["accent"], fg="white", activebackground=pal["accent_dark"], activeforeground="white", relief=tk.FLAT, bd=0)
                elif cls == "Listbox":
                    widget.configure(bg="white", fg="#111111", selectbackground=pal["accent"], selectforeground="white", highlightbackground=pal["border"], highlightcolor=pal["accent"])
                elif cls == "Entry":
                    widget.configure(bg="white", fg="black", insertbackground="black")
            except Exception:
                pass
            for child in widget.winfo_children():
                self._apply_tk_theme(child)

        def show_about(self):
            win = tk.Toplevel(self.root)
            win.title(f"About - {APP_NAME}")
            win.transient(self.root)
            win.grab_set()
            win.resizable(False, False)
            win.configure(bg=self.pal["panel"])
            frm = tk.Frame(win, bg=self.pal["panel"], padx=26, pady=22)
            frm.pack(fill=tk.BOTH, expand=True)
            tk.Label(frm, text=APP_NAME, font=("Segoe UI", 24, "bold"), fg=self.pal["accent_dark"], bg=self.pal["panel"]).pack(anchor=tk.W)
            tk.Label(frm, text="Phần mềm tính toán móng cọc", font=("Segoe UI", 12, "italic"), fg=self.pal["muted"], bg=self.pal["panel"]).pack(anchor=tk.W, pady=(2, 14))
            tk.Frame(frm, height=1, bg=self.pal["border"]).pack(fill=tk.X, pady=(0, 14))
            tk.Label(frm, text="Tác giả", font=("Segoe UI", 11, "bold"), fg=self.pal["text"], bg=self.pal["panel"]).pack(anchor=tk.W)
            tk.Label(frm, text="Nguyễn Ngọc Dũng", font=("Segoe UI", 14, "bold"), fg=self.pal["accent_dark"], bg=self.pal["panel"]).pack(anchor=tk.W, pady=(2, 8))
            tk.Label(frm, text="Phòng QLTK - Khối XD PPP\nTập đoàn SunGroup", justify=tk.LEFT, font=("Segoe UI", 10), fg=self.pal["text"], bg=self.pal["panel"]).pack(anchor=tk.W, pady=(0, 16))
            ttk.Button(frm, text="OK", style="Big.TButton", command=win.destroy).pack(fill=tk.X)
            self._center_window(win)

        def show_license(self):
            win = tk.Toplevel(self.root)
            win.title(f"License - {APP_NAME}")
            win.geometry("560x260")
            win.resizable(False, False)
            win.configure(bg=self.pal["bg"])
            try:
                win.transient(self.root)
                win.grab_set()
            except Exception:
                pass

            frm = ttk.Frame(win, padding=18)
            frm.pack(fill=tk.BOTH, expand=True)
            ttk.Label(frm, text="License", font=("Segoe UI", 20, "bold"), foreground=self.pal["accent_dark"]).grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=(0, 14))

            days_txt = self.days_left
            try:
                days_txt = int(float(days_txt))
            except Exception:
                pass
            ttk.Label(frm, text="Số ngày sử dụng còn lại:", font=("Segoe UI", 10, "bold")).grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=6)
            ttk.Label(frm, text=f"{days_txt} ngày", font=("Segoe UI", 11, "bold"), foreground="#16a34a").grid(row=1, column=1, sticky=tk.W, pady=6)

            try:
                full_machine_id = get_machine_id()
            except Exception as exc:
                full_machine_id = f"Không lấy được HWID: {exc}"

            ttk.Label(frm, text="HWID / Mã thiết bị:", font=("Segoe UI", 10, "bold")).grid(row=2, column=0, sticky=tk.W, padx=(0, 10), pady=6)
            hwid_var = tk.StringVar(value=full_machine_id)
            hwid_entry = ttk.Entry(frm, textvariable=hwid_var, width=48, state="readonly")
            hwid_entry.grid(row=2, column=1, sticky=tk.EW, pady=6)

            def copy_hwid():
                self.root.clipboard_clear()
                self.root.clipboard_append(full_machine_id)
                self.status_var.set("Đã copy HWID vào clipboard")

            ttk.Button(frm, text="Copy", command=copy_hwid).grid(row=2, column=2, sticky=tk.W, padx=(8, 0), pady=6)
            ttk.Label(frm, text="Gửi HWID này cho người quản lý license khi cần kích hoạt/gia hạn.", style="Muted.TLabel", wraplength=500).grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=(10, 4))
            ttk.Button(frm, text="Đóng", style="Big.TButton", command=win.destroy).grid(row=4, column=1, sticky=tk.E, pady=(16, 0))
            frm.columnconfigure(1, weight=1)
            self._center_window(win)

        def show_settings(self):
            win = tk.Toplevel(self.root)
            win.title(f"Settings - {APP_NAME}")
            win.transient(self.root)
            win.grab_set()
            win.resizable(False, False)
            frm = ttk.Frame(win, padding=18)
            frm.pack(fill=tk.BOTH, expand=True)
            ttk.Label(frm, text="Settings", font=("Segoe UI", 20, "bold"), foreground=self.pal["accent_dark"]).grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 12))
            ttk.Label(frm, text="Theme giao diện", font=("Segoe UI", 11, "bold")).grid(row=1, column=0, sticky=tk.W, pady=6, padx=(0, 12))
            theme_var = tk.StringVar(value=ui_theme_label(self.theme_key))
            cb_theme = ttk.Combobox(frm, textvariable=theme_var, values=UI_THEME_CHOICES, state="readonly", width=48)
            cb_theme.grid(row=1, column=1, sticky=tk.EW, pady=6)
            frm.columnconfigure(1, weight=1)
            def apply_settings(close_after=False):
                self.apply_theme(normalize_ui_theme(theme_var.get()))
                self.status_var.set(f"Đã đổi theme: {ui_theme_label(self.theme_key)}")
                if close_after:
                    win.destroy()
            btns = ttk.Frame(frm)
            btns.grid(row=2, column=0, columnspan=2, sticky=tk.EW, pady=(12, 0))
            ttk.Button(btns, text="OK", style="Big.TButton", command=lambda: apply_settings(True)).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=3)
            ttk.Button(btns, text="Apply", command=lambda: apply_settings(False)).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=3)
            ttk.Button(btns, text="Cancel", command=win.destroy).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=3)
            win.bind("<Return>", lambda _e: apply_settings(False))
            win.bind("<Escape>", lambda _e: win.destroy())
            self._center_window(win)

        def _center_window(self, win):
            try:
                win.update_idletasks()
                x = self.root.winfo_x() + (self.root.winfo_width() - win.winfo_width()) // 2
                y = self.root.winfo_y() + (self.root.winfo_height() - win.winfo_height()) // 2
                win.geometry(f"+{max(x,0)}+{max(y,0)}")
            except Exception:
                pass

        def _log(self, message, level="INFO"):
            stamp = time.strftime("%H:%M:%S")
            line = f"[{stamp}] [{level}] {message}"
            def append():
                if hasattr(self, "log_text"):
                    self.log_text.config(state=tk.NORMAL)
                    self.log_text.insert(tk.END, line + "\n")
                    self.log_text.see(tk.END)
                    self.log_text.config(state=tk.DISABLED)
                if hasattr(self, "status_var"):
                    self.status_var.set(message)
            self.root.after(0, append)

        # ------------------------------------------------------------------
        # Lưu / mở dự án TS-PILE dạng JSON nhẹ
        # Chỉ lưu danh sách input và các lựa chọn trên giao diện; không đổi lõi tính toán.
        # ------------------------------------------------------------------
        def _append_input_file(self, path, display_name=None):
            path = str(path or "")
            if not path:
                return
            label = str(display_name or os.path.basename(path) or path).strip() or os.path.basename(path)
            self.file_list.append(path)
            if not hasattr(self, "file_display_names"):
                self.file_display_names = []
            self.file_display_names.append(label)
            try:
                self.listbox_files.insert(tk.END, label)
            except Exception:
                pass

        def _refresh_input_listbox(self):
            try:
                self.listbox_files.delete(0, tk.END)
                names = getattr(self, "file_display_names", []) or []
                if len(names) != len(self.file_list):
                    names = [os.path.basename(x) for x in self.file_list]
                    self.file_display_names = list(names)
                for name in names:
                    self.listbox_files.insert(tk.END, name)
            except Exception:
                pass

        def _project_snapshot(self):
            def bget(name, default=False):
                try:
                    return bool(getattr(self, name).get())
                except Exception:
                    return bool(default)
            def sget(name, default=""):
                try:
                    return str(getattr(self, name).get())
                except Exception:
                    return str(default)
            def iget(name, default=0):
                try:
                    return int(getattr(self, name).get())
                except Exception:
                    return int(default)
            return {
                "file_type": 'TS-PILE project',
                "format_version": 1,
                "app_name": APP_NAME,
                "saved_at": datetime.datetime.now().isoformat(timespec="seconds"),
                "theme_key": str(getattr(self, "theme_key", UI_THEME_AQUA)),
                "file_list": list(getattr(self, "file_list", []) or []),
                "file_display_names": list(getattr(self, "file_display_names", []) or []),
                "out_dir": sget("out_dir_var"),
                "options": {
                    "group_main": bget("chk_group_main"),
                    "group_method": self.cbo_group_method.get() if hasattr(self, "cbo_group_method") else "",
                    "shadow_effect": bget("chk_shadow_effect"),
                    "stiffness": bget("chk_stiffness"),
                    "stiff_direct": bget("chk_stiff_direct"),
                    "rm": bget("chk_rm"),
                    "doc": bget("chk_doc", True),
                    "txt": bget("chk_txt"),
                    "excel": bget("chk_excel"),
                    "pdf": bget("chk_pdf"),
                    "col_csv": bget("chk_col_csv"),
                    "merge": bget("chk_merge"),
                    "print_now": bget("chk_print_now"),
                    "report_form": self.cbo_report_form.get() if hasattr(self, "cbo_report_form") else "Form MCOC Classic",
                    "txt_classic_when_new": bget("chk_txt_classic_when_new"),
                    "p1": bget("chk_p1", True), "p2": bget("chk_p2", True),
                    "p3": bget("chk_p3", True), "p4": bget("chk_p4", True),
                    "p5": bget("chk_p5", True), "p6": bget("chk_p6", True),
                    "p7": bget("chk_p7", True), "p7_ttgh": bget("chk_p7_ttgh", True), "p8": bget("chk_p8", True),
                    "p9": bget("chk_p9", True), "p8_opt": iget("var_p8_opt", 1),
                    "printer": self.cbo_printer.get() if hasattr(self, "cbo_printer") else "Default Printer",
                    "paper_size": self.cbo_paper_size.get() if hasattr(self, "cbo_paper_size") else "A4",
                },
                "report_info": {
                    "company": sget("company_var", "TẬP ĐOÀN SUNGROUP"),
                    "department": sget("department_var", "KHỐI XD PPP - PHÒNG QLTK"),
                    "author": sget("author_var", 'TSDUNGVN - 2026'),
                },
            }

        def save_project_file(self):
            path = getattr(self, "project_file_path", "") or ""
            if not path:
                return self.save_project_file_as()
            try:
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(self._project_snapshot(), f, ensure_ascii=False, indent=2)
                self.status_var.set(f"Đã lưu dự án: {os.path.basename(path)}")
            except Exception as exc:
                messagebox.showerror("Save", f"Không lưu được dự án:\n{exc}")

        def save_project_file_as(self):
            initial = 'TS_PILE_Project.tspile'
            try:
                if self.file_display_names:
                    initial = self._sanitize_filename(str(self.file_display_names[0])) + ".tspile"
            except Exception:
                pass
            path = filedialog.asksaveasfilename(
                title='Lưu dự án TS-PILE',
                defaultextension=".tspile",
                initialfile=initial,
                filetypes=[('TS-PILE Project', "*.tspile"), ("JSON", "*.json"), ("All files", "*.*")],
            )
            if not path:
                return
            old = getattr(self, "project_file_path", "")
            self.project_file_path = path
            try:
                self.save_project_file()
                messagebox.showinfo("Save As", f"Đã lưu dự án:\n{path}")
            except Exception:
                self.project_file_path = old

        def open_project_file(self):
            path = filedialog.askopenfilename(
                title='Mở dự án TS-PILE',
                filetypes=[('TS-PILE Project', "*.tspile"), ("JSON", "*.json"), ("All files", "*.*")],
            )
            if not path:
                return
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self._apply_project_snapshot(data)
                self.project_file_path = path
                self.status_var.set(f"Đã mở dự án: {os.path.basename(path)}")
                messagebox.showinfo("Open", "Đã mở dự án.")
            except Exception as exc:
                messagebox.showerror("Open", f"Không mở được dự án:\n{exc}")

        def _apply_project_snapshot(self, data):
            if not isinstance(data, dict):
                raise ValueError("File dự án không đúng định dạng JSON.")
            if str(data.get("file_type", "")).strip() and 'TS-PILE' not in str(data.get("file_type", "")):
                raise ValueError('File này không phải dự án TS-PILE.')
            self.file_list = [str(x) for x in (data.get("file_list", []) or [])]
            names = [str(x) for x in (data.get("file_display_names", []) or [])]
            if len(names) != len(self.file_list):
                names = [os.path.basename(x) for x in self.file_list]
            self.file_display_names = names
            self._refresh_input_listbox()
            try:
                self.out_dir_var.set(str(data.get("out_dir", "") or ""))
            except Exception:
                pass
            opts = data.get("options", {}) or {}
            def set_bool(name, key):
                try:
                    if key in opts:
                        getattr(self, name).set(bool(opts.get(key)))
                except Exception:
                    pass
            for name, key in [
                ("chk_group_main", "group_main"), ("chk_shadow_effect", "shadow_effect"),
                ("chk_stiffness", "stiffness"), ("chk_stiff_direct", "stiff_direct"), ("chk_rm", "rm"),
                ("chk_doc", "doc"), ("chk_txt", "txt"), ("chk_excel", "excel"), ("chk_pdf", "pdf"),
                ("chk_col_csv", "col_csv"), ("chk_merge", "merge"), ("chk_print_now", "print_now"),
                ("chk_txt_classic_when_new", "txt_classic_when_new"),
                ("chk_p1", "p1"), ("chk_p2", "p2"), ("chk_p3", "p3"), ("chk_p4", "p4"),
                ("chk_p5", "p5"), ("chk_p6", "p6"), ("chk_p7", "p7"), ("chk_p7_ttgh", "p7_ttgh"), ("chk_p8", "p8"), ("chk_p9", "p9"),
            ]:
                set_bool(name, key)
            try:
                if "p8_opt" in opts:
                    self.var_p8_opt.set(int(opts.get("p8_opt", 1)))
            except Exception:
                pass
            try:
                if opts.get("group_method"):
                    self.cbo_group_method.set(opts.get("group_method"))
            except Exception:
                pass
            try:
                if opts.get("report_form"):
                    self.cbo_report_form.set(opts.get("report_form"))
                    self.report_form_var.set("new" if str(opts.get("report_form")).lower().startswith("new") else "classic")
            except Exception:
                pass
            try:
                if opts.get("printer"):
                    self.cbo_printer.set(opts.get("printer"))
                if opts.get("paper_size"):
                    self.cbo_paper_size.set(opts.get("paper_size"))
            except Exception:
                pass
            info = data.get("report_info", {}) or {}
            for attr, key in [("company_var", "company"), ("department_var", "department"), ("author_var", "author")]:
                try:
                    if key in info:
                        getattr(self, attr).set(info.get(key, ""))
                except Exception:
                    pass
            try:
                self.toggle_group_method(); self.toggle_stiffness(); self.toggle_print_options(); self.toggle_pdf(); self.toggle_report_form()
            except Exception:
                pass

        def _go_to_workflow_step(self, idx: int):
            """Chọn tab theo nút Menu ở thanh trái."""
            try:
                if hasattr(self, "notebook"):
                    self.notebook.select(idx)
                    self._update_workflow_buttons()
            except Exception:
                pass

        def _update_workflow_buttons(self):
            """Cập nhật màu nút Menu theo tab đang mở."""
            pal = getattr(self, "pal", THEME_PRESETS[UI_THEME_AQUA])
            try:
                active_idx = self.notebook.index(self.notebook.select()) if hasattr(self, "notebook") else 0
            except Exception:
                active_idx = 0

            # Style sidebar theo kiểu TS-CAP: nút sáng, có viền, cùng hàng/cùng bề rộng.
            btn_bg = pal.get("button", "#E0F2FE")
            btn_active = pal.get("button_active", "#BAE6FD")
            btn_fg = pal.get("accent_dark", pal.get("accent", "#1D4ED8"))
            border = pal.get("sidebar2", pal.get("border", "#93C5FD"))

            for btn in getattr(self, "file_buttons", []):
                try:
                    btn.configure(
                        bg=btn_bg,
                        fg=btn_fg,
                        activebackground=btn_active,
                        activeforeground=btn_fg,
                        font=("Segoe UI", 9, "bold"),
                        relief=tk.RIDGE,
                        bd=1,
                        cursor="hand2",
                        anchor="w",
                        justify=tk.LEFT,
                        padx=10,
                        pady=7,
                        highlightthickness=1,
                        highlightbackground=border,
                        highlightcolor=border,
                    )
                except Exception:
                    pass

            for i, btn in enumerate(getattr(self, "workflow_buttons", [])):
                label = self.workflow_steps[i][0] if i < len(getattr(self, "workflow_steps", [])) else btn.cget("text")
                is_active = (i == active_idx)
                try:
                    btn.configure(
                        text=label,
                        bg=(pal.get("panel", "white") if is_active else btn_bg),
                        fg=btn_fg,
                        activebackground=btn_active,
                        activeforeground=btn_fg,
                        font=("Segoe UI", 9, "bold" if is_active else "normal"),
                        relief=(tk.SUNKEN if is_active else tk.RIDGE),
                        bd=1,
                        cursor="hand2",
                        anchor="w",
                        justify=tk.LEFT,
                        padx=10,
                        pady=7,
                        highlightthickness=1,
                        highlightbackground=border,
                        highlightcolor=border,
                    )
                except Exception:
                    pass

        def create_widgets(self):
            self.left_frame = tk.Frame(self.root, width=240, bg=self.pal["sidebar"])
            left_frame = self.left_frame
            left_frame.pack(side=tk.LEFT, fill=tk.Y)
            left_frame.pack_propagate(False)

            self.sidebar_title = tk.Label(left_frame, text=APP_NAME, font=("Segoe UI", 19, "bold"), fg="white", bg=self.pal["sidebar"], justify=tk.LEFT)
            self.sidebar_title.pack(anchor=tk.W, padx=20, pady=(26, 18))
            self.sidebar_separator = tk.Frame(left_frame, height=1, bg=self.pal["sidebar2"])
            self.sidebar_separator.pack(fill=tk.X, padx=20, pady=(0, 22))
            tk.Label(left_frame, text="File", font=("Segoe UI", 10, "bold"), fg="white", bg=self.pal["sidebar"]).pack(anchor=tk.W, padx=20, pady=(0, 8))
            self.file_buttons = []
            for label, cmd in [
                ("Open (Ctrl+O)", self.open_project_file),
                ("Save (Ctrl+S)", self.save_project_file),
                ("Save As", self.save_project_file_as),
            ]:
                btn = tk.Button(left_frame, text=label, command=cmd, wraplength=190)
                btn.pack(fill=tk.X, padx=20, pady=4)
                self.file_buttons.append(btn)

            tk.Label(left_frame, text="Menu", font=("Segoe UI", 10, "bold"), fg="white", bg=self.pal["sidebar"]).pack(anchor=tk.W, padx=20, pady=(16, 8))
            # Các nút bên trái liên kết trực tiếp với đúng tab bên phải.
            # Giữ thứ tự này đồng bộ với thứ tự self.notebook.add(...) phía dưới.
            self.workflow_steps = [
                ("Nhập số liệu", 0),
                ("Tính nội lực cọc", 1),
                ("Xuất BC nội lực", 2),
                ("KT Nén uốn", 3),
                ("KT SCT cọc theo đất nền", 4),
            ]
            self.workflow_buttons = []
            for label, tab_idx in self.workflow_steps:
                btn = tk.Button(
                    left_frame,
                    text=label,
                    command=lambda idx=tab_idx: self._go_to_workflow_step(idx),
                    wraplength=190,
                )
                btn.pack(fill=tk.X, padx=20, pady=4)
                self.workflow_buttons.append(btn)

            author_frame = tk.Frame(left_frame, bg=self.pal["sidebar"])
            author_frame.pack(side=tk.BOTTOM, anchor=tk.W, fill=tk.X, padx=20, pady=22)
            author_lines = [line.strip() for line in APP_AUTHOR.split("\n") if line.strip()]
            if author_lines:
                tk.Label(author_frame, text=author_lines[0], font=("Segoe UI", 10, "bold"), fg="#FFF7EA", bg=self.pal["sidebar"], wraplength=210, justify=tk.LEFT).pack(anchor=tk.W, pady=(0, 8))
                for line in author_lines[1:]:
                    tk.Label(author_frame, text=line, font=("Segoe UI", 10), fg="#FFF7EA", bg=self.pal["sidebar"], wraplength=210, justify=tk.LEFT).pack(anchor=tk.W, pady=(2, 2))

            right_frame = tk.Frame(self.root, padx=18, pady=14, bg=self.pal["bg"])
            right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

            self.header_frame = tk.Frame(right_frame, bg=self.pal["bg"])
            header_frame = self.header_frame
            header_frame.pack(fill=tk.X, pady=(0, 12))
            self.title_frame = tk.Frame(header_frame, bg=self.pal["bg"])
            title_frame = self.title_frame
            title_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
            tk.Label(title_frame, text=APP_NAME, font=("Segoe UI", 20, "bold"), fg=self.pal["accent_dark"], bg=self.pal["bg"]).pack(anchor=tk.W)
            self.about_frame = tk.Frame(header_frame, bg=self.pal["bg"])
            about_frame = self.about_frame
            about_frame.pack(side=tk.RIGHT, anchor=tk.NE)
            ttk.Button(about_frame, text="Settings", style="About.TButton", command=self.show_settings).pack(side=tk.LEFT, padx=(0, 6))
            ttk.Button(about_frame, text="License", style="About.TButton", command=self.show_license).pack(side=tk.LEFT, padx=(0, 6))
            ttk.Button(about_frame, text="About", style="About.TButton", command=self.show_about).pack(side=tk.LEFT)
            self.notebook = ttk.Notebook(right_frame)
            self.notebook.pack(fill=tk.BOTH, expand=True)

            tab_data = ttk.Frame(self.notebook)
            tab_algo = ttk.Frame(self.notebook)
            tab_n2d_col = ttk.Frame(self.notebook)
            tab_sct = ttk.Frame(self.notebook)
            tab_report = ttk.Frame(self.notebook)

            self.notebook.add(tab_data, text=' Nhập số liệu ')
            self.notebook.add(tab_algo, text=' Tính nội lực cọc ')
            self.notebook.add(tab_report, text=' Xuất BC nội lực ')
            self.notebook.add(tab_n2d_col, text=' KT Nén uốn ')
            self.notebook.add(tab_sct, text=' KT SCT cọc theo đất nền ')
            self.notebook.bind("<<NotebookTabChanged>>", lambda _e: self._update_workflow_buttons())
            self._update_workflow_buttons()

            self._build_n2d_col_tab(tab_n2d_col)
            self._build_sct_placeholder_tab(tab_sct)

            # ==========================================
            # === TAB 1: DATA ===
            # ==========================================
            lf_input_source = ttk.LabelFrame(tab_data, text="Nguồn số liệu đầu vào")
            lf_input_source.pack(fill=tk.X, pady=(10, 6), padx=10, ipadx=5, ipady=5)
            ttk.Label(
                lf_input_source,
                text="Chọn file/thư mục hoặc nhập trực tiếp bằng các nút bên phải của danh sách dữ liệu.",
                style="Muted.TLabel",
            ).pack(anchor=tk.W, padx=8, pady=(2, 6))

            tk.Label(tab_data, text="Danh sách dữ liệu sẽ tính:", font=("Arial", 9, "bold")).pack(anchor=tk.W, pady=(6, 5), padx=10)
            
            input_frame = tk.Frame(tab_data, padx=10)
            input_frame.pack(fill=tk.BOTH, expand=True)
            self.listbox_files = tk.Listbox(input_frame, height=3) 
            self.listbox_files.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            btn_frame = tk.Frame(input_frame, padx=10)
            btn_frame.pack(side=tk.RIGHT, fill=tk.Y)
            ttk.Button(btn_frame, text="Thêm File", command=self.add_files, width=15).pack(pady=2)
            ttk.Button(btn_frame, text="Thêm Thư mục", command=self.add_folder, width=15).pack(pady=2)
            ttk.Button(btn_frame, text="Xóa file chọn", command=self.delete_selected, width=15).pack(pady=2)
            ttk.Button(btn_frame, text="Xóa toàn bộ", command=self.clear_files, width=15).pack(pady=2)
            ttk.Button(btn_frame, text="Nhập trực tiếp", command=self.open_direct_input_dialog, width=15).pack(pady=(10, 2))

            tk.Label(tab_data, text="Thư mục lưu kết quả xuất:", font=("Arial", 9, "bold")).pack(anchor=tk.W, pady=(15, 5), padx=10)
            
            outf = tk.Frame(tab_data, padx=10)
            outf.pack(fill=tk.X, pady=(0, 10))
            
            self.out_dir_var = tk.StringVar()
            self.entry_out_dir = ttk.Entry(outf, textvariable=self.out_dir_var)
            self.entry_out_dir.pack(side=tk.LEFT, fill=tk.X, expand=True)
            
            out_btn_frame = tk.Frame(outf, padx=10)
            out_btn_frame.pack(side=tk.RIGHT, fill=tk.Y)
            ttk.Button(out_btn_frame, text="Duyệt", command=self.choose_out_dir, width=15).pack(pady=0)

            # ==========================================
            # === TAB 2: ALGORITHM ===
            # ==========================================
            lf_core = ttk.LabelFrame(tab_algo, text="Lựa chọn các thông số tính toán")
            lf_core.pack(fill=tk.X, pady=10, padx=10, ipadx=5, ipady=5)
            
            self.chk_group_main = tk.BooleanVar(value=False)
            self.cb_group_main = ttk.Checkbutton(lf_core, text="Xét đến hiệu ứng nhóm cọc:", variable=self.chk_group_main, command=self.toggle_group_method)
            self.cb_group_main.grid(row=0, column=0, sticky=tk.W, pady=5, padx=10)
            
            self.cbo_group_method = ttk.Combobox(lf_core, values=["Hệ số Poulos", "Hệ số Converse-Labarre", "Móng khối tương đương"], state="disabled", width=25)
            self.cbo_group_method.current(0)
            self.cbo_group_method.grid(row=0, column=1, sticky=tk.W, pady=5)

            self.chk_shadow_effect = tk.BooleanVar(value=False)
            ttk.Checkbutton(lf_core, text="Xét hiệu ứng Bóng râm (P-multiplier)", variable=self.chk_shadow_effect).grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=5, padx=10)

            lf_stiff = ttk.LabelFrame(tab_algo, text="Tính Độ cứng tương đương của móng cọc")
            lf_stiff.pack(fill=tk.X, pady=5, padx=10, ipadx=5, ipady=5)
            
            self.chk_stiffness = tk.BooleanVar(value=False)
            ttk.Checkbutton(lf_stiff, text="Tính độ cứng tương đương của móng cọc (Đã khử liên kết - Mặc định)", variable=self.chk_stiffness, command=self.toggle_stiffness).grid(row=0, column=0, sticky=tk.W, pady=5, padx=10)
            
            self.chk_stiff_direct = tk.BooleanVar(value=False)
            self.cb_stiff_direct = ttk.Checkbutton(lf_stiff, text="Xuất thêm độ cứng trực tiếp (Chưa khử liên kết - Tham khảo)", variable=self.chk_stiff_direct)
            self.cb_stiff_direct.grid(row=1, column=0, sticky=tk.W, pady=2, padx=35)
            
            self.chk_rm = tk.BooleanVar(value=False)
            self.cb_rm = ttk.Checkbutton(lf_stiff, text="Xuất thêm độ cứng quy đổi cho RM Bridge (Đổi trục, đơn vị kN/m)", variable=self.chk_rm)
            self.cb_rm.grid(row=2, column=0, sticky=tk.W, pady=2, padx=35)
            
            self.toggle_stiffness()

            # ==========================================
            # === TAB 3: REPORT ===
            # ==========================================
            lf_format = ttk.LabelFrame(tab_report, text="Định dạng file xuất")
            lf_format.pack(fill=tk.X, pady=5, padx=10, ipadx=5, ipady=5)
            
            self.chk_doc = tk.BooleanVar(value=True); self.chk_txt = tk.BooleanVar(value=False)
            self.chk_excel = tk.BooleanVar(value=False); self.chk_pdf = tk.BooleanVar(value=False)
            self.chk_col_csv = tk.BooleanVar(value=False)
            
            ttk.Checkbutton(lf_format, text=".DOC", variable=self.chk_doc).grid(row=0, column=0, padx=10, pady=2)
            ttk.Checkbutton(lf_format, text=".TXT", variable=self.chk_txt).grid(row=0, column=1, padx=10, pady=2)
            ttk.Checkbutton(lf_format, text=".XLSX", variable=self.chk_excel).grid(row=0, column=2, padx=10, pady=2)
            ttk.Checkbutton(lf_format, text=".PDF", variable=self.chk_pdf, command=self.toggle_pdf).grid(row=0, column=3, padx=10, pady=2)
            
            self.report_form_var = tk.StringVar(value="classic")
            ttk.Label(lf_format, text="Form báo cáo:").grid(row=1, column=0, padx=10, pady=2, sticky=tk.W)
            self.cbo_report_form = ttk.Combobox(lf_format, values=["Form MCOC Classic", "New form"], state="readonly", width=20)
            self.cbo_report_form.current(0)
            self.cbo_report_form.grid(row=1, column=1, columnspan=2, padx=5, pady=2, sticky=tk.W)
            self.cbo_report_form.bind("<<ComboboxSelected>>", self.toggle_report_form)

            self.chk_txt_classic_when_new = tk.BooleanVar(value=False)
            self.cb_txt_classic_when_new = ttk.Checkbutton(
                lf_format,
                text="New form: xuất thêm TXT Classic (_C)",
                variable=self.chk_txt_classic_when_new
            )
            self.cb_txt_classic_when_new.grid(row=2, column=1, columnspan=3, padx=5, pady=2, sticky=tk.W)
            ttk.Checkbutton(lf_format, text='Xuất CSV nội lực cho TS-COL (_TSCOL_INPUT)', variable=self.chk_col_csv).grid(row=3, column=1, columnspan=3, padx=5, pady=2, sticky=tk.W)
            self.toggle_report_form()
            
            self.chk_merge = tk.BooleanVar(value=False)
            self.cb_merge = ttk.Checkbutton(lf_format, text="Gộp chung thành 1 file PDF", variable=self.chk_merge)
            self.cb_merge.grid(row=1, column=3, padx=10, pady=2, sticky=tk.W)
            self.toggle_pdf()

            lf_content = ttk.LabelFrame(tab_report, text="Nội dung thuyết minh (Áp dụng cho xuất DOC, TXT, PDF, EXCEL)")
            lf_content.pack(fill=tk.X, pady=5, padx=10, ipadx=5, ipady=2)
            
            self.chk_p1 = tk.BooleanVar(value=True); self.chk_p2 = tk.BooleanVar(value=True)
            self.chk_p3 = tk.BooleanVar(value=True); self.chk_p4 = tk.BooleanVar(value=True)
            self.chk_p5 = tk.BooleanVar(value=True); self.chk_p6 = tk.BooleanVar(value=True)
            self.chk_p7 = tk.BooleanVar(value=True); self.chk_p7_ttgh = tk.BooleanVar(value=True)
            self.chk_p8 = tk.BooleanVar(value=True); self.chk_p9 = tk.BooleanVar(value=True)
            self.var_p8_opt = tk.IntVar(value=1)

            ttk.Checkbutton(lf_content, text="1. Thông tin cơ bản", variable=self.chk_p1).grid(row=0, column=0, sticky=tk.W, padx=10, pady=2)
            ttk.Checkbutton(lf_content, text="2. Tổ hợp tải trọng", variable=self.chk_p2).grid(row=0, column=1, sticky=tk.W, padx=10, pady=2)
            ttk.Checkbutton(lf_content, text="3. Thông số cọc", variable=self.chk_p3).grid(row=1, column=0, sticky=tk.W, padx=10, pady=2)
            ttk.Checkbutton(lf_content, text="4. Tọa độ cọc", variable=self.chk_p4).grid(row=1, column=1, sticky=tk.W, padx=10, pady=2)
            ttk.Checkbutton(lf_content, text="5. Chuyển vị bệ", variable=self.chk_p5).grid(row=2, column=0, sticky=tk.W, padx=10, pady=2)
            ttk.Checkbutton(lf_content, text="6. Nội lực cọc", variable=self.chk_p6).grid(row=2, column=1, sticky=tk.W, padx=10, pady=2)
            ttk.Checkbutton(lf_content, text="7. Bảng tổng kết chung", variable=self.chk_p7).grid(row=3, column=0, sticky=tk.W, padx=10, pady=2)
            ttk.Checkbutton(lf_content, text="8. Tính toán kiểm tra (P = K * Delta)", variable=self.chk_p8).grid(row=3, column=1, sticky=tk.W, padx=10, pady=2)
            ttk.Checkbutton(lf_content, text="7A. Tổng hợp Max/Min riêng theo TTGH", variable=self.chk_p7_ttgh).grid(row=4, column=0, sticky=tk.W, padx=10, pady=2)
            ttk.Checkbutton(lf_content, text="9. Độ cứng tương đương (nếu có tính)", variable=self.chk_p9).grid(row=5, column=0, sticky=tk.W, padx=10, pady=2)
            
            p8_opt_frame = tk.Frame(lf_content)
            p8_opt_frame.grid(row=4, column=1, rowspan=2, sticky=tk.W, padx=25)
            ttk.Radiobutton(p8_opt_frame, text="8.1 Chỉ in kết quả tính ngược", variable=self.var_p8_opt, value=1).pack(anchor=tk.W)
            ttk.Radiobutton(p8_opt_frame, text="8.2 So sánh chi tiết sai số", variable=self.var_p8_opt, value=2).pack(anchor=tk.W)

            lf_info = ttk.LabelFrame(tab_report, text="Thông tin Tiêu đề / Dự án")
            lf_info.pack(fill=tk.X, pady=5, padx=10, ipadx=5, ipady=2)
            self.company_var = tk.StringVar(value="TẬP ĐOÀN SUNGROUP")
            self.department_var = tk.StringVar(value="KHỐI XD PPP - PHÒNG QLTK")
            self.author_var = tk.StringVar(value='TSDUNGVN - 2026')
            ttk.Label(lf_info, text="Công ty:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=2)
            ttk.Entry(lf_info, textvariable=self.company_var, width=40).grid(row=0, column=1, padx=5, pady=2)
            ttk.Label(lf_info, text="Phòng ban:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=2)
            ttk.Entry(lf_info, textvariable=self.department_var, width=40).grid(row=1, column=1, padx=5, pady=2)
            ttk.Label(lf_info, text="Người lập:").grid(row=2, column=0, sticky=tk.W, padx=10, pady=2)
            ttk.Entry(lf_info, textvariable=self.author_var, width=40).grid(row=2, column=1, padx=5, pady=2)

            # --- BOTTOM FRAME (RUN & PROGRESS) ---
            bottom_frame = tk.Frame(right_frame)
            bottom_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(10, 0))

            print_opt_frame = tk.Frame(bottom_frame)
            print_opt_frame.pack(anchor=tk.W, pady=2)
            
            self.chk_print_now = tk.BooleanVar(value=False)
            ttk.Checkbutton(print_opt_frame, text="In kết quả (Mở hộp thoại In)", variable=self.chk_print_now, command=self.toggle_print_options).pack(side=tk.LEFT, padx=(0, 10))
            
            tk.Label(print_opt_frame, text="Máy in:").pack(side=tk.LEFT, padx=2)
            self.cbo_printer = ttk.Combobox(print_opt_frame, values=get_windows_printers(), state="disabled", width=22)
            self.cbo_printer.set("Default Printer")
            self.cbo_printer.pack(side=tk.LEFT, padx=5)
            
            tk.Label(print_opt_frame, text="Khổ giấy:").pack(side=tk.LEFT, padx=2)
            self.cbo_paper_size = ttk.Combobox(print_opt_frame, values=["A4", "A3", "Letter"], state="disabled", width=8)
            self.cbo_paper_size.set("A4")
            self.cbo_paper_size.pack(side=tk.LEFT, padx=5)

            self.status_var = tk.StringVar(value="Sẵn sàng...")
            self.status_label = tk.Label(bottom_frame, textvariable=self.status_var, fg=self.pal["accent_dark"], bg=self.pal["bg"])
            self.status_label.pack(anchor=tk.W, pady=(5, 0))
            
            self.progress = ttk.Progressbar(bottom_frame, mode='determinate', style="Status.Horizontal.TProgressbar")
            self.progress.pack(fill=tk.X, pady=5)
            
            self.btn_run = tk.Button(bottom_frame, text="TÍNH TOÁN & XUẤT BÁO CÁO", font=("Segoe UI", 12, "bold"), height=2, bg=self.pal["accent"], fg="white", activebackground=self.pal["accent_dark"], activeforeground="white", relief=tk.FLAT, bd=0, command=self.run_calculation)
            self.btn_run.pack(fill=tk.X, pady=(2, 8))

            log_frame = ttk.LabelFrame(bottom_frame, text="Log xử lý")
            log_frame.pack(fill=tk.BOTH, expand=False, pady=(2, 0))
            self.log_text = tk.Text(log_frame, height=6, wrap=tk.WORD, font=("Consolas", 9), bg="#0F172A", fg="#E5E7EB", insertbackground="white", relief=tk.FLAT)
            log_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
            self.log_text.configure(yscrollcommand=log_scroll.set)
            self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(4, 0), pady=4)
            log_scroll.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 4), pady=4)
            self.log_text.config(state=tk.DISABLED)

        def _build_placeholder_panel(self, parent, title, description):
            frame = ttk.Frame(parent, padding=22)
            frame.pack(fill=tk.BOTH, expand=True)
            ttk.Label(frame, text=title, font=("Segoe UI", 16, "bold"), foreground=self.pal["accent_dark"]).pack(anchor=tk.W, pady=(0, 10))
            ttk.Label(frame, text=description, wraplength=720, justify=tk.LEFT).pack(anchor=tk.W, pady=(0, 14))
            ttk.Separator(frame).pack(fill=tk.X, pady=(4, 14))
            ttk.Label(frame, text="Mục này được bố trí sẵn để phát triển tiếp, không ảnh hưởng đến engine tính nội lực hiện tại.", foreground=self.pal["muted"], wraplength=720, justify=tk.LEFT).pack(anchor=tk.W)

        def _build_geo_placeholder_tab(self, parent):
            self._build_placeholder_panel(
                parent,
                "2. Nhập file input địa chất",
                "Khu vực dành cho dữ liệu địa chất, lớp đất và các thông số phục vụ kiểm toán sức chịu tải theo đất nền. Phần này đang để sẵn cho bước phát triển tiếp theo."
            )

        def _build_sct_placeholder_tab(self, parent):
            frame = ttk.Frame(parent, padding=22)
            frame.pack(fill=tk.BOTH, expand=True)
            ttk.Label(frame, text="5. Kiểm toán sức chịu tải theo đất nền", font=("Segoe UI", 16, "bold"), foreground=self.pal["accent_dark"]).pack(anchor=tk.W, pady=(0, 10))
            ttk.Label(
                frame,
                text=('Module này liên kết với TS-CAP để kiểm toán sức chịu tải cọc theo đất nền. '
                      f"{APP_NAME} chỉ mở TS-CAP khi cần dùng, tương tự tab KT Nén uốn, nên chương trình chính vẫn nhẹ và không nạp module SCT lúc khởi động."),
                wraplength=760,
                justify=tk.LEFT
            ).pack(anchor=tk.W, pady=(0, 16))
            btn_row = ttk.Frame(frame)
            btn_row.pack(fill=tk.X, pady=(0, 14))
            ttk.Button(btn_row, text='Mở TS-CAP', style="Big.TButton", command=self._launch_n2d_cap).pack(side=tk.LEFT, padx=(0, 8))
            ttk.Button(btn_row, text='Chọn file TS-CAP...', command=lambda: self._launch_n2d_cap(True)).pack(side=tk.LEFT, padx=(0, 8))
            ttk.Separator(frame).pack(fill=tk.X, pady=(6, 14))
            ttk.Label(
                frame,
                text=("Bộ cài TS Foundation Suite sẽ tự tìm TS-CAP tại thư mục TS_CAP cùng cấp với TS_PILE. "
                      "Khi chạy source vẫn có thể chọn file TS-CAP thủ công."),
                foreground=self.pal["muted"],
                wraplength=760,
                justify=tk.LEFT
            ).pack(anchor=tk.W)

        def _find_n2d_col_script(self):
            """Tìm TS-COL trong bộ TS Foundation Suite hoặc khi chạy source.

            Bản cài chuẩn đặt ba app ở ba thư mục anh em:
              TS_PILE\\TS_PILE.exe
              TS_COL\\TS_COL.exe
              TS_CAP\\TS_CAP.exe
            Vì vậy ưu tiên tuyệt đối ..\\TS_COL\\TS_COL.exe trước khi dùng cơ chế
            dò version/file source cũ. Không nạp TS-COL vào tiến trình TS-PILE.
            """
            base_dirs = []
            for getter in (
                lambda: os.path.dirname(os.path.abspath(__file__)),
                lambda: os.path.dirname(os.path.abspath(sys.executable)),
                lambda: os.getcwd(),
                lambda: resource_path(""),
            ):
                try:
                    d = os.path.abspath(getter())
                    if d and d not in base_dirs:
                        base_dirs.append(d)
                except Exception:
                    pass

            # 1) Bộ cài chuẩn: TS_PILE / TS_COL / TS_CAP là các thư mục cùng cấp.
            preferred = []
            for d in base_dirs:
                parent = os.path.dirname(d)
                preferred.extend([
                    os.path.join(parent, "TS_COL", "TS_COL.exe"),
                    os.path.join(d, "TS_COL", "TS_COL.exe"),
                    os.path.join(d, "TS_COL.exe"),
                ])
            seen = set()
            for pp in preferred:
                pp = os.path.abspath(pp)
                if pp not in seen:
                    seen.add(pp)
                    if os.path.isfile(pp):
                        return pp

            # 2) Chế độ phát triển/source: dò thêm thư mục cùng cấp và tên cũ.
            search_dirs = list(base_dirs)
            for d in list(base_dirs):
                parent = os.path.dirname(d)
                for extra in (os.path.join(parent, "TS_COL"), os.path.join(d, "TS_COL")):
                    extra = os.path.abspath(extra)
                    if extra not in search_dirs:
                        search_dirs.append(extra)

            explicit_names = [
                'TS_COL.exe', 'TS_COL.py',
                'TS_COL_V70_1_20260629_integration_cache_license.py',
                'TS_COL_V70_20260628_fix_argparse_percent_help.py',
                'TS_COL_V70_20260628_fix_argparse_percent_help(1).py',
                'TS_COL_V70.py', "n2d_col_v70.py",
                # tên cũ giữ lại để tương thích:
                'TS_COL_V56_20260628_button_update.py', "n2d_col_v56.py", 'TS_COL_v56.py',
            ]
            patterns = [
                'TS_COL*.exe', "n2d_col*.exe", 'TS_COL*.py', "n2d_col*.py",
                "pile_section_capacity_tcvn11823_aashto_v*.exe",
                "pile_section_capacity_tcvn11823_aashto_v*.py",
            ]
            candidates = []
            seen = set()
            for d in search_dirs:
                for name in explicit_names:
                    pp = os.path.abspath(os.path.join(d, name))
                    if pp not in seen:
                        seen.add(pp)
                        candidates.append(pp)
                for pat in patterns:
                    for pp in glob.glob(os.path.join(d, pat)):
                        pp = os.path.abspath(pp)
                        if pp not in seen:
                            seen.add(pp)
                            candidates.append(pp)

            def score(path):
                base = os.path.basename(path).lower()
                # Ưu tiên exe để bản đóng gói không vô tình mở source .py bên cạnh.
                is_exe = 1 if base.endswith(".exe") else 0
                exact_exe = 1 if base == "ts_col.exe" else 0
                v = 0
                m = re.search(r"v[_-]?(\d+)", base)
                if m:
                    try:
                        v = int(m.group(1))
                    except Exception:
                        v = 0
                date_score = 0
                d = re.search(r"(20\d{6})", base)
                if d:
                    try:
                        date_score = int(d.group(1))
                    except Exception:
                        date_score = 0
                try:
                    mtime = os.path.getmtime(path)
                except Exception:
                    mtime = 0
                return (is_exe, exact_exe, v, date_score, mtime)

            existing = [pp for pp in candidates if os.path.isfile(pp)]
            return max(existing, key=score) if existing else None

        def _launch_n2d_col(self, force_choose=False):
            script = None if force_choose else self._find_n2d_col_script()
            if force_choose or not script:
                script = filedialog.askopenfilename(
                    title='Chọn file TS-COL',
                    filetypes=[("Python file", "*.py"), ("Executable", "*.exe"), ("All files", "*.*")]
                )
            if not script:
                return
            try:
                if script.lower().endswith(".py"):
                    subprocess.Popen([sys.executable, script], cwd=os.path.dirname(script) or None)
                else:
                    subprocess.Popen([script], cwd=os.path.dirname(script) or None)
                self._log(f"Đã mở TS-COL: {os.path.basename(script)}", "OK")
            except Exception as exc:
                messagebox.showerror('Không mở được TS-COL', f"Không mở được file:\n{script}\n\nLỗi: {exc}")

        def _find_n2d_cap_script(self):
            """Tìm TS-CAP trong bộ TS Foundation Suite hoặc khi chạy source.

            Ưu tiên ..\\TS_CAP\\TS_CAP.exe của bộ cài chung; sau đó mới dò các
            tên/version cũ để giữ tương thích khi phát triển bằng file .py.
            """
            base_dirs = []
            for getter in (
                lambda: os.path.dirname(os.path.abspath(__file__)),
                lambda: os.path.dirname(os.path.abspath(sys.executable)),
                lambda: os.getcwd(),
                lambda: resource_path(""),
            ):
                try:
                    d = os.path.abspath(getter())
                    if d and d not in base_dirs:
                        base_dirs.append(d)
                except Exception:
                    pass

            # 1) Bộ cài chuẩn: tìm executable ở thư mục TS_CAP cùng cấp TS_PILE.
            preferred = []
            for d in base_dirs:
                parent = os.path.dirname(d)
                preferred.extend([
                    os.path.join(parent, "TS_CAP", "TS_CAP.exe"),
                    os.path.join(d, "TS_CAP", "TS_CAP.exe"),
                    os.path.join(d, "TS_CAP.exe"),
                ])
            seen = set()
            for pp in preferred:
                pp = os.path.abspath(pp)
                if pp not in seen:
                    seen.add(pp)
                    if os.path.isfile(pp):
                        return pp

            # 2) Chế độ source/phát triển.
            search_dirs = list(base_dirs)
            for d in list(base_dirs):
                parent = os.path.dirname(d)
                for extra in (os.path.join(parent, "TS_CAP"), os.path.join(d, "TS_CAP")):
                    extra = os.path.abspath(extra)
                    if extra not in search_dirs:
                        search_dirs.append(extra)

            explicit_names = [
                'TS_CAP.exe', 'TS_CAP.py',
                'TS_CAP_V1_0_20260704.py',
                'TS_CAP_V1_0_66_20260703_fatal_validation_warning_panel.py',
                "n2d_cap.py",
            ]
            patterns = [
                'TS_CAP*.exe', 'TS-CAP*.exe', "n2d_cap*.exe",
                'TS_CAP*.py', 'TS-CAP*.py', "n2d_cap*.py",
            ]
            candidates = []
            seen = set()
            for d in search_dirs:
                for name in explicit_names:
                    pp = os.path.abspath(os.path.join(d, name))
                    if pp not in seen:
                        seen.add(pp)
                        candidates.append(pp)
                for pat in patterns:
                    for pp in glob.glob(os.path.join(d, pat)):
                        pp = os.path.abspath(pp)
                        if pp not in seen:
                            seen.add(pp)
                            candidates.append(pp)

            def score(path):
                base = os.path.basename(path).lower()
                is_exe = 1 if base.endswith(".exe") else 0
                exact_exe = 1 if base == "ts_cap.exe" else 0
                m = re.search(r"v[_-]?([0-9][0-9_\.\-]*)", base)
                version_nums = []
                if m:
                    version_nums = [int(x) for x in re.findall(r"\d+", m.group(1))[:4]]
                version_nums = (version_nums + [0, 0, 0, 0])[:4]
                date_score = 0
                d = re.search(r"(20\d{6})", base)
                if d:
                    try:
                        date_score = int(d.group(1))
                    except Exception:
                        date_score = 0
                try:
                    mtime = os.path.getmtime(path)
                except Exception:
                    mtime = 0
                return (is_exe, exact_exe, *version_nums, date_score, mtime)

            existing = [pp for pp in candidates if os.path.isfile(pp)]
            return max(existing, key=score) if existing else None

        def _launch_n2d_cap(self, force_choose=False):
            script = None if force_choose else self._find_n2d_cap_script()
            if force_choose or not script:
                script = filedialog.askopenfilename(
                    title='Chọn file TS-CAP',
                    filetypes=[("Python file", "*.py"), ("Executable", "*.exe"), ("All files", "*.*")]
                )
            if not script:
                return
            try:
                if script.lower().endswith(".py"):
                    subprocess.Popen([sys.executable, script], cwd=os.path.dirname(script) or None)
                else:
                    subprocess.Popen([script], cwd=os.path.dirname(script) or None)
                self._log(f"Đã mở TS-CAP: {os.path.basename(script)}", "OK")
            except Exception as exc:
                messagebox.showerror('Không mở được TS-CAP', f"Không mở được file:\n{script}\n\nLỗi: {exc}")

        def _build_n2d_col_tab(self, parent):
            frame = ttk.Frame(parent, padding=22)
            frame.pack(fill=tk.BOTH, expand=True)
            ttk.Label(frame, text="4. Kiểm toán nén uốn cọc", font=("Segoe UI", 16, "bold"), foreground=self.pal["accent_dark"]).pack(anchor=tk.W, pady=(0, 10))
            ttk.Label(
                frame,
                text=('Module này liên kết dữ liệu với TS-COL để kiểm toán nén-uốn mặt cắt cọc BTCT. '
                      f"{APP_NAME} chỉ mở TS-COL khi cần dùng, nên chương trình chính vẫn nhẹ và không nạp các thư viện tính mặt cắt lúc khởi động."),
                wraplength=760,
                justify=tk.LEFT
            ).pack(anchor=tk.W, pady=(0, 16))
            btn_row = ttk.Frame(frame)
            btn_row.pack(fill=tk.X, pady=(0, 14))
            ttk.Button(btn_row, text='Mở TS-COL', style="Big.TButton", command=self._launch_n2d_col).pack(side=tk.LEFT, padx=(0, 8))
            ttk.Button(btn_row, text='Chọn file TS-COL...', command=lambda: self._launch_n2d_col(True)).pack(side=tk.LEFT, padx=(0, 8))
            ttk.Button(btn_row, text="Mở CSV Col gần nhất", command=self._open_last_col_csv).pack(side=tk.LEFT)
            ttk.Separator(frame).pack(fill=tk.X, pady=(6, 14))
            ttk.Label(
                frame,
                text=("Bộ cài TS Foundation Suite sẽ tự tìm TS-COL tại thư mục TS_COL cùng cấp với TS_PILE. "
                      "Khi chạy source vẫn có thể chọn file TS-COL thủ công."),
                foreground=self.pal["muted"],
                wraplength=760,
                justify=tk.LEFT
            ).pack(anchor=tk.W)

        def _open_last_col_csv(self):
            if not self.last_col_csv or not os.path.exists(self.last_col_csv):
                messagebox.showinfo("Chưa có CSV", 'Chưa có file CSV nội lực cho TS-COL. Hãy chạy tính toán trước hoặc bật lựa chọn xuất CSV TS-COL.')
                return
            try:
                if platform.system() == "Windows":
                    os.startfile(self.last_col_csv)
                else:
                    subprocess.Popen(["xdg-open", self.last_col_csv])
            except Exception as exc:
                messagebox.showerror("Không mở được CSV", f"Không mở được file:\n{self.last_col_csv}\n\nLỗi: {exc}")


        # ==========================================
        # NHẬP SỐ LIỆU TRỰC TIẾP - TẠO FILE INPUT MCOC TỪ BẢNG
        # ==========================================
        def _safe_float(self, value, default=0.0):
            try:
                s = str(value).strip().replace(",", ".")
                if s == "":
                    return float(default)
                return float(s)
            except Exception:
                return float(default)

        def _sanitize_filename(self, name):
            s = str(name or "input").strip()
            # Giữ tên ngắn gọn, cho phép khoảng trắng/dấu tiếng Việt/ngoặc; chỉ thay ký tự cấm trong tên file Windows.
            s = re.sub(r'[<>:"/\\|?*]+', "_", s, flags=re.UNICODE)
            s = re.sub(r"\s+", " ", s, flags=re.UNICODE).strip(" ._")
            return s or "input"

        def _next_direct_input_name(self, project_name):
            base = str(project_name or "").strip() or "input"
            base = re.sub(r"\s+", " ", base, flags=re.UNICODE).strip() or "input"
            existing = set(str(x) for x in getattr(self, "file_display_names", []) or [])
            try:
                out_dir = self._default_direct_input_dir()
            except Exception:
                out_dir = tempfile.gettempdir()
            candidate = base
            i = 1
            while True:
                fp = os.path.join(out_dir, self._sanitize_filename(candidate) + ".txt")
                if candidate not in existing and not os.path.exists(fp):
                    return candidate
                candidate = f"{base}({i})"
                i += 1

        def _default_direct_input_dir(self):
            try:
                if self.out_dir_var.get().strip():
                    return self.out_dir_var.get().strip()
            except Exception:
                pass
            docs = os.path.join(os.path.expanduser("~"), "Documents")
            base = docs if os.path.isdir(docs) else tempfile.gettempdir()
            return os.path.join(base, 'TS_PILE_Direct_Input')

        def _write_direct_input_file(self, project_id, global_values, combo_rows, pile_rows):
            out_dir = self._default_direct_input_dir()
            os.makedirs(out_dir, exist_ok=True)
            base = self._sanitize_filename(project_id)
            fp = os.path.join(out_dir, f"{base}.txt")
            if os.path.exists(fp):
                i = 1
                while True:
                    fp_try = os.path.join(out_dir, f"{base}({i}).txt")
                    if not os.path.exists(fp_try):
                        fp = fp_try
                        break
                    i += 1

            n_piles = len(pile_rows)
            n_combos = len(combo_rows)
            # Format 18 số đầu theo parser MCOC hiện tại:
            # [0] n_piles, [1] n_combos, [2:6] dự phòng/legacy, [6] Kn, [7] dự phòng,
            # [8] Bx, [9] By, [10] Cz, [11] EI_uon, [12] Er_uon, [13] EA_nen, [14] Er_nen,
            # [15] md, [16] mq, [17] m.
            header_nums = [
                # Giữ đúng trật tự cột A sheet Output của file Data MCOC.xls:
                # A2 n_cọc, A3 n_tổ_hợp, A4=0, A5=2, A6=0, A7=n_cọc,
                # A8 Kn, A9=0, A10 Bx, A11 By, A12 Cz, A13 Ev_uốn,
                # A14=0, A15 Ev_nén, A16=0, A17 md, A18 mq, A19 m.
                n_piles, n_combos, 0, global_values.get("Legacy_Header_Flag", 2),
                global_values.get("Legacy_Reserved_1", 0),
                global_values.get("Pile_Count_Copy", n_piles),
                global_values.get("Kn", 1.0), global_values.get("Legacy_Reserved_2", 0),
                global_values.get("Bx", 0.0), global_values.get("By", 0.0), global_values.get("Cz", 0.0),
                global_values.get("EI_uon", 0.0), global_values.get("Er_uon", 0.0),
                global_values.get("EA_nen", 0.0), global_values.get("Er_nen", 0.0),
                global_values.get("md", 0.0), global_values.get("mq", 0.0), global_values.get("m", 0.0),
            ]
            def fmt(x, decimals=None, strip=True):
                try:
                    xf = float(x)
                    if decimals is None:
                        if abs(xf - round(xf)) < 1e-10:
                            return str(int(round(xf)))
                        return f"{xf:.10g}"
                    s = f"{xf:.{int(decimals)}f}"
                    if strip and int(decimals) > 0:
                        s = s.rstrip("0").rstrip(".")
                    if s == "-0":
                        s = "0"
                    return s
                except Exception:
                    return str(x)

            def fmt_header(i, x):
                # 18 số đầu: E/Ev không lấy thập phân; kích thước và hệ số nền làm gọn.
                if i in (11, 12, 13, 14):
                    return fmt(x, 0, strip=False)
                if i in (8, 9, 10, 15, 16, 17):
                    return fmt(x, 3)
                return fmt(x, 0, strip=False)

            def fmt_pile_value(key, x):
                # Fo/Io giữ 12 chữ số để không mất độ cứng của cọc nhỏ.
                if key in ("Area", "J_xy"):
                    return fmt(x, 12, strip=True)
                if key in ("Bpx", "Bpy", "d_ngoai", "d_trong", "day_vo", "Co", "Ct", "Phi"):
                    return fmt(x, 3)
                if key in ("Lo", "H", "Po", "X", "Y", "Xi"):
                    return fmt(x, 6)
                return fmt(x)

            # Ghi theo đúng bố cục file INPUT MCOC mẫu:
            # - Dòng 1: tên công trình.
            # - Dòng 2: 11 số đầu của header: n_cọc ... Cz.
            # - Dòng 3: 7 số còn lại của header: Ev_uốn, Er_uốn, Ev_nén, Er_nén, md, mq, m.
            # - Mỗi tổ hợp tải trọng: 1 dòng gồm 6 số.
            # - Mỗi cọc: 16 thông số, mỗi thông số một dòng, đúng thứ tự parser MCOC đang đọc.
            with open(fp, "w", encoding="utf-8") as f:
                f.write(str(project_id or "TS_DIRECT_INPUT").strip() + "\n")
                f.write(" ".join(fmt_header(i, x) for i, x in enumerate(header_nums[:11])) + "\n")
                f.write(" ".join(fmt_header(i + 11, x) for i, x in enumerate(header_nums[11:])) + "\n")
                for row in combo_rows:
                    vals = [row.get(k, 0.0) for k in ("Hx", "Hy", "N_load", "Mx", "My", "Mz")]
                    f.write(" ".join(fmt(x) for x in vals) + "\n")
                pile_keys = ("Lo", "H", "Bpx", "Bpy", "d_ngoai", "d_trong", "day_vo", "Area", "J_xy", "Po", "Co", "Ct", "X", "Y", "Phi", "Xi")
                for row in pile_rows:
                    for k in pile_keys:
                        f.write(fmt_pile_value(k, row.get(k, 0.0)) + "\n")
            return fp

        def _parse_numeric_table_text(self, text, min_cols=6):
            rows = []
            for line in str(text or "").splitlines():
                raw = line.strip()
                if not raw or raw.startswith("#"):
                    continue
                # Bỏ chữ, giữ các số; hỗ trợ số âm và số thập phân kiểu 1.23 hoặc 1,23.
                nums = []
                for m in re.finditer(r"[-+]?\d+(?:[\.,]\d+)?(?:[eE][-+]?\d+)?", raw):
                    try:
                        nums.append(float(m.group(0).replace(",", ".")))
                    except Exception:
                        pass
                if len(nums) >= min_cols:
                    rows.append(nums)
            return rows

        def _calc_area_j_from_geometry(self, row_dict):
            try:
                A = float(row_dict.get("d_ngoai", 0.0))
                B = float(row_dict.get("d_trong", 0.0))
                t = float(row_dict.get("day_vo", 0.0))
                if abs(t) < 1e-12:
                    # Cọc tròn/ống tròn: d_ngoai=D, d_trong=d_trong. Nếu d_trong=0 là cọc đặc.
                    D = max(A, 0.0)
                    d = B if 0.0 < B < D else 0.0
                    area = math.pi * (D**2 - d**2) / 4.0
                    j = math.pi * (D**4 - d**4) / 64.0
                else:
                    # Cọc chữ nhật/vuông rỗng: d_ngoai=A, d_trong=B, day_vo=t. Nếu t không hợp lệ thì xem là đặc.
                    bw = max(A, 0.0)
                    h = max(B, 0.0)
                    if t > 0 and 2*t < bw and 2*t < h:
                        bi = bw - 2*t
                        hi = h - 2*t
                        area = bw*h - bi*hi
                        # Dùng moment quán tính trung bình để phù hợp biến J_xy một giá trị trong MCOC.
                        ix = (bw*h**3 - bi*hi**3) / 12.0
                        iy = (h*bw**3 - hi*bi**3) / 12.0
                        j = 0.5 * (ix + iy)
                    else:
                        area = bw*h
                        ix = bw*h**3 / 12.0
                        iy = h*bw**3 / 12.0
                        j = 0.5 * (ix + iy)
                return max(area, 0.0), max(j, 0.0)
            except Exception:
                return 0.0, 0.0

        def _direct_ec_from_concrete_mark(self, mark_mpa):
            """Mô đun E bê tông theo bảng Excel Data MCOC của anh.
            Đơn vị trả về: T/m². Giữ đúng các mốc đang dùng trong file Excel:
            M250 = 2.90e6, M300 = 3.15e6, M350 = 3.325e6, M400 = 3.50e6.
            """
            try:
                m = float(mark_mpa)
            except Exception:
                return 3150000.0
            pts = [(250.0, 2900000.0), (300.0, 3150000.0), (350.0, 3325000.0), (400.0, 3500000.0)]
            if m <= pts[0][0]:
                return pts[0][1]
            if m >= pts[-1][0]:
                return pts[-1][1]
            for (x0, y0), (x1, y1) in zip(pts[:-1], pts[1:]):
                if x0 <= m <= x1:
                    t = (m - x0) / max(x1 - x0, 1e-12)
                    return y0 + t * (y1 - y0)
            return 3150000.0

        def _direct_ec_tcvn11823_tfm2(self, fc_mpa, k1=1.0, wc_kg_m3=2450.0):
            """Mô đun đàn hồi bê tông theo TCVN 11823-5, Eq. Ec = 0.0017*K1*wc^2*f'c^0.33.

            Trả về đơn vị T/m² để đưa trực tiếp vào engine MCOC.
            Quy đổi: 1 MPa = 1000 kN/m² = 1000/9.80665 T/m².
            """
            try:
                fc = float(fc_mpa)
                if fc <= 0:
                    return 0.0
                ec_mpa = 0.0017 * float(k1) * (float(wc_kg_m3) ** 2.0) * (fc ** 0.33)
                return ec_mpa * (1000.0 / 9.80665)
            except Exception:
                return 0.0

        def _direct_bpx_bpy_from_excel_rule(self, pile_type, d):
            """Bpx/Bpy theo công thức trong sheet Output của file Data MCOC.xls."""
            try:
                typ = int(round(float(pile_type)))
                d = float(d)
            except Exception:
                return 0.0
            if d <= 0:
                return 0.0
            if typ == 0:  # cọc đóng, tiết diện vuông trong template Excel
                return 1.5 * d + 0.5
            # cọc khoan nhồi
            return 2.25 * d if d < 0.9 else 0.9 * (d + 1.0)

        def _direct_co_ct_from_excel_rule(self, h_embed, m_pile_tip, d):
            """Co/Ct theo công thức trong file Excel:
            Co = IF(H<=10, 50*m/D, 5*m*H/D), Ct = Co/2.
            """
            try:
                h = float(h_embed)
                mp = float(m_pile_tip)
                d = float(d)
                if d <= 0:
                    return 0.0, 0.0
                co = (50.0 * mp / d) if h <= 10.0 else (5.0 * mp * h / d)
                return co, co / 2.0
            except Exception:
                return 0.0, 0.0

        def _direct_section_props_from_excel_rule(self, pile_type, d):
            """Trả về d_ngoai, d_trong, day_vo, Area, J_xy theo đúng mapping Output sheet.
            pile_type=1: cọc khoan nhồi tròn đặc.
            pile_type=0: cọc đóng tiết diện vuông đặc.
            """
            try:
                typ = int(round(float(pile_type)))
                d = float(d)
            except Exception:
                return 0.0, 0.0, 0.0, 0.0, 0.0
            if d <= 0:
                return 0.0, 0.0, 0.0, 0.0, 0.0
            if typ == 1:
                return d, 0.0, 0.0, math.pi * d * d / 4.0, math.pi * d**4 / 64.0
            return d, d, d / 2.0, d * d, d**4 / 12.0

        def open_direct_input_dialog(self):
            'Cửa sổ nhập trực tiếp theo đúng logic Data MCOC, nhưng nhập bảng lớn kiểu TS-COL.\n\n            - Tổ hợp tải trọng và tọa độ cọc dùng bảng Treeview, paste được hàng trăm dòng từ Excel/Clipboard.\n            - Các đại lượng có công thức trong sheet Output được tự tính, chỉ để người dùng nhập dữ kiện gốc.\n            '
            win = tk.Toplevel(self.root)
            win.title('Nhập số liệu trực tiếp cho TS-PILE - bảng dữ liệu')
            win.geometry("1120x720")
            win.minsize(980, 620)
            try:
                win.transient(self.root)
                win.grab_set()
            except Exception:
                pass

            pal = getattr(self, "pal", THEME_PRESETS[UI_THEME_AQUA])
            win.configure(bg=pal.get("bg", APP_BG))
            main = ttk.Frame(win, padding=6)
            main.pack(fill=tk.BOTH, expand=True)

            header = ttk.Frame(main)
            header.pack(fill=tk.X, pady=(0, 8))
            ttk.Label(header, text="Nhập số liệu trực tiếp", style="Header.TLabel").pack(side=tk.LEFT)
            ttk.Label(
                header,
                text="Tổ hợp tải trọng và tọa độ cọc dạng bảng lớn: có thể paste trực tiếp từ Excel/Clipboard",
                style="SubHeader.TLabel",
            ).pack(side=tk.RIGHT)

            nb = ttk.Notebook(main)
            nb.pack(fill=tk.BOTH, expand=True)
            tab_general = ttk.Frame(nb, padding=6)
            tab_load = ttk.Frame(nb, padding=6)
            tab_piles = ttk.Frame(nb, padding=6)
            tab_preview = ttk.Frame(nb, padding=6)
            nb.add(tab_general, text=" 1. Thông số chung ")
            nb.add(tab_load, text=" 2. Tổ hợp tải trọng ")
            nb.add(tab_piles, text=" 3. Tọa độ cọc ")
            nb.add(tab_preview, text=" 4. Xem trước INPUT ")

            def fvar(var, default=0.0):
                try:
                    return float(str(var.get()).strip().replace(",", "."))
                except Exception:
                    return float(default)

            def fmt_num(x, decimals=None, strip=True):
                """Format số gọn cho giao diện nhập trực tiếp.
                decimals=None: giữ dạng ngắn; decimals=N: làm tròn N chữ số.
                """
                try:
                    xf = float(x)
                    if decimals is None:
                        if abs(xf - round(xf)) < 1e-10:
                            return str(int(round(xf)))
                        return f"{xf:.6g}"
                    s = f"{xf:.{int(decimals)}f}"
                    if strip and int(decimals) > 0:
                        s = s.rstrip("0").rstrip(".")
                    if s == "-0":
                        s = "0"
                    return s
                except Exception:
                    return "" if x is None else str(x)

            # Dữ liệu gốc người dùng cần nhập. Các thông số khác được suy ra theo sheet Output.
            project_var = tk.StringVar(value="")
            pile_type_var = tk.StringVar(value="1 - Cọc khoan nhồi tròn")      # 0: cọc đóng vuông, 1: cọc khoan nhồi tròn
            pile_d_var = tk.StringVar(value="")
            pile_l_var = tk.StringVar(value="")
            pile_lo_var = tk.StringVar(value="")
            kn_var = tk.StringVar(value="")
            sct_pile_var = tk.StringVar(value="")
            m_pile_tip_var = tk.StringVar(value="")
            bx_var = tk.StringVar(value="")
            by_var = tk.StringVar(value="")
            cz_var = tk.StringVar(value="")
            fc_var = tk.StringVar(value="")
            md_var = tk.StringVar(value="")
            mq_var = tk.StringVar(value="")
            m_work_var = tk.StringVar(value="")
            manual_other_var = tk.BooleanVar(value=False)  # QA UI: tick để người dùng tự nhập các thông số khác.

            # Biến tự tính / chỉ hiển thị.
            h_embed_var = tk.StringVar(value="")
            bpx_var = tk.StringVar(value="")
            bpy_var = tk.StringVar(value="")
            area_var = tk.StringVar(value="")
            jxy_var = tk.StringVar(value="")
            co_var = tk.StringVar(value="")
            ct_var = tk.StringVar(value="")
            ec_var = tk.StringVar(value="")
            e_axial_var = tk.StringVar(value="")
            e_bending_var = tk.StringVar(value="")

            E_BENDING_FACTOR = 0.80  # giữ đúng số chết ở sheet Output: Ev uốn = 0.8*E

            def derived_common_values(update_vars=True):
                pile_type = _parse_choice_int(pile_type_var.get(), 1)
                # Đường kính/cạnh cọc nhập trực tiếp luôn dùng mét.
                d = fvar(pile_d_var, 0.0)
                L = fvar(pile_l_var, 0.0)
                Lo = fvar(pile_lo_var, 0.0)
                H = max(L - Lo, 0.0)
                fc = fvar(fc_var, 0.0)

                ec_tfm2 = self._direct_ec_tcvn11823_tfm2(fc)
                e_axial = ec_tfm2
                # Trong TCVN/AASHTO chỉ có Ec vật liệu. File Data MCOC của anh đang giảm E uốn bằng hệ số 0.8.
                # Bản này giữ 0.8 để tái tạo đúng INPUT MCOC từ Excel; nếu muốn dùng Ec nguyên thì đổi E_BENDING_FACTOR = 1.0.
                e_bending = E_BENDING_FACTOR * ec_tfm2

                bpx = self._direct_bpx_bpy_from_excel_rule(pile_type, d)
                bpy = bpx
                d_ngoai, d_trong, day_vo, area, jxy = self._direct_section_props_from_excel_rule(pile_type, d)
                co, ct = self._direct_co_ct_from_excel_rule(H, fvar(m_pile_tip_var, 0.0), d)
                auto_common = {
                    "Lo": Lo, "H": H,
                    "Bpx": bpx, "Bpy": bpy,
                    "d_ngoai": d_ngoai, "d_trong": d_trong, "day_vo": day_vo,
                    "Area": area, "J_xy": jxy,
                    "Po": fvar(sct_pile_var, 0.0), "Co": co, "Ct": ct,
                    "EA_nen": e_axial, "EI_uon": e_bending,
                    "Ec": ec_tfm2,
                }
                if bool(manual_other_var.get()):
                    # Khi tick "Tự nhập", dùng trực tiếp các giá trị người dùng gõ trong khung "Các thông số khác".
                    # Không ghi đè lại các ô này khi người dùng sửa dữ kiện gốc bên trái.
                    common = dict(auto_common)
                    common.update({
                        "H": fvar(h_embed_var, auto_common["H"]),
                        "Bpx": fvar(bpx_var, auto_common["Bpx"]),
                        "Bpy": fvar(bpy_var, auto_common["Bpy"]),
                        "Area": fvar(area_var, auto_common["Area"]),
                        "J_xy": fvar(jxy_var, auto_common["J_xy"]),
                        "Co": fvar(co_var, auto_common["Co"]),
                        "Ct": fvar(ct_var, auto_common["Ct"]),
                        "Ec": fvar(ec_var, auto_common["Ec"]),
                        "EA_nen": fvar(e_axial_var, auto_common["EA_nen"]),
                        "EI_uon": fvar(e_bending_var, auto_common["EI_uon"]),
                    })
                else:
                    common = auto_common
                    if update_vars:
                        h_embed_var.set(fmt_num(H, 3))
                        bpx_var.set(fmt_num(bpx, 3))
                        bpy_var.set(fmt_num(bpy, 3))
                        area_var.set(fmt_num(area, 6, strip=False))
                        jxy_var.set(fmt_num(jxy, 6, strip=False))
                        co_var.set(fmt_num(co, 3))
                        ct_var.set(fmt_num(ct, 3))
                        ec_var.set(fmt_num(ec_tfm2, 0, strip=False))
                        e_axial_var.set(fmt_num(e_axial, 0, strip=False))
                        e_bending_var.set(fmt_num(e_bending, 0, strip=False))
                return common

            # --- Tab thông số chung ---
            form_wrap = ttk.Frame(tab_general)
            form_wrap.pack(fill=tk.BOTH, expand=True)
            left = ttk.LabelFrame(form_wrap, text="Dữ kiện nhập tay", padding=6)
            left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))
            right = ttk.LabelFrame(form_wrap, text="Các thông số khác", padding=6)
            right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0))
            # Căn form nhập tay/tự tính: cột nhãn đủ rộng, ô nhập đồng đều và dịch sang phải vừa phải.
            left._ts_label_col_minsize = 275
            right._ts_label_col_minsize = 325
            left._ts_entry_width = 24
            right._ts_entry_width = 24

            def add_row(parent, r, label, var, width=18, values=None, readonly=False, note=""):
                # label có thể là chuỗi thường hoặc list [(text, kind)] để hiển thị chỉ số dưới.
                # Chỉ dùng ở giao diện nhập tay; không ảnh hưởng mapping dữ liệu/tính toán.
                entry_width = int(getattr(parent, "_ts_entry_width", width))
                try:
                    parent.columnconfigure(0, minsize=int(getattr(parent, "_ts_label_col_minsize", 0)), weight=0)
                    parent.columnconfigure(1, weight=0)
                    parent.columnconfigure(2, weight=1)
                except Exception:
                    pass
                if isinstance(label, (list, tuple)) and label and isinstance(label[0], (list, tuple)):
                    label_box = ttk.Frame(parent)
                    for part in label:
                        txt = str(part[0] if len(part) > 0 else "")
                        kind = str(part[1] if len(part) > 1 and part[1] is not None else "")
                        lbl = ttk.Label(label_box, text=txt)
                        if kind == "sub":
                            try:
                                lbl.configure(font=("Segoe UI", 8))
                            except Exception:
                                pass
                            # Giữ chỉ số dưới sát ký hiệu chính hơn, tránh tụt quá xa.
                            lbl.pack(side=tk.LEFT, anchor="s", pady=(1, 0))
                        elif kind == "sup":
                            try:
                                lbl.configure(font=("Segoe UI", 8))
                            except Exception:
                                pass
                            lbl.pack(side=tk.LEFT, anchor="n", pady=(0, 1))
                        else:
                            lbl.pack(side=tk.LEFT, anchor="s", pady=(0, 0))
                    label_box.grid(row=r, column=0, sticky="w", padx=(4, 14), pady=3)
                else:
                    ttk.Label(parent, text=label).grid(row=r, column=0, sticky="w", padx=(4, 14), pady=3)
                if values is not None:
                    w = ttk.Combobox(parent, textvariable=var, values=values, width=entry_width, state="readonly" if readonly else "normal")
                else:
                    w = ttk.Entry(parent, textvariable=var, width=entry_width, state="readonly" if readonly else "normal")
                w.grid(row=r, column=1, sticky="w", padx=(12, 4), pady=3)
                if note:
                    ttk.Label(parent, text=note, style="Muted.TLabel").grid(row=r, column=2, sticky="w", padx=(8, 4), pady=3)
                return w

            r = 0
            add_row(left, r, "Tên công trình", project_var, 22); r += 1
            add_row(left, r, "Loại cọc", pile_type_var, 24, values=["1 - Cọc khoan nhồi tròn", "0 - Cọc đóng vuông"], readonly=True); r += 1
            add_row(left, r, "Đường kính/cạnh cọc D (m)", pile_d_var, 14); r += 1
            add_row(left, r, "Chiều dài cọc L (m)", pile_l_var, 14); r += 1
            add_row(left, r, "Chiều dài tự do Lo (m)", pile_lo_var, 14); r += 1
            add_row(left, r, "Hệ số tỷ lệ nền Kn", kn_var, 14, values=["0", "1"], readonly=True); r += 1
            add_row(left, r, "SCT cọc theo đất nền P0 (T)", sct_pile_var, 14); r += 1
            add_row(left, r, "Cường độ bê tông f'c (MPa)", fc_var, 14); r += 1
            add_row(left, r, "Chiều dài bệ Bx - dọc cầu (m)", bx_var, 14); r += 1
            add_row(left, r, "Chiều rộng bệ By - ngang cầu (m)", by_var, 14); r += 1
            add_row(left, r, "Chiều cao bệ Cz (m)", cz_var, 14); r += 1
            add_row(left, r, [("Hệ số nền đáy bệ m", None), ("d", "sub"), (" (t/m⁴)", None)], md_var, 14); r += 1
            add_row(left, r, [("Hệ số nền quanh bệ m", None), ("q", "sub"), (" (T/m⁴)", None)], mq_var, 14); r += 1
            add_row(left, r, "Hệ số nền khu vực cọc làm việc m (T/m⁴)", m_work_var, 14); r += 1
            add_row(left, r, [("Hệ số nền tại mũi cọc m", None), ("c", "sub"), (" (T/m⁴)", None)], m_pile_tip_var, 14); r += 1

            right_input_widgets = []

            def toggle_manual_other_state(*_):
                manual = bool(manual_other_var.get())
                for widget in list(right_input_widgets):
                    try:
                        widget.configure(state=(tk.NORMAL if manual else "readonly"))
                    except Exception:
                        pass
                if not manual:
                    derived_common_values(True)

            r = 0
            ttk.Label(right, text="Chế độ", style="Muted.TLabel").grid(row=r, column=0, sticky="w", padx=(4, 14), pady=3)
            ttk.Checkbutton(
                right,
                text="Tự nhập các thông số này\n(bỏ tick = phần mềm tự tính)",
                variable=manual_other_var,
                command=toggle_manual_other_state,
            ).grid(row=r, column=1, columnspan=2, sticky="w", padx=(12, 4), pady=3)
            r += 1

            def add_other_row(row, label, var, width=20):
                widget = add_row(right, row, label, var, width, readonly=not bool(manual_other_var.get()))
                right_input_widgets.append(widget)
                return widget

            add_other_row(r, "Chiều dài ngàm H = L - Lo (m)", h_embed_var, 20); r += 1
            add_other_row(r, "Bề rộng tính toán theo phương x - Bpx(m)", bpx_var, 20); r += 1
            add_other_row(r, "Bề rộng tính toán theo phương y - Bpy (m)", bpy_var, 20); r += 1
            add_other_row(r, "Diện tích mũi cọc Fo(m²)", area_var, 20); r += 1
            add_other_row(r, "Moment quán tính tại mũi cọc Jo (m⁴)", jxy_var, 20); r += 1
            add_other_row(r, "Hệ số nền chịu nén tại mũi cọc Co (T/m³)", co_var, 20); r += 1
            add_other_row(r, "Hệ số nền chịu trượt tại mũi cọc Ct (T/m³)", ct_var, 20); r += 1
            add_other_row(r, "Modulus đàn hồi của vật liệu cọc Ec (T/m²)", ec_var, 20); r += 1
            add_other_row(r, "Modulus của vỏ cọc khi chịu nén Ev-nén (T/m²)", e_axial_var, 20); r += 1
            add_other_row(r, "Modulus của vỏ cọc khi chịu uốn Ev-uốn (T/m²)", e_bending_var, 20); r += 1
            toggle_manual_other_state()
            def normalize_kn(*_):
                s = str(kn_var.get()).strip()
                target = "1" if s.startswith("1") else "0"
                if target != s:
                    kn_var.set(target)

            for var in [pile_type_var, pile_d_var, pile_l_var, pile_lo_var, sct_pile_var, m_pile_tip_var, fc_var, kn_var]:
                try:
                    var.trace_add("write", lambda *_: derived_common_values(True))
                except Exception:
                    pass
            try:
                kn_var.trace_add("write", normalize_kn)
            except Exception:
                pass

            # --- Table helpers ---
            def create_tree_table(parent, columns, widths, height=18):
                frame = ttk.Frame(parent)
                frame.pack(fill=tk.BOTH, expand=True)
                tree = ttk.Treeview(frame, columns=columns, show="headings", height=height, selectmode="extended")
                vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
                hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
                tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
                tree.grid(row=0, column=0, sticky="nsew")
                vsb.grid(row=0, column=1, sticky="ns")
                hsb.grid(row=1, column=0, sticky="ew")
                frame.rowconfigure(0, weight=1)
                frame.columnconfigure(0, weight=1)
                for col, width in zip(columns, widths):
                    tree.heading(col, text=col)
                    tree.column(col, width=width, minwidth=max(60, min(width, 120)), anchor="center", stretch=True)
                return tree

            def clear_tree(tree):
                for item in tree.get_children():
                    tree.delete(item)
                try:
                    if tree is load_tree:
                        update_load_count()
                except Exception:
                    pass

            def insert_rows(tree, rows, clear=True):
                if clear:
                    clear_tree(tree)
                for row in rows:
                    tree.insert("", tk.END, values=[fmt_num(x) for x in row])

            def add_blank_rows(tree, n, start_index=None):
                cols = tree["columns"]
                existing = len(tree.get_children())
                for i in range(max(int(n), 1)):
                    vals = [""] * len(cols)
                    if start_index is not None:
                        vals[0] = str(existing + i + 1)
                    tree.insert("", tk.END, values=vals)
                try:
                    if tree is load_tree:
                        update_load_count()
                except Exception:
                    pass

            def add_rows_prompt(tree, title="Thêm dòng"):
                n = simpledialog.askinteger(
                    title,
                    "Nhập số dòng cần thêm:",
                    parent=win,
                    initialvalue=10,
                    minvalue=1,
                    maxvalue=10000,
                )
                if n:
                    add_blank_rows(tree, n, start_index=True)

            def delete_selected_rows(tree):
                sel = tree.selection()
                if not sel:
                    return
                for item in sel:
                    tree.delete(item)
                try:
                    if tree is load_tree:
                        update_load_count()
                except Exception:
                    pass

            def parse_numeric_lines(text, min_cols=1):
                rows = []
                for line in str(text or "").splitlines():
                    raw = line.strip()
                    if not raw:
                        continue
                    # Hỗ trợ paste từ Excel có tab, csv, hoặc text cách nhau bằng khoảng trắng.
                    nums = []
                    for m in re.finditer(r"[-+]?\d+(?:[\.,]\d+)?(?:[eE][-+]?\d+)?", raw):
                        try:
                            nums.append(float(m.group(0).replace(",", ".")))
                        except Exception:
                            pass
                    if len(nums) >= min_cols:
                        rows.append(nums)
                return rows

            def looks_like_index(x):
                try:
                    xf = float(x)
                    return xf >= 0 and abs(xf - round(xf)) < 1e-8 and xf < 100000
                except Exception:
                    return False

            load_columns = ["TH", "Hx", "Hy", "N", "Mx", "My", "Mz"]
            load_tree = None
            load_count_var = tk.StringVar(value="Chưa có tổ hợp")
            coord_columns = ["Cọc", "X", "Y", "Độ dốc đứng a", "f / Xi (độ)"]
            coord_tree = None

            def update_load_count():
                count = len(load_tree.get_children()) if load_tree is not None else 0
                load_count_var.set(f"Đã nạp {count} tổ hợp" if count else "Chưa có tổ hợp")
                return count

            def load_rows_from_numeric(nums_rows):
                rows = []
                for nums in nums_rows:
                    if len(nums) >= 7 and looks_like_index(nums[0]):
                        row = nums[:7]
                    elif len(nums) >= 6:
                        row = [len(rows) + 1] + nums[:6]
                    else:
                        continue
                    # Bỏ dòng 0 rỗng kiểu file Excel.
                    if abs(float(row[0])) < 1e-12 and all(abs(float(v)) < 1e-12 for v in row[1:]):
                        continue
                    rows.append(row)
                if rows:
                    insert_rows(load_tree, rows, clear=True)
                    children = load_tree.get_children()
                    if children:
                        last = children[-1]
                        load_tree.selection_set(last)
                        load_tree.focus(last)
                        load_tree.see(last)
                    update_load_count()
                else:
                    update_load_count()
                return len(rows)

            def coord_rows_from_numeric(nums_rows):
                rows = []
                for nums in nums_rows:
                    if len(nums) >= 5 and looks_like_index(nums[0]):
                        name, x, y, slope, xi = nums[0], nums[1], nums[2], nums[3], nums[4]
                    elif len(nums) >= 4:
                        name, x, y, slope, xi = len(rows) + 1, nums[0], nums[1], nums[2], nums[3]
                    else:
                        continue
                    if abs(float(name)) < 1e-12 and all(abs(float(v)) < 1e-12 for v in (x, y, slope, xi)):
                        continue
                    rows.append([name, x, y, slope, xi])
                if rows:
                    insert_rows(coord_tree, rows, clear=True)

            def paste_loads():
                try:
                    text = win.clipboard_get()
                except Exception as exc:
                    messagebox.showerror("Lỗi Clipboard", f"Không đọc được Clipboard: {exc}", parent=win)
                    return

                rows, diag = parse_load_clipboard_text(text)
                loaded = load_rows_from_numeric(rows)
                total = int(diag.get("nonempty_lines", 0) or 0)
                rejected = list(diag.get("rejected_rows", []) or [])

                # Hiển thị số thực tế đã nhận từ Clipboard, không chỉ số dòng đang nhìn thấy.
                if total:
                    load_count_var.set(f"Đã nạp {loaded}/{total} dòng Clipboard")
                else:
                    load_count_var.set("Clipboard không có dữ liệu")

                if rejected:
                    preview = "\n".join(
                        f"Dòng {line_no}: {raw[:160]}" for line_no, raw in rejected[:8]
                    )
                    more = "" if len(rejected) <= 8 else f"\n... và {len(rejected) - 8} dòng khác"
                    messagebox.showwarning(
                        "Paste chưa đủ dữ liệu",
                        f"Clipboard có {total} dòng không rỗng nhưng chỉ nạp được {loaded} tổ hợp.\n\n"
                        "Các dòng bị bỏ qua do thiếu/không nhận đủ 6 giá trị tải:\n"
                        f"{preview}{more}\n\n"
                        "Dấu '-' hoặc ô trống trong bảng Excel đã được hiểu là 0.",
                        parent=win,
                    )
                elif loaded == 0:
                    messagebox.showwarning(
                        "Không nhận được tổ hợp",
                        "Không tìm thấy dòng nào có đủ 6 giá trị tải. "
                        "Hãy copy 6 cột Hx-Hy-N-Mx-My-Mz hoặc 7 cột TH-Hx-Hy-N-Mx-My-Mz.",
                        parent=win,
                    )

            def paste_coords():
                try:
                    text = win.clipboard_get()
                except Exception:
                    text = ""
                rows = parse_numeric_lines(text, min_cols=4)
                coord_rows_from_numeric(rows)

            def import_numeric_file(kind):
                fp = filedialog.askopenfilename(parent=win, title="Chọn file dữ liệu", filetypes=[("CSV/TXT/DAT", "*.csv *.txt *.dat"), ("Tất cả", "*.*")])
                if not fp:
                    return
                try:
                    with open(fp, "r", encoding="utf-8", errors="ignore") as f:
                        rows = parse_numeric_lines(f.read(), min_cols=4 if kind == "coord" else 6)
                    if kind == "coord":
                        coord_rows_from_numeric(rows)
                    else:
                        load_rows_from_numeric(rows)
                except Exception as exc:
                    messagebox.showerror("Lỗi đọc file", str(exc), parent=win)

            def begin_cell_edit(tree, event, readonly_cols=None, after_edit=None):
                readonly_cols = set(readonly_cols or [])
                if tree.identify("region", event.x, event.y) != "cell":
                    return
                item = tree.identify_row(event.y)
                col_id = tree.identify_column(event.x)
                if not item or not col_id:
                    return
                col_index = int(col_id.replace("#", "")) - 1
                columns = list(tree["columns"])
                col_name = columns[col_index]
                if col_name in readonly_cols:
                    return
                bbox = tree.bbox(item, col_id)
                if not bbox:
                    return
                x, y, w, h = bbox
                old = tree.set(item, col_name)
                editor = ttk.Entry(tree)
                editor.insert(0, old)
                editor.select_range(0, tk.END)
                editor.focus_set()
                editor.place(x=x, y=y, width=w, height=h)

                def commit(_=None):
                    try:
                        tree.set(item, col_name, editor.get().strip())
                        if after_edit:
                            after_edit(item)
                    finally:
                        editor.destroy()

                def cancel(_=None):
                    editor.destroy()

                editor.bind("<Return>", commit)
                editor.bind("<FocusOut>", commit)
                editor.bind("<Escape>", cancel)

            def update_coord_phi(item=None):
                return

            # --- Menu chuột phải cho bảng tổ hợp tải trọng và tọa độ cọc ---
            def _tree_ordered_selection(tree):
                selected = list(tree.selection())
                children = list(tree.get_children())
                return [r for r in children if r in selected]

            def _tree_row_values(tree, rowid):
                vals = list(tree.item(rowid, "values"))
                ncol = len(tree["columns"])
                return (vals + [""] * ncol)[:ncol]

            def _tree_set_cell(tree, rowid, col_index, value):
                vals = _tree_row_values(tree, rowid)
                if 0 <= int(col_index) < len(vals):
                    vals[int(col_index)] = value
                    tree.item(rowid, values=vals)

            def _copy_selected_rows(tree):
                rows = [_tree_row_values(tree, r) for r in _tree_ordered_selection(tree)]
                if not rows:
                    return
                text = "\n".join("\t".join(str(v) for v in row) for row in rows)
                try:
                    win.clipboard_clear()
                    win.clipboard_append(text)
                except Exception:
                    pass

            def _fill_clicked_cell_to_selected(tree, rowid, col_index):
                if not rowid or col_index is None:
                    return
                selected = _tree_ordered_selection(tree)
                if rowid not in selected:
                    selected = [rowid]
                children = list(tree.get_children())
                try:
                    pos = children.index(rowid)
                except ValueError:
                    pos = -1
                targets = [r for r in selected if r != rowid and (pos < 0 or children.index(r) > pos)]
                if not targets:
                    targets = [r for r in selected if r != rowid]
                if not targets:
                    messagebox.showinfo("Copy giá trị ô đang chọn", "Hãy chọn thêm các dòng cần copy giá trị.", parent=win)
                    return
                value = _tree_row_values(tree, rowid)[col_index]
                for r0 in targets:
                    _tree_set_cell(tree, r0, col_index, value)

            def _fill_selected_empty_from_neighbor(tree, direction="above"):
                children = list(tree.get_children())
                selected = _tree_ordered_selection(tree)
                if not selected:
                    return
                iterable = selected if direction == "above" else list(reversed(selected))
                for rowid in iterable:
                    try:
                        idx = children.index(rowid)
                    except ValueError:
                        continue
                    neighbor_idx = idx - 1 if direction == "above" else idx + 1
                    if neighbor_idx < 0 or neighbor_idx >= len(children):
                        continue
                    vals = _tree_row_values(tree, rowid)
                    nvals = _tree_row_values(tree, children[neighbor_idx])
                    changed = False
                    for c in range(len(vals)):
                        if str(vals[c]).strip() == "" and str(nvals[c]).strip() != "":
                            vals[c] = nvals[c]
                            changed = True
                    if changed:
                        tree.item(rowid, values=vals)

            def _show_table_context_menu(tree, event, paste_command):
                rowid = tree.identify_row(event.y)
                colid = tree.identify_column(event.x)
                try:
                    col_index = int(str(colid).replace("#", "")) - 1
                except Exception:
                    col_index = None
                if rowid:
                    if rowid not in tree.selection():
                        tree.selection_set(rowid)
                    tree.focus(rowid)
                menu = tk.Menu(tree, tearoff=0)
                menu.add_command(label="Edit ô", command=lambda e=event: begin_cell_edit(tree, e))
                menu.add_command(label="Paste từ Clipboard", command=paste_command)
                menu.add_command(label="Copy dòng đã chọn", command=lambda: _copy_selected_rows(tree))
                menu.add_command(label="Copy giá trị ô đang chọn", command=lambda: _fill_clicked_cell_to_selected(tree, rowid, col_index))
                menu.add_command(label="Copy ô trên", command=lambda: _fill_selected_empty_from_neighbor(tree, "above"))
                menu.add_command(label="Copy ô dưới", command=lambda: _fill_selected_empty_from_neighbor(tree, "below"))
                menu.add_separator()
                menu.add_command(label="Xóa dòng", command=lambda: delete_selected_rows(tree))
                menu.add_command(label="Xóa tất cả", command=lambda: clear_tree(tree))
                try:
                    menu.tk_popup(event.x_root, event.y_root)
                finally:
                    menu.grab_release()

            # --- Tab tổ hợp tải trọng ---
            load_top = ttk.Frame(tab_load)
            load_top.pack(fill=tk.X, pady=(0, 6))
            ttk.Label(load_top, text="Có thể paste dạng TH-Hx-Hy-N-Mx-My-Mz hoặc Hx-Hy-N-Mx-My-Mz.", style="Muted.TLabel").pack(side=tk.LEFT)
            ttk.Label(load_top, textvariable=load_count_var, style="Muted.TLabel").pack(side=tk.LEFT, padx=(14, 0))
            ttk.Button(load_top, text="Import TH", command=lambda: import_numeric_file("load")).pack(side=tk.RIGHT, padx=3)
            ttk.Button(load_top, text="Paste từ Clipboard", command=paste_loads).pack(side=tk.RIGHT, padx=3)
            ttk.Button(load_top, text="+(...) dòng", command=lambda: add_rows_prompt(load_tree, "Thêm dòng tổ hợp tải trọng")).pack(side=tk.RIGHT, padx=3)
            ttk.Button(load_top, text="Xóa dòng", command=lambda: delete_selected_rows(load_tree)).pack(side=tk.RIGHT, padx=3)
            ttk.Button(load_top, text="Xóa hết", command=lambda: clear_tree(load_tree)).pack(side=tk.RIGHT, padx=3)
            load_tree = create_tree_table(tab_load, load_columns, [62, 105, 105, 115, 115, 115, 115], height=18)
            load_tree.bind("<Double-1>", lambda e: begin_cell_edit(load_tree, e))
            load_tree.bind("<Button-3>", lambda e: _show_table_context_menu(load_tree, e, paste_loads))
            load_tree.bind("<Control-v>", lambda e: (paste_loads(), "break")[1])
            load_tree.bind("<Control-V>", lambda e: (paste_loads(), "break")[1])
            load_tree.bind("<Shift-Insert>", lambda e: (paste_loads(), "break")[1])

            # --- Tab tọa độ cọc ---
            coord_top = ttk.Frame(tab_piles)
            coord_top.pack(fill=tk.X, pady=(0, 6))
            ttk.Label(coord_top, text="Có thể paste dạng Cọc-X-Y-a-f hoặc X-Y-a-f.", style="Muted.TLabel").pack(side=tk.LEFT)
            ttk.Button(coord_top, text="Import tọa độ", command=lambda: import_numeric_file("coord")).pack(side=tk.RIGHT, padx=3)
            ttk.Button(coord_top, text="Paste từ Clipboard", command=paste_coords).pack(side=tk.RIGHT, padx=3)
            ttk.Button(coord_top, text="+(...) dòng", command=lambda: add_rows_prompt(coord_tree, "Thêm dòng tọa độ cọc")).pack(side=tk.RIGHT, padx=3)
            ttk.Button(coord_top, text="Xóa dòng", command=lambda: delete_selected_rows(coord_tree)).pack(side=tk.RIGHT, padx=3)
            ttk.Button(coord_top, text="Xóa hết", command=lambda: clear_tree(coord_tree)).pack(side=tk.RIGHT, padx=3)
            coord_tree = create_tree_table(tab_piles, coord_columns, [70, 140, 140, 150, 150], height=18)
            coord_tree.bind("<Double-1>", lambda e: begin_cell_edit(coord_tree, e))
            coord_tree.bind("<Button-3>", lambda e: _show_table_context_menu(coord_tree, e, paste_coords))

            # --- Preview ---
            preview_text = tk.Text(tab_preview, height=20, wrap="none", font=("Consolas", 9))
            preview_text.pack(fill=tk.BOTH, expand=True)
            preview_text.config(state=tk.DISABLED)
            prev_x = ttk.Scrollbar(tab_preview, orient="horizontal", command=preview_text.xview)
            prev_y = ttk.Scrollbar(tab_preview, orient="vertical", command=preview_text.yview)
            preview_text.configure(xscrollcommand=prev_x.set, yscrollcommand=prev_y.set)
            prev_y.pack(side=tk.RIGHT, fill=tk.Y)
            prev_x.pack(side=tk.BOTTOM, fill=tk.X)

            def collect_combos():
                combos = []
                for item in load_tree.get_children():
                    vals = list(load_tree.item(item, "values"))
                    while len(vals) < 7:
                        vals.append("")
                    th = self._safe_float(vals[0], len(combos) + 1)
                    data = [self._safe_float(vals[i], 0.0) for i in range(1, 7)]
                    # Bỏ dòng trắng hoặc dòng 0 rỗng.
                    if not any(str(v).strip() for v in vals):
                        continue
                    if abs(th) < 1e-12 and all(abs(v) < 1e-12 for v in data):
                        continue
                    # Nếu người dùng thêm dòng trắng có số thứ tự nhưng chưa nhập tải thì bỏ qua.
                    if all(abs(v) < 1e-12 for v in data) and str(vals[0]).strip() and len(str(vals[0]).strip()) < 8:
                        continue
                    combos.append({
                        "Name": str(int(th)) if abs(th - round(th)) < 1e-8 else fmt_num(th),
                        "Hx": data[0], "Hy": data[1], "N_load": data[2],
                        "Mx": data[3], "My": data[4], "Mz": data[5],
                    })
                return combos

            def collect_piles():
                common = derived_common_values(True)
                piles = []
                for item in coord_tree.get_children():
                    vals = list(coord_tree.item(item, "values"))
                    while len(vals) < 5:
                        vals.append("")
                    if not any(str(v).strip() for v in vals):
                        continue
                    name = self._safe_float(vals[0], len(piles) + 1)
                    x = self._safe_float(vals[1], 0.0)
                    y = self._safe_float(vals[2], 0.0)
                    slope = self._safe_float(vals[3], 0.0)
                    xi = self._safe_float(vals[4], 0.0)
                    if abs(name) < 1e-12 and all(abs(v) < 1e-12 for v in (x, y, slope, xi)):
                        continue
                    phi = math.degrees(math.atan(slope))
                    piles.append({
                        "Lo": common["Lo"], "H": common["H"],
                        "Bpx": common["Bpx"], "Bpy": common["Bpy"],
                        "Sec1": common["d_ngoai"], "Sec2": common["d_trong"], "Sec3": common["day_vo"],
                        "d_ngoai": common["d_ngoai"], "d_trong": common["d_trong"], "day_vo": common["day_vo"],
                        "Area": common["Area"], "J_xy": common["J_xy"],
                        "Po": common["Po"], "Co": common["Co"], "Ct": common["Ct"],
                        "X": x, "Y": y, "Phi": phi, "Xi": xi,
                    })
                return piles, common

            def collect_globals(common=None):
                if common is None:
                    common = derived_common_values(True)
                return {
                    "Kn": fvar(kn_var, 0.0),
                    "Bx": fvar(bx_var, 0.0), "By": fvar(by_var, 0.0), "Cz": fvar(cz_var, 0.0),
                    "EI_uon": common.get("EI_uon", 0.0), "Er_uon": 0.0,
                    "EA_nen": common.get("EA_nen", 0.0), "Er_nen": 0.0,
                    "md": fvar(md_var, 0.0), "mq": fvar(mq_var, 0.0), "m": fvar(m_work_var, 0.0),
                }

            def format_preview_input(project, gvals, combos, piles):
                n_piles = len(piles)
                n_combos = len(combos)
                header_nums = [
                    n_piles, n_combos, 0, gvals.get("Legacy_Header_Flag", 2),
                    gvals.get("Legacy_Reserved_1", 0),
                    gvals.get("Pile_Count_Copy", n_piles),
                    gvals.get("Kn", 0.0), gvals.get("Legacy_Reserved_2", 0),
                    gvals.get("Bx", 0.0), gvals.get("By", 0.0), gvals.get("Cz", 0.0),
                    gvals.get("EI_uon", 0.0), gvals.get("Er_uon", 0.0),
                    gvals.get("EA_nen", 0.0), gvals.get("Er_nen", 0.0),
                    gvals.get("md", 0.0), gvals.get("mq", 0.0), gvals.get("m", 0.0),
                ]
                def fmt_header(i, x):
                    if i in (11, 12, 13, 14):
                        return fmt_num(x, 0, strip=False)
                    if i in (8, 9, 10, 15, 16, 17):
                        return fmt_num(x, 3)
                    return fmt_num(x, 0, strip=False)

                def fmt_pile_value(key, x):
                    if key in ("Area", "J_xy"):
                        return fmt_num(x, 12, strip=True)
                    if key in ("Bpx", "Bpy", "d_ngoai", "d_trong", "day_vo", "Co", "Ct", "Phi"):
                        return fmt_num(x, 3)
                    if key in ("Lo", "H", "Po", "X", "Y", "Xi"):
                        return fmt_num(x, 6)
                    return fmt_num(x)

                lines = [str(project or "TS_DIRECT_INPUT")]
                lines.append(" ".join(fmt_header(i, x) for i, x in enumerate(header_nums[:11])))
                lines.append(" ".join(fmt_header(i + 11, x) for i, x in enumerate(header_nums[11:])))
                for row in combos:
                    lines.append(" ".join(fmt_num(row.get(k, 0.0)) for k in ("Hx", "Hy", "N_load", "Mx", "My", "Mz")))
                pile_keys = ("Lo", "H", "Bpx", "Bpy", "d_ngoai", "d_trong", "day_vo", "Area", "J_xy", "Po", "Co", "Ct", "X", "Y", "Phi", "Xi")
                for row in piles:
                    for k in pile_keys:
                        lines.append(fmt_pile_value(k, row.get(k, 0.0)))
                return "\n".join(lines) + "\n"

            def update_preview():
                try:
                    project = project_var.get().strip() or "input"
                    piles, common = collect_piles()
                    combos = collect_combos()
                    gvals = collect_globals(common)
                    text = format_preview_input(project, gvals, combos, piles)
                    preview_text.config(state=tk.NORMAL)
                    preview_text.delete("1.0", tk.END)
                    preview_text.insert(tk.END, text)
                    preview_text.config(state=tk.DISABLED)
                    nb.select(tab_preview)
                except Exception as exc:
                    messagebox.showerror("Lỗi xem trước", str(exc), parent=win)

            def collect_and_add():
                project = project_var.get().strip()
                try:
                    piles, common = collect_piles()
                    combos = collect_combos()
                    gvals = collect_globals(common)
                except Exception as exc:
                    messagebox.showerror("Lỗi thu thập dữ liệu", str(exc), parent=win)
                    return
                bad = []
                if int(round(gvals.get("Kn", 0.0))) not in (0, 1):
                    bad.append("Kn phải bằng 0 hoặc 1 theo engine hiện tại")
                if not combos:
                    bad.append("Chưa có tổ hợp tải trọng")
                if not piles:
                    bad.append("Chưa có tọa độ cọc")
                try:
                    d_check_m = fvar(pile_d_var, 0.0)
                except Exception:
                    d_check_m = 0.0
                if d_check_m <= 0:
                    bad.append("Đường kính/cạnh cọc D phải > 0")
                elif d_check_m > 10.0:
                    bad.append("Đường kính/cạnh cọc D > 10 m; kiểm tra lại giá trị nhập")
                if common.get("EA_nen", 0.0) <= 0:
                    bad.append("Ev nén tự tính <= 0; kiểm tra f'c")
                for i, p in enumerate(piles, start=1):
                    if p.get("Lo", 0.0) + p.get("H", 0.0) <= 0:
                        bad.append(f"Cọc {i}: Lo+H <= 0")
                    if p.get("Area", 0.0) <= 0:
                        bad.append(f"Cọc {i}: Fo <= 0")
                    if p.get("J_xy", 0.0) <= 0:
                        bad.append(f"Cọc {i}: Io/J_xy <= 0")
                if bad:
                    messagebox.showerror("Dữ liệu chưa hợp lệ", "\n".join(bad[:12]), parent=win)
                    return
                try:
                    display_name = self._next_direct_input_name(project)
                    fp = self._write_direct_input_file(display_name, gvals, combos, piles)
                    if fp not in self.file_list:
                        self._append_input_file(fp, display_name)
                    if not self.out_dir_var.get().strip():
                        self.out_dir_var.set(os.path.dirname(fp))
                    self._log(f"Đã tạo dữ liệu trực tiếp theo Data MCOC: {os.path.basename(fp)}", "OK")
                    messagebox.showinfo("Đã thêm dữ liệu", f"Đã tạo và thêm file INPUT:\n{fp}", parent=win)
                    win.destroy()
                except Exception as exc:
                    messagebox.showerror("Lỗi tạo INPUT", str(exc), parent=win)

            # Không tự nạp dữ liệu mẫu khi mở cửa sổ nhập trực tiếp.

            btns = ttk.Frame(main)
            btns.pack(fill=tk.X, pady=(8, 0))
            ttk.Button(btns, text="Hủy", command=win.destroy).pack(side=tk.RIGHT, padx=4)
            ttk.Button(btns, text="Tạo INPUT và đưa vào danh sách tính", style="Primary.TButton", command=collect_and_add).pack(side=tk.RIGHT, padx=4)
            ttk.Button(btns, text="Cập nhật/Xem trước", command=update_preview).pack(side=tk.RIGHT, padx=4)
            ttk.Button(btns, text="Paste TH", command=lambda: [nb.select(tab_load), paste_loads()]).pack(side=tk.LEFT, padx=4)
            ttk.Button(btns, text="Paste tọa độ", command=lambda: [nb.select(tab_piles), paste_coords()]).pack(side=tk.LEFT, padx=4)
            ttk.Button(btns, text="Cập nhật các thông số khác", command=lambda: derived_common_values(True)).pack(side=tk.LEFT, padx=4)

            self._center_window(win)

        def toggle_group_method(self):
            if self.chk_group_main.get(): self.cbo_group_method.config(state="readonly")
            else: self.cbo_group_method.config(state="disabled")

        def toggle_stiffness(self):
            if self.chk_stiffness.get():
                self.cb_stiff_direct.config(state=tk.NORMAL)
                self.cb_rm.config(state=tk.NORMAL)
            else:
                self.chk_stiff_direct.set(False)
                self.chk_rm.set(False)
                self.cb_stiff_direct.config(state=tk.DISABLED)
                self.cb_rm.config(state=tk.DISABLED)

        def toggle_print_options(self):
            if self.chk_print_now.get():
                self.cbo_printer.config(state="readonly")
                self.cbo_paper_size.config(state="readonly")
            else:
                self.cbo_printer.config(state="disabled")
                self.cbo_paper_size.config(state="disabled")

        def toggle_pdf(self):
            if self.chk_pdf.get(): self.cb_merge.config(state=tk.NORMAL)
            else: self.chk_merge.set(False); self.cb_merge.config(state=tk.DISABLED)

        def toggle_report_form(self, *_args):
            is_new = self.cbo_report_form.get().lower().startswith("new")
            self.report_form_var.set("new" if is_new else "classic")
            if hasattr(self, "cb_txt_classic_when_new"):
                self.cb_txt_classic_when_new.config(state=(tk.NORMAL if is_new else tk.DISABLED))
                if not is_new:
                    self.chk_txt_classic_when_new.set(False)

        def is_valid_input_file(self, fp):
            """Chỉ nhận file INPUT MCOC dạng .txt/.TXT hoặc file không có đuôi.
            Không nhận các file kết quả/báo cáo hoặc file trung gian có đuôi khác.
            """
            try:
                if not fp or not os.path.isfile(fp):
                    return False
                b = os.path.basename(fp)
                bl = b.lower()
                if "_out" in bl or "merged_" in bl or "_tscol_input" in bl:
                    return False
                root, ext = os.path.splitext(b)
                return ext == "" or ext.lower() == ".txt"
            except Exception:
                return False

        def _input_file_dialog_types(self):
            # File "không có phần mở rộng" nghĩa là tên file không có dấu chấm/đuôi mở rộng.
            # Pattern "*." là cách Tk/Windows lọc nhóm file không có extension;
            # tuyệt đối không dùng pattern như "*.khongduoi" hay "*.không đuôi".
            return [
                ("File TXT (*.txt)", "*.txt *.TXT"),
                ("File không có phần mở rộng (*.)", "*."),
                ("Tất cả file", "*.*"),
            ]
        
        def add_files(self):
            files = filedialog.askopenfilenames(
                title="Chọn file INPUT MCOC (.txt hoặc không đuôi)",
                filetypes=self._input_file_dialog_types()
            )
            skipped = 0
            added = 0
            for f in files:
                if self.is_valid_input_file(f) and f not in self.file_list:
                    self._append_input_file(f, os.path.basename(f))
                    added += 1
                else:
                    skipped += 1
            if skipped and not added:
                messagebox.showwarning("Không đúng định dạng", "Chỉ nhận file .txt/.TXT hoặc file không có đuôi.")
                    
        def add_folder(self):
            folder = filedialog.askdirectory(title="Chọn thư mục chứa file INPUT MCOC")
            if folder:
                added = 0
                skipped = 0
                for fn in os.listdir(folder):
                    fp = os.path.join(folder, fn)
                    if os.path.isfile(fp) and self.is_valid_input_file(fp) and fp not in self.file_list:
                        self._append_input_file(fp, os.path.basename(fp))
                        added += 1
                    else:
                        skipped += 1
                if hasattr(self, "_log"):
                    self._log(f"Thêm thư mục: nhận {added} file INPUT MCOC (.txt/không đuôi), bỏ qua {skipped} file khác.")
                        
        def delete_selected(self):
            idx = self.listbox_files.curselection()
            if not idx: messagebox.showwarning("Cảnh báo", "Vui lòng chọn file để xóa."); return
            for i in reversed(idx):
                self.listbox_files.delete(i)
                del self.file_list[i]
                try:
                    del self.file_display_names[i]
                except Exception:
                    pass
            
        def clear_files(self):
            self.file_list.clear()
            try:
                self.file_display_names.clear()
            except Exception:
                self.file_display_names = []
            self.listbox_files.delete(0, tk.END)
        def choose_out_dir(self):
            folder = filedialog.askdirectory()
            if folder: self.out_dir_var.set(folder)

        def _ask_ttgh_grouping(self, combo_counts):
            """Hỏi và thu thập phạm vi tổ hợp theo ba trạng thái giới hạn."""
            answer = messagebox.askyesnocancel(
                "Phân loại tổ hợp theo TTGH",
                "Có tách bảng tổng hợp Max/Min theo trạng thái giới hạn không?\n\n"
                "Có: chọn tổ hợp cho Sử dụng, Cường độ và Đặc biệt.\n"
                "Không: xuất báo cáo như bình thường.\n"
                "Hủy: dừng lệnh tính.",
                parent=self.root,
            )
            if answer is None:
                return None
            if answer is False:
                return []

            valid_counts = [int(v) for v in (combo_counts or []) if int(v) > 0]
            max_combo = max(valid_counts) if valid_counts else None
            win = tk.Toplevel(self.root)
            win.title("Phân loại tổ hợp theo trạng thái giới hạn")
            win.resizable(False, False)
            try:
                win.transient(self.root)
                win.grab_set()
            except Exception:
                pass

            frame = ttk.Frame(win, padding=14)
            frame.pack(fill=tk.BOTH, expand=True)
            ttk.Label(frame, text="Phân loại tổ hợp tải trọng", style="Header.TLabel").grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 8))
            guide = (
                "Nhập số hoặc khoảng, ví dụ: 1-4 hoặc 1-4,7,9-11.\n"
                "Nhập 0 để bỏ qua một trạng thái giới hạn. Mỗi tổ hợp chỉ được thuộc một nhóm."
            )
            ttk.Label(frame, text=guide, style="Muted.TLabel", justify=tk.LEFT).grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(0, 10))
            if valid_counts:
                count_text = ", ".join(map(str, sorted(set(valid_counts))))
                suffix = "" if len(set(valid_counts)) == 1 else " (áp dụng theo số TH; TH không tồn tại trong một file sẽ tự bỏ qua)"
                ttk.Label(frame, text=f"Số tổ hợp trong các file đang chọn: {count_text}{suffix}", style="Muted.TLabel", wraplength=570).grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(0, 10))

            last_specs = getattr(self, "_last_ttgh_specs", {}) or {}
            variables = {
                "SD": tk.StringVar(value=str(last_specs.get("SD", "0"))),
                "CD": tk.StringVar(value=str(last_specs.get("CD", "0"))),
                "DB": tk.StringVar(value=str(last_specs.get("DB", "0"))),
            }
            labels = (
                ("SD", "TTGHSD - Sử dụng:"),
                ("CD", "TTGHCĐ - Cường độ:"),
                ("DB", "TTGHĐB - Đặc biệt:"),
            )
            first_entry = None
            for row_idx, (code, label) in enumerate(labels, start=3):
                ttk.Label(frame, text=label).grid(row=row_idx, column=0, sticky=tk.W, padx=(0, 12), pady=5)
                entry = ttk.Entry(frame, textvariable=variables[code], width=34)
                entry.grid(row=row_idx, column=1, sticky=tk.EW, pady=5)
                if first_entry is None:
                    first_entry = entry

            result = {"value": None}
            def accept():
                specs = {code: var.get().strip() or "0" for code, var in variables.items()}
                try:
                    groups = _build_ttgh_groups(specs, max_combo=max_combo)
                except ValueError as exc:
                    messagebox.showerror("Phân loại TTGH chưa hợp lệ", str(exc), parent=win)
                    return
                self._last_ttgh_specs = dict(specs)
                result["value"] = groups
                win.destroy()

            buttons = ttk.Frame(frame)
            buttons.grid(row=6, column=0, columnspan=2, sticky=tk.E, pady=(14, 0))
            ttk.Button(buttons, text="Hủy", command=win.destroy).pack(side=tk.RIGHT, padx=(6, 0))
            ttk.Button(buttons, text="Áp dụng", style="Primary.TButton", command=accept).pack(side=tk.RIGHT)
            win.protocol("WM_DELETE_WINDOW", win.destroy)
            win.bind("<Return>", lambda _e: accept())
            win.bind("<Escape>", lambda _e: win.destroy())
            self._center_window(win)
            if first_entry is not None:
                first_entry.focus_set()
                first_entry.selection_range(0, tk.END)
            self.root.wait_window(win)
            return result["value"]

        def run_calculation(self):
            if not self.file_list: messagebox.showwarning("Lỗi", "Chưa chọn file đầu vào!"); return
            extra_classic_txt = (
                hasattr(self, "chk_txt_classic_when_new")
                and self.report_form_var.get().lower() == "new"
                and self.chk_txt_classic_when_new.get()
            )
            if not (self.chk_doc.get() or self.chk_excel.get() or self.chk_pdf.get() or self.chk_txt.get() or extra_classic_txt or self.chk_col_csv.get()):
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất 1 định dạng kết quả.")
                return

            ttgh_groups = []
            if hasattr(self, "chk_p7_ttgh") and self.chk_p7_ttgh.get():
                combo_counts = []
                for fp in self.file_list:
                    parsed = parse_foundation_input(fp)
                    if isinstance(parsed, dict):
                        combo_counts.append(len(parsed.get("Load_Combos", [])))
                ttgh_groups = self._ask_ttgh_grouping(combo_counts)
                if ttgh_groups is None:
                    return

            out_dir = self.out_dir_var.get() or os.path.dirname(self.file_list[0])
            if out_dir: os.makedirs(out_dir, exist_ok=True)
            self.progress['maximum'] = len(self.file_list)
            self.progress.config(value=0)
            self._log(f"Bắt đầu xử lý {len(self.file_list)} file. Thư mục xuất: {out_dir}")
            self.btn_run.config(state=tk.DISABLED)
            threading.Thread(target=self._calculation_worker, args=(out_dir, ttgh_groups), daemon=True).start()

        def _calculation_worker(self, out_dir, ttgh_groups=None):
            pdf_files_to_merge, individual_print_targets, temp_files_to_delete = [], [], []
            
            c_company = self.company_var.get().strip() or "TẬP ĐOÀN SUNGROUP"
            c_dept = self.department_var.get().strip() or "KHỐI XD PPP - PHÒNG QLTK"
            c_author = self.author_var.get().strip() or 'TSDUNGVN - 2026'

            report_config = {
                'p1': self.chk_p1.get(), 'p2': self.chk_p2.get(), 'p3': self.chk_p3.get(),
                'p4': self.chk_p4.get(), 'p5': self.chk_p5.get(), 'p6': self.chk_p6.get(),
                'p7': self.chk_p7.get(), 'p7_ttgh': self.chk_p7_ttgh.get(),
                'ttgh_groups': list(ttgh_groups or []),
                'p8': self.chk_p8.get(), 'p8_opt': self.var_p8_opt.get(),
                'p9': self.chk_p9.get(), 'p9_direct': self.chk_stiff_direct.get(),
                'form': self.report_form_var.get(),
                'txt_classic_when_new': (
                    hasattr(self, "chk_txt_classic_when_new")
                    and self.report_form_var.get().lower() == "new"
                    and self.chk_txt_classic_when_new.get()
                ),
                'col_csv': self.chk_col_csv.get()
            }

            try:
                if report_config.get("p7_ttgh") and report_config.get("ttgh_groups"):
                    group_text = "; ".join(
                        f"{g.get('short_label')}: {g.get('spec')}"
                        for g in report_config.get("ttgh_groups", [])
                        if g.get("combo_ids")
                    )
                    self._log(f"Phân loại tổ hợp theo TTGH: {group_text}")
                for idx, fp in enumerate(self.file_list):
                    self._log(f"Đang tính toán {os.path.basename(fp)}...")
                    data = parse_foundation_input(fp)
                    if isinstance(data, str):
                        self._log(f"Lỗi đọc {os.path.basename(fp)}: {data}", "ERROR")
                        self.root.after(0, lambda e=data, f=fp: messagebox.showerror("Lỗi", f"Lỗi đọc {os.path.basename(f)}:\n{e}"))
                        continue
                    
                    self._log(f"Đã đọc {os.path.basename(fp)}: {len(data.get('Piles', []))} cọc, {len(data.get('Load_Combos', []))} tổ hợp tải. Kiểm tra lại nếu khác kỳ vọng (file lệch cột sẽ trôi dữ liệu).")
                    data['Config'] = {
                        'Group_Effect_Enabled': self.chk_group_main.get(),
                        'Group_Method': self.cbo_group_method.get(),
                        'Ignore_Cz': False,
                        'Shadow_Effect': self.chk_shadow_effect.get()
                    }
                    
                    try:
                        solver = PiledRaftFoundation(data)
                        disp, forces, verifications, effs = solver.solve()
                        for wmsg in getattr(solver, "stiffness_warnings", []) or []:
                            self._log(wmsg, "WARN")
                    except ValueError as ve:
                        self._log(f"Lỗi tính toán {os.path.basename(fp)}: {ve}", "ERROR")
                        self.root.after(0, lambda e_msg=str(ve), f=fp: messagebox.showerror("Lỗi Tính Toán", f"File: {os.path.basename(f)}\n{e_msg}"))
                        continue
                        
                    eq_k, rm_k = None, None
                    if self.chk_stiffness.get():
                        eq_k = compute_equivalent_stiffness(solver.K)
                        if self.chk_rm.get(): rm_k = convert_stiffness_to_rmbridge(eq_k)

                    base = os.path.splitext(os.path.basename(fp))[0]

                    if report_config.get('col_csv', False):
                        out_col_csv = os.path.join(out_dir, f"{base}_TSCOL_INPUT.csv")
                        export_forces_to_n2d_col_csv(out_col_csv, forces, source_file=os.path.basename(fp), item=base, limit_state="CĐ")
                        self.last_col_csv = out_col_csv
                        self.last_col_forces = list(forces)
                        self._log(f"Đã xuất CSV cho TS-COL: {os.path.basename(out_col_csv)}", "OK")
                    
                    report_form = str(report_config.get('form', 'classic')).lower()
                    doc_export_func = export_to_new_doc if report_form == 'new' else export_to_legacy_doc
                    txt_export_func = export_to_new_text if report_form == 'new' else export_to_legacy_doc

                    out_doc = os.path.join(out_dir, f"{base}_OUT.doc")
                    success = safe_save_file(self.root, out_doc, doc_export_func, data, disp, forces, verifications, effs, out_doc, eq_k, rm_k, c_company, c_dept, c_author, report_config)
                    if not success: continue
                    if not self.chk_doc.get(): temp_files_to_delete.append(out_doc)

                    out_txt = os.path.join(out_dir, f"{base}_OUT.txt")
                    if self.chk_txt.get() or self.chk_print_now.get():
                        safe_save_file(self.root, out_txt, txt_export_func, data, disp, forces, verifications, effs, out_txt, eq_k, rm_k, c_company, c_dept, c_author, report_config)
                        if not self.chk_txt.get(): temp_files_to_delete.append(out_txt)

                    if report_form == 'new' and report_config.get('txt_classic_when_new', False):
                        out_txt_classic = os.path.join(out_dir, f"{base}_OUT_C.txt")
                        safe_save_file(self.root, out_txt_classic, export_to_legacy_doc, data, disp, forces, verifications, effs, out_txt_classic, eq_k, rm_k, c_company, c_dept, c_author, report_config)

                    individual_print_targets.append(out_doc)

                    if self.chk_excel.get():
                        out_excel = os.path.join(out_dir, f"{base}_OUT.xlsx")
                        safe_save_file(self.root, out_excel, export_to_excel, data, disp, forces, verifications, effs, out_excel, eq_k, rm_k, c_company, c_dept, c_author, report_config)

                    if self.chk_pdf.get():
                        out_pdf = os.path.join(out_dir, f"{base}_OUT.pdf")
                        if str(report_config.get('form', 'classic')).lower() == 'new':
                            safe_save_file(self.root, out_pdf, export_to_new_pdf, data, disp, forces, verifications, effs, out_pdf, eq_k, rm_k, c_company, c_dept, c_author, report_config)
                        else:
                            safe_save_file(self.root, out_pdf, export_to_pdf_from_text, out_doc, out_pdf)
                        pdf_files_to_merge.append(out_pdf)

                    self.root.after(0, lambda i=idx: self.progress.config(value=i + 1))
                    self._log(f"Hoàn thành {os.path.basename(fp)}", "OK")

                merged_success = False
                if self.chk_merge.get() and len(pdf_files_to_merge) > 1:
                    out_merged = os.path.join(out_dir, "Merged_BaoCao_TongHop.pdf")
                    merge_pdfs(pdf_files_to_merge, out_merged); merged_success = True

                self._log("Hoàn thành tính toán. Đang xử lý in/gộp PDF nếu có...", "OK")

                if self.chk_print_now.get() and platform.system() == "Windows":
                    printer_name = self.cbo_printer.get()
                    paper_str = self.cbo_paper_size.get()
                    paper_code = 7
                    if "A3" in paper_str: paper_code = 6
                    elif "Letter" in paper_str: paper_code = 0
                    
                    for p_target in individual_print_targets:
                        if os.path.exists(p_target):
                            filepath_ps = os.path.abspath(p_target).replace("'", "''")
                            printer_arg = f"$w.ActivePrinter = '{printer_name}'" if printer_name != "Default Printer" else ""
                            
                            ps_script = f"""
                            try {{
                                $w = [System.Runtime.InteropServices.Marshal]::GetActiveObject('Word.Application')
                            }} catch {{
                                $w = New-Object -ComObject Word.Application
                            }}
                            try {{
                                $w.Visible = $false
                                {printer_arg}
                                $d = $w.Documents.Open('{filepath_ps}')
                                foreach ($sec in $d.Sections) {{
                                    $sec.PageSetup.PaperSize = {paper_code}
                                }}
                                $w.Options.PrintBackground = $false
                                $d.PrintOut()
                                $d.Close($false)
                            }} catch {{ }}
                            """
                            subprocess.run(["powershell", "-ExecutionPolicy", "Bypass", "-Command", ps_script], creationflags=0x08000000)

                for f in temp_files_to_delete:
                    try: os.remove(f)
                    except: pass

                self._log(f"Đã xuất báo cáo tại {out_dir}", "OK")
                self.root.after(0, lambda: messagebox.showinfo("Thành công", f"Đã xử lý và kết xuất báo cáo thành công tại:\n{out_dir}"))

            except Exception as e:
                self._log(f"Lỗi hệ thống: {e}", "ERROR")
                self.root.after(0, lambda msg=traceback.format_exc(), e_str=str(e): messagebox.showerror("Lỗi Cấu Trúc", f"Có lỗi hệ thống xảy ra:\n{e_str}\n\n{msg}"))
            finally:
                self.root.after(0, lambda: self.btn_run.config(state=tk.NORMAL))

    def launch_application():
        if platform.system() == "Windows":
            multiprocessing.freeze_support()

        splash = tk.Tk()
        splash.title("Đang khởi động")
        splash.geometry("460x220")
        splash.resizable(False, False)
        apply_app_icon(splash, "ts_pile")

        container = tk.Frame(splash, bg=APP_BG, padx=24, pady=22)
        container.pack(fill=tk.BOTH, expand=True)
        tk.Label(container, text=APP_NAME, font=("Segoe UI", 15, "bold"), fg="#0F2742", bg=APP_BG).pack(anchor=tk.W)
        status = tk.StringVar(value="Đang kiểm tra bản quyền và chuẩn bị môi trường...")
        tk.Label(container, textvariable=status, font=("Segoe UI", 10), fg="#1F6FEB", bg=APP_BG, wraplength=400, justify=tk.LEFT).pack(anchor=tk.W, pady=(0, 16))
        pb = ttk.Progressbar(container, mode="indeterminate")
        pb.pack(fill=tk.X)
        pb.start(12)

        def worker():
            ok, days_left_info = check_server_trial()
            def done():
                pb.stop()
                splash.destroy()
                if not ok:
                    root_hidden = tk.Tk(); root_hidden.withdraw()
                    if N2D_LICENSE_INFO.get("reason") in ("network", "offline_trial_expired"):
                        messagebox.showerror(
                            "Không kiểm tra được bản quyền",
                            N2D_LICENSE_INFO.get("message") or
                            "Không kết nối được máy chủ bản quyền. Vui lòng kết nối mạng rồi mở lại phần mềm.")
                    else:
                        messagebox.showerror("Hết hạn", "Hết thời gian sử dụng hoặc chưa kích hoạt bản quyền.")
                    root_hidden.destroy()
                    sys.exit()
                root = tk.Tk()
                app = PileDesignToolApp(root, days_left_info)
                def _safe_close():
                    try:
                        root.destroy()
                    except Exception:
                        sys.exit(0)
                root.protocol("WM_DELETE_WINDOW", _safe_close)
                root.mainloop()
            splash.after(0, done)

        threading.Thread(target=worker, daemon=True).start()
        splash.mainloop()

    if __name__ == "__main__":
        launch_application()

except Exception:
    # QA fix L3: nếu lỗi khởi động do chính tkinter thì không dựng được hộp thoại;
    # ghi ra stderr thay vì crash thứ cấp.
    import traceback as _tb
    _details = _tb.format_exc()
    try:
        root_hidden = tk.Tk(); root_hidden.withdraw()
        from tkinter import messagebox
        messagebox.showerror("Lỗi", f"Lỗi khởi động:\n{_details}")
        root_hidden.destroy()
    except Exception:
        sys.stderr.write("Loi khoi dong:\n" + _details + "\n")
    sys.exit(1)